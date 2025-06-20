from django.shortcuts import render, redirect , HttpResponse
from django.contrib.auth import authenticate, login,logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import JsonResponse
import pymysql
import json
from .models import Complaint
from django.http import JsonResponse
from django.db import connection
from django.contrib import messages
from datetime import datetime, timedelta
from django.views.decorators.cache import never_cache   
from django.views.decorators.cache import cache_control
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
import logging
from django.urls import reverse
from django.http import HttpResponseNotFound
from openpyxl.styles import PatternFill
from datetime import date
from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter
import io
from openpyxl.styles import Font, Alignment,PatternFill
import pytz
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from django.core.mail import send_mail
from django.conf import settings
from django.utils.html import strip_tags
from django.views.decorators.csrf import csrf_protect
import base64
import os

# Create your views here.
def connection():
    db = pymysql.connect(
        host='localhost',
        user='root',
        password = 'root',
        database='complaints_db',
    )
    
    cur = db.cursor()
    return cur,db
cur,db = connection()

@never_cache  # Prevent caching of this view
def login(request):
    # Check if user is already logged in
    if request.session.get('loggedIn', False):
        user_role = request.session.get('user_role')
        if user_role == 'superadmin':
            return redirect('superadmindashboard')
        elif user_role == 'admin':
            return redirect('resolverdashboard')
        elif user_role == 'user':
            return redirect('userdashboard')

    if request.method == 'POST':
        # Parse the incoming JSON request body
        data = json.loads(request.body)
        email = data.get('email')
        password = data.get('password')

        # Initialize database connection
        cur, db = connection()  

        try:
            # Query to fetch user info from the database, excluding soft-deleted users
            query = """
                SELECT id, name, role, location, department, password 
                FROM user 
                WHERE email = %s AND is_deleted = 0
            """
            cur.execute(query, (email,))
            user = cur.fetchone()

            if user is not None:
                user_id, name, role, location, department, user_password = user

                # Compare the entered password with the password stored in the database (plain text comparison)
                if password != user_password:
                    return JsonResponse({'success': False, 'message': 'Invalid credentials'}, status=401)

                # Set session variables
                request.session['user_id'] = user_id
                request.session['user_name'] = name
                request.session['user_role'] = role
                request.session['user_location'] = location
                request.session['user_department'] = department
                request.session['loggedIn'] = True

                # Return success response
                return JsonResponse({
                    'success': True,
                    'message': f'Welcome, {name}!',
                    'role': role 
                }, status=200)
            else:
                return JsonResponse({'success': False, 'message': 'Invalid credentials'}, status=401)

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

        finally:
            cur.close()
            db.close()

    # If the request is GET (initial page load)
    response = render(request, 'login.html')
    response['Cache-Control'] = 'no-store'  # Prevent caching
    return response

@never_cache
def logout_view(request):
    logout(request)  # Log out the user
    request.session.flush()  # Clear the session data
    return redirect('login')  # Redirect to the login page

@never_cache
def superadmindashboard(request):
    # Check if the user is logged in
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    
    # Get the logged-in user's name and role from the session
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')

    # Establish database connection
    cur, db = connection()
    
    # Fetch Graphical additional data
    cur.execute("SELECT location, COUNT(*) FROM complaint GROUP BY location")
    location_data = cur.fetchall()

    cur.execute("SELECT status, COUNT(*) FROM complaint GROUP BY status")
    status_data = cur.fetchall()

    cur.execute("SELECT location, status, COUNT(*) FROM complaint GROUP BY location, status")
    state_complaints_data = cur.fetchall()

    cur.execute("SELECT complaint_type, COUNT(*) FROM complaint GROUP BY complaint_type")
    category_data = cur.fetchall()

    cur.execute("SELECT resolved_by_id, COUNT(*) FROM complaint WHERE status = 'resolved' GROUP BY resolved_by_id")
    resolver_performance_data = cur.fetchall()

    cur.execute("SELECT department, COUNT(*) FROM complaint GROUP BY department")
    department_data = cur.fetchall()

    cur.execute("SELECT location, COUNT(*) FROM complaint WHERE status = 'resolved' GROUP BY location")
    locations_data = cur.fetchall()

    # Prepare data for graphs
    
    locations = [item[0] for item in location_data]
    location_counts = [item[1] for item in location_data]

    statuses = [item[0] for item in status_data]
    status_counts = [item[1] for item in status_data]

    states = [item[0] for item in state_complaints_data]
    states_status = [item[1] for item in state_complaints_data]
    state_counts = [item[2] for item in state_complaints_data]

    categories = [item[0] for item in category_data]
    category_counts = [item[1] for item in category_data]

    resolvers = [item[0] for item in resolver_performance_data]
    resolver_counts = [item[1] for item in resolver_performance_data]

    departments = [row[0] for row in department_data]
    department_counts = [row[1] for row in department_data]

    locations2 = [item[0] for item in locations_data]
    location_counts2 = [item[1] for item in locations_data]
    
    # Initialize an empty error message
    error_message = None

    # Get date range from request parameters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    date_filter = ""
    params = []

    # Handle the from_date and to_date filter
    if from_date and to_date:
        try:
            # Parse from_date and to_date
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d")
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(microseconds=1)

            # Create date filter with inclusive dates and add parameters
            date_filter = "AND c.issue_raise_date BETWEEN %s AND %s"
            params.extend([from_date_obj, to_date_obj])

        except ValueError:
            error_message = "Invalid date format. Please use YYYY-MM-DD."

    # Handle the last 7 days filter if no from_date and to_date are provided
    elif not from_date and not to_date:
        seven_days_ago = datetime.now() - timedelta(days=7)
        date_filter = "AND c.issue_raise_date >= %s"
        params.append(seven_days_ago)

    try:
        # Update `complain_status` for overdue complaints
        cur.execute(""" 
            UPDATE complaint 
            SET complain_status = 'overdue' 
            WHERE updated_at > due_date AND complain_status IS NOT NULL;
        """)
        db.commit()

        # Fetch statistics
        cur.execute("SELECT COUNT(*) FROM complaint")
        total_complaints = cur.fetchone()[0]

        cur.execute("SELECT COUNT(*) FROM complaint WHERE status = 'work in progress' ")
        total_in_progress = cur.fetchone()[0]

        cur.execute("SELECT COUNT(*) FROM complaint WHERE status = 'resolved' AND complain_status = 'done'")
        total_resolved = cur.fetchone()[0]
        
        # Count complaints where status is 'resolved' and complain_status is 'overdue'
        cur.execute("SELECT COUNT(*) FROM complaint WHERE status = 'resolved' AND complain_status = 'overdue'")
        resolved_overdue = cur.fetchone()[0]

        cur.execute("SELECT COUNT(*) FROM complaint WHERE status = 'pending'")
        total_overdue = cur.fetchone()[0]
               
        statistics = {
            'total_complaints': total_complaints,
            'total_in_progress': total_in_progress,
            'total_resolved': total_resolved,
            'total_overdue': total_overdue,
            'resolved_overdue': resolved_overdue,
        } 

        # Fetch resolution reports with date filter
        query = f"""
            SELECT 
                c.id AS complaint_id,
                u.name AS complainant_name, 
                u.department, 
                u.location, 
                COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
                c.issue_raise_date AS date,
                c.due_date,
                c.status,
                c.resolution,
                c.description,
                c.complaint_type,
                CASE 
                    WHEN c.status = 'resolved' THEN c.updated_at
                    ELSE NULL
                END AS resolved_at,
                c.updated_due_date,  -- Add updated_due_date column
                c.reason            -- Add reason column
            FROM 
                complaint AS c
            JOIN 
                user AS u ON c.user_id = u.id
            LEFT JOIN 
                user AS u2 ON c.resolved_by_id = u2.id AND u2.role = 'admin' 
            WHERE 
                (c.status IN ('resolved', 'work in progress', 'pending') OR c.complain_status = 'overdue') 
                {date_filter}
            ORDER BY 
                c.issue_raise_date DESC;
        """

        cur.execute(query, params)
        resolution_reports = cur.fetchall()

        # Process report data and calculate TAT
        report_data = []
        for report in resolution_reports:
            complaint_id = report[0]
            complainant_name = report[1]
            department = report[2]
            location = report[3].capitalize()
            resolved_by = report[4]
            date = report[5]
            due_date = report[6]  # Get the due_date from the query result
            status = report[7].capitalize()
            resolution = report[8]
            description = report[9]
            complaint_type = report[10]
            resolved_at = report[11]
            updated_due_date = report[12]  # Updated due date from the query result
            reason = report[13]  # Reason from the query result
                    
            # Calculate TAT
            issue_raise_date = date
            if resolved_at:
                tat_duration = resolved_at - issue_raise_date
            else:
                tat_duration = datetime.now() - issue_raise_date

            total_seconds = tat_duration.total_seconds()
            days = int(total_seconds // 86400)
            hours = int((total_seconds % 86400) // 3600)
            minutes = int((total_seconds % 3600) // 60)
            seconds = int(total_seconds % 60)
            # Format TAT based on the conditions
            if days > 0:
        # If more than a day, show days, hours, minutes, and seconds
                tat_formatted = f"{days}d {hours}h {minutes}m {seconds}s"
            elif hours > 0:
                # If less than a day but more than an hour, show hours, minutes, and seconds
                tat_formatted = f"{hours}h {minutes}m {seconds}s"
            else:
                # If less than an hour, show minutes and seconds
                tat_formatted = f"{minutes}m {seconds}s"
                
            report_data.append({
                'id':complaint_id,
                'name': complainant_name, 
                'department': department,
                'location': location,
                'resolved_by': resolved_by,
                'date': date.strftime("%Y-%m-%d %H:%M:%S"),
                'due_date': due_date.strftime("%Y-%m-%d %H:%M:%S"),
                'status': status,
                'resolution': resolution,
                'description': description,
                'complaint_type': complaint_type,  # Added complaint_type to the report
                'updated_at': resolved_at.strftime("%Y-%m-%d %H:%M:%S") if resolved_at else 'None',
                'tat': tat_formatted if resolved_at else 'None',  # Add TAT to the report data
                'updated_due_date': updated_due_date.strftime("%Y-%m-%d %H:%M:%S") if updated_due_date else 'No revised due date assigned yet.',  # Added updated_due_date
                'reason': reason if reason else 'No reason specified.'  # Added reason
            })

        # 1. Fetch all admins with their id, location, and department
        cur.execute("SELECT id, location, department FROM user WHERE role = 'admin'")
        admins = cur.fetchall()

        # 2. Iterate over each admin (resolver) and check for complaints assigned to them in their respective location and department
        for admin in admins:
            admin_id, admin_location, admin_department = admin

            # 3. Select complaints that are resolved, belong to the admin's location and department, 
            # and are assigned to the admin (resolver)
            cur.execute(""" 
                SELECT id, location, department, assigned_to
                FROM complaint
                WHERE status = 'resolved' 
                AND resolution IS NOT NULL 
                AND resolution <> '' 
                AND location = %s
                AND department = %s
                AND assigned_to = %s  -- Match complaints assigned to the current admin
            """, (admin_location, admin_department, admin_id))
            
            complaints_to_update = cur.fetchall()

            if complaints_to_update:
                # 4. Update the resolved_by_id for each complaint with the current resolver's admin_id
                for complaint in complaints_to_update:
                    complaint_id = complaint[0]


                    cur.execute(""" 
                        UPDATE complaint
                        SET resolved_by_id = %s 
                        WHERE id = %s
                    """, (admin_id, complaint_id))


                db.commit()
                print(f"Updated resolved_by_id for complaints in {admin_location}, department {admin_department} by admin {admin_id}")
            else:
                print(f"No resolved complaints assigned to admin {admin_id} in {admin_location}, department {admin_department}")

    except Exception as e:
        print(f"Error: {e}")
        db.rollback()
        error_message = str(e)

    finally:
        cur.close()
        db.close()

    # Render the dashboard with statistics and filtered report data
    return render(request, 'superadmindashboard.html', {
        'stats': statistics,
        'reports': report_data,
        'user_name': user_name,
        'user_role': user_role,
        'error_message': error_message,
        'from_date': from_date,
        'to_date': to_date,
        'departments': departments,
        'department_counts': department_counts,
        'locations': locations,
        'location_counts': location_counts,
        'locations2': locations2,
        'location_counts2': location_counts2,
        'statuses': statuses,
        'status_counts': status_counts,
        'states': states,
        'states_status': states_status,
        'state_counts': state_counts,
        'categories': categories,
        'category_counts': category_counts,
        'resolvers': resolvers,
        'resolver_counts': resolver_counts,
    })

# def superadmin(request):
#     return render(request,'superadmin.html')

@never_cache
def superadmin(request):
    user_name = request.session.get('user_name')
    cur, db = connection()
    try:
        if request.method == "POST" and "add_user" in request.POST:
            name = request.POST.get('name')
            location = request.POST.get('location')
            department = request.POST.get('department')
            issue_raise_date = request.POST.get('date')

            # Insert new user into the database
            cur.execute("""
                INSERT INTO user (name, location, department, issue_raise_date) 
                VALUES (%s, %s, %s, %s)
            """, (name, location, department, issue_raise_date))
            db.commit()

            return redirect('superadmin')

        # Fetch user records
        cur.execute("SELECT * FROM user WHERE is_deleted = 0")
        users = cur.fetchall()

        # Fetch complaints with additional details
        cur.execute("""
            SELECT 
                c.id AS complaint_id,  -- Include complaint_id to link to the resolver view
                u.name AS user_name, 
                c.department AS user_department, 
                c.complaint_type, 
                c.description, 
                c.location, 
                c.issue_raise_date, 
                c.status,
                c.resolution,
                c.due_date
            FROM 
                complaint c 
            JOIN 
                user u ON c.user_id = u.id
            ORDER BY 
                c.issue_raise_date DESC;  -- Order by complaint_id in ascending order
        """)
        complaints = cur.fetchall()  # Fetch all results

        # Prepare complaints data for the template
        complaints_data = [
            {
                'complaint_id': row[0],  # Include complaint_id for linking to resolver view
                'user_name': row[1],
                'user_department': row[2],
                'complaint_type': row[3],
                'description': row[4],
                'location': row[5].capitalize(),
                'issue_raise_date': row[6].strftime("%Y-%m-%d %H:%M:%S"),  # Format date for display
                'status': row[7].capitalize(),
                'action': 'View', # Placeholder for the action column
                'resolution':row[8],
                'due_date':row[9].strftime("%Y-%m-%d %H:%M:%S"),
            }
            for row in complaints
        ]

        return render(request, 'superadmin.html', {'users': users, 'complaints': complaints_data,'user_name':user_name})

    finally:
        cur.close()
        db.close()

def delete_user(request, user_id):
    cur, db = connection()
    
    try:
        # Check if the user exists
        cur.execute("SELECT * FROM user WHERE id = %s", (user_id,))
        user = cur.fetchone()
        
        if not user:
            messages.error(request, f"User with ID {user_id} does not exist.")
            return redirect('superadmin')

        # Move user to `deleted_users` table
        cur.execute("""
            INSERT INTO deleted_users (id, name, department, location, email, emp_code, emp_level, created_at)
            SELECT id, name, department, location, email, emp_code, emp_level, created_at
            FROM user
            WHERE id = %s
        """, (user_id,))

        # Update the `is_deleted` flag in the `user` table
        cur.execute("UPDATE user SET is_deleted = 1 WHERE id = %s", (user_id,))

        db.commit()
        messages.success(request, f"User with ID {user_id} soft-deleted successfully!")
    except Exception as e:
        db.rollback()
        messages.error(request, f"Error deleting user: {str(e)}")
    finally:
        cur.close()
        db.close()

    return redirect('superadmin')

@never_cache
def resolver_dashboard(request):
    # Check if the user is logged in
    if 'user_id' not in request.session:
        return redirect('login')  # Redirect to the login page if not logged in

    # Get the logged-in user's ID, name, role, location, and department from the session
    user_id = request.session.get('user_id')
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')
    
    # Establish the connection and get the cursor
    cur, db = connection()  # Ensure `connection()` is correctly defined in your project

    # Get the current resolver's ID (e.g., from session or user model)
    current_resolver_id = user_id  # Using logged-in user's ID (assumes it's stored in the session)

    # 1. Fetch all complaints that are assigned to the current resolver
    cur.execute(""" 
        SELECT id, location, department, assigned_to
        FROM complaint
        WHERE status = 'resolved' 
        AND resolution IS NOT NULL 
        AND resolution <> '' 
        AND assigned_to = %s  -- Match complaints assigned to the current resolver
    """, (current_resolver_id,))

    complaints_to_update = cur.fetchall()

    if complaints_to_update:
        # 2. Update the resolved_by_id for each complaint with the current resolver's ID
        for complaint in complaints_to_update:
            complaint_id = complaint[0]

            cur.execute(""" 
                UPDATE complaint
                SET resolved_by_id = %s 
                WHERE id = %s
            """, (current_resolver_id, complaint_id))

        db.commit()
        print(f"Updated resolved_by_id for complaints resolved by Resolver with ID : {current_resolver_id}")
    else:
        print(f"No resolved complaints assigned to Resolver with ID :  {current_resolver_id}")
    
    # Fetch Graphical additional data
    cur.execute("SELECT location, COUNT(*) FROM complaint WHERE assigned_to = %s GROUP BY location", [current_resolver_id])
    location_data = cur.fetchall()

    cur.execute("SELECT status, COUNT(*) FROM complaint WHERE assigned_to = %s GROUP BY status", [current_resolver_id])
    status_data = cur.fetchall()

    cur.execute("SELECT location, status, COUNT(*) FROM complaint WHERE assigned_to = %s GROUP BY location, status", [current_resolver_id])
    state_complaints_data = cur.fetchall()

    cur.execute("SELECT complaint_type, COUNT(*) FROM complaint WHERE assigned_to = %s GROUP BY complaint_type", [current_resolver_id])
    category_data = cur.fetchall()

    # Fetch performance data for the current resolver
    cur.execute("SELECT resolved_by_id, COUNT(*) FROM complaint WHERE assigned_to = %s AND status = 'resolved' GROUP BY resolved_by_id", [current_resolver_id])
    resolver_performance_data = cur.fetchall()

    cur.execute("SELECT department, COUNT(*) FROM complaint WHERE assigned_to = %s GROUP BY department", [current_resolver_id])
    department_data = cur.fetchall()

    cur.execute("SELECT location, COUNT(*) FROM complaint WHERE assigned_to = %s AND status = 'resolved' GROUP BY location", [current_resolver_id])
    locations_data = cur.fetchall()

    # Execute the SQL query to get both resolved and total counts
   
    cur.execute("""
        SELECT 
            COUNT(CASE WHEN status = 'resolved' THEN 1 END) AS resolved_count,
            COUNT(*) AS total_count
        FROM complaint
        WHERE assigned_to = %s
    """, [current_resolver_id])

    # Fetch the data
    result = cur.fetchall()

    if result:
        resolved_count = result[0][0]  # resolved_count
        total_count = result[0][1]     # total_count
        
        print(f"Resolved count for resolver is : {resolved_count}")
        print(f"Total complaints for resolver is : {total_count}")
        
        # Calculate the percentage score if total_count > 0
        if total_count > 0:
            resolver_score_percentage = (resolved_count / total_count) * 100
        else:
            resolver_score_percentage = 0  # Handle no complaints case
    else:
        resolver_score_percentage = 0  # Handle no complaints for the resolver

    # Calculate total complaints and performance score for the current resolver
    resolver_score = 0
    total_complaints = 0
    for resolver, count in resolver_performance_data:
        if resolver == current_resolver_id:
            resolver_score = count
        total_complaints += count

    # Prepare data for graphs
    locations = [item[0] for item in location_data]
    location_counts = [item[1] for item in location_data]

    statuses = [item[0] for item in status_data]
    status_counts = [item[1] for item in status_data]

    states = [item[0] for item in state_complaints_data]
    states_status = [item[1] for item in state_complaints_data]
    state_counts = [item[2] for item in state_complaints_data]

    categories = [item[0] for item in category_data]
    category_counts = [item[1] for item in category_data]

    resolvers = [item[0] for item in resolver_performance_data]
    resolver_counts = [item[1] for item in resolver_performance_data]

    departments = [row[0] for row in department_data]
    department_counts = [row[1] for row in department_data]

    locations2 = [item[0] for item in locations_data]
    location_counts2 = [item[1] for item in locations_data]

    # Fetch resolver's location and department from the database
    cur.execute("SELECT location, department FROM user WHERE id = %s", [user_id])
    resolver_data = cur.fetchone()
    if resolver_data:
        resolver_location = resolver_data[0]  # Extract the location value
        resolver_department = resolver_data[1]  # Extract the department value
    else:
        resolver_location = None  # Default to None if no location found
        resolver_department = None  # Default to None if no department found

    # Get 'fromdate' and 'todate' filters from the request
    fromdate_str = request.GET.get('fromdate', '')
    todate_str = request.GET.get('todate', '')

    # Calculate default date range (last 7 days)
    today = datetime.now()
    default_fromdate = (today - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
    default_todate = today.replace(hour=23, minute=59, second=59, microsecond=999999)

    # Convert 'fromdate' and 'todate' to datetime objects
    try:
        if fromdate_str:
            fromdate = datetime.strptime(fromdate_str, "%Y-%m-%d")
            full_fromdate = fromdate.replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            full_fromdate = default_fromdate

        if todate_str:
            todate = datetime.strptime(todate_str, "%Y-%m-%d")
            full_todate = todate.replace(hour=23, minute=59, second=59, microsecond=999999)
        else:
            full_todate = default_todate
    except ValueError:
        full_fromdate = default_fromdate
        full_todate = default_todate

    # Raw SQL query to get complaints assigned to the admin (resolver) and filter by date and department
    query = """
    SELECT 
        c.id AS complaint_id,  -- Include complaint_id to link to the resolver view
        u.name AS user_name,   -- Assuming 'user' table has a 'name' field
        c.department AS user_department, 
        c.complaint_type, 
        c.description, 
        TRIM(UPPER(c.location)) AS location,  -- Standardize location using TRIM and UPPER
        c.issue_raise_date, 
        c.status,
        c.resolution,
        c.updated_at,
        c.due_date,
        c.updated_due_date,     -- Include the updated_due_date column
        c.reason                -- Add the reason column here
    FROM 
        complaint c 
    JOIN 
        user u ON c.user_id = u.id
    WHERE 
        TRIM(UPPER(c.location)) = %s  -- Ensure consistency for location comparison
        AND c.department = %s
        AND c.assigned_to = %s  -- Filter complaints by the logged-in resolver (admin)
        AND c.issue_raise_date BETWEEN %s AND %s
    ORDER BY 
        c.issue_raise_date DESC;
    """

    # Execute the query with resolver's location, department, admin ID, and date filters
    cur.execute(query, [resolver_location.upper().strip(), resolver_department, user_id, full_fromdate, full_todate])
    dashboard_data = cur.fetchall()
    
    print(f'This is the dashboard data : {dashboard_data}')

    # Update overdue complaints
    cur.execute(""" 
        UPDATE complaint 
        SET complain_status = 'overdue' 
        WHERE updated_at > due_date AND complain_status IS NOT NULL;
    """)
    db.commit()

    # Fetch statistics with standardized location values
    cur.execute(""" 
        SELECT COUNT(*) 
        FROM complaint 
        WHERE TRIM(UPPER(location)) = %s 
        AND department = %s 
        AND assigned_to = %s
    """, [resolver_location.upper().strip(), resolver_department, user_id])
    total_complaints = cur.fetchone()[0]

    cur.execute(""" 
        SELECT COUNT(*) 
        FROM complaint 
        WHERE status = 'work in progress' 
        AND TRIM(UPPER(location)) = %s 
        AND department = %s 
        AND assigned_to = %s
    """, [resolver_location.upper().strip(), resolver_department, user_id])
    total_in_progress = cur.fetchone()[0]

    cur.execute(""" 
        SELECT COUNT(*) 
        FROM complaint 
        WHERE status = 'resolved' 
        AND complain_status = 'done' 
        AND TRIM(UPPER(location)) = %s 
        AND department = %s 
        AND assigned_to = %s
    """, [resolver_location.upper().strip(), resolver_department, user_id])
    total_resolved = cur.fetchone()[0]

    cur.execute(""" 
        SELECT COUNT(*) 
        FROM complaint 
        WHERE status = 'resolved' 
        AND complain_status = 'overdue' 
        AND TRIM(UPPER(location)) = %s 
        AND department = %s 
        AND assigned_to = %s
    """, [resolver_location.upper().strip(), resolver_department, user_id])
    resolved_overdue = cur.fetchone()[0]

    cur.execute(""" 
        SELECT COUNT(*) 
        FROM complaint 
        WHERE status = 'pending' 
        AND TRIM(UPPER(location)) = %s 
        AND department = %s 
        AND assigned_to = %s
    """, [resolver_location.upper().strip(), resolver_department, user_id])
    total_overdue = cur.fetchone()[0]

    # Close the cursor and database connection
    cur.close()
    db.close()
    
    # Prepare data for the template
    context = {
        'dashboard_data': [
            {
                'complaint_id': row[0],  # Include complaint_id for linking to resolver view
                'user_name': row[1],
                'department': row[2],
                'complaint_type': row[3],
                'description': row[4],
                'location': row[5].capitalize() if row[5] else 'Not Applicable',
                'issue_raise_date': row[6].strftime("%Y-%m-%d %H:%M:%S") if row[6] else 'Not Applicable',  # Format date for display
                'status': row[7].capitalize(),
                'resolution': row[8] if row[8] else 'No Resolution Yet',
                'action': 'View',
                'updated_at': row[9].strftime("%Y-%m-%d %H:%M:%S") if row[7] == 'resolved' else 'Not Resolved',  # Format date for display
                'due_date': row[10].strftime("%Y-%m-%d %H:%M:%S") if row[10] else 'Not Applicable',
                'updated_due_date': row[11].strftime("%Y-%m-%d %H:%M:%S") if row[11] else 'No revised due date yet',  # Include updated_due_date
                'reason': row[12] if row[12] else 'No reason yet'  # Include reason field, default to 'N/A' if None
            }
            for row in dashboard_data
        ],
        'user_name': user_name,
        'user_role': user_role,
        'departments': departments,
        'department_counts': department_counts,
        'locations': locations,
        'location_counts': location_counts,
        'locations2': locations2,
        'location_counts2': location_counts2,
        'statuses': statuses,
        'status_counts': status_counts,
        'states': states,
        'states_status': states_status,
        'state_counts': state_counts,
        'categories': categories,
        'category_counts': category_counts,
        'resolvers': resolvers,
        'resolver_counts': resolver_counts,
        'resolver_score_percentage': resolver_score_percentage,
        'resolver_score': resolver_score,
        'total_complaints': total_complaints, # Add total complaints to the context
        'resolved_count':resolved_count,
        'total_count':total_count,           
        'statistics': {
            'total_complaints': total_complaints,
            'total_in_progress': total_in_progress,
            'total_resolved': total_resolved,
            'total_overdue': total_overdue,
            'resolved_overdue': resolved_overdue,
        },
        'fromdate': full_fromdate.strftime("%Y-%m-%d"),
        'todate': full_todate.strftime("%Y-%m-%d"),

    }
    return render(request, 'resolverdashboard.html', context)

# def user_dashboard(request):
#     return render(request,'userdashboard.html')

def format_resolution_time(delta):
    days = delta.days
    hours, remainder = divmod(delta.seconds, 3600)
    minutes = remainder // 60

    if days > 0:
        return f"{days}d {hours}h {minutes}m"
    elif hours > 0:
        return f"{hours}h {minutes}m"
    else:
        return f"{minutes}m"

@never_cache
def user_dashboard(request):
    # Check if the user is logged in
    if 'user_id' not in request.session:
        return redirect('login')  # Redirect to the login page if not logged in
    
    # Get the logged-in user's ID, name, and role from the session
    user_id = request.session.get('user_id')
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')
    
    # Print session keys for debugging
    print("Current session keys:", request.session.keys())
    
    # Connect to the database
    cur, db = connection()
        
    # Fetch complaint counts by status for the logged-in user
    cur.execute("""
        SELECT status, COUNT(*) 
        FROM complaint 
        WHERE user_id = %s
        GROUP BY status
    """, (user_id,))

    status_data = cur.fetchall()

# Fetch complaint counts by department for the logged-in user, including department from complaint_type_master
    cur.execute("""
        SELECT ct.department, COUNT(*) 
        FROM complaint c
        JOIN complaint_type_master ct ON c.complaint_type = ct.complaint_type
        WHERE c.user_id = %s
        GROUP BY ct.department
    """, (user_id,))

    department_data = cur.fetchall()

    # Fetch complaint counts by category with priority levels for the logged-in user
    cur.execute("""
        SELECT 
            complaint_type AS category, 
            COUNT(*) AS total_complaints,
            COUNT(CASE WHEN status = 'Pending' THEN 1 END) AS pending_complaints,
            COUNT(CASE WHEN status = 'Work in Progress' THEN 1 END) AS work_in_progress_complaints,
            CASE
                WHEN COUNT(*) > 100 THEN 'High Priority'
                WHEN COUNT(*) BETWEEN 50 AND 100 THEN 'Medium Priority'
                ELSE 'Low Priority'
            END AS priority
        FROM complaint
        WHERE user_id = %s
        GROUP BY category
        ORDER BY total_complaints DESC
    """, (user_id,))

    category_data = cur.fetchall()

   # Fetch data from the database
    cur.execute("""
        SELECT complaint_type, 
            ROUND(AVG(TIMESTAMPDIFF(SECOND, issue_raise_date, updated_at) / 3600), 2) AS avg_resolution_hours
        FROM complaint
        WHERE status = 'resolved' AND user_id = %s
        GROUP BY complaint_type;
    """, [user_id])
    avg_resolution_time = cur.fetchall()

    # Process data for graph rendering (Complaint Categories and Priorities)
    categories1 = [row[0] for row in category_data]  # Complaint categories (e.g., 'IT', 'Support')
    category_counts1 = [row[1] for row in category_data]  # Total complaints per category
    priorities1 = [row[4] for row in category_data]  # Priority levels
    print(f"Priorities: {priorities1}")

    # Extract complaint types and average resolution times
    categories2 = [row[0] for row in avg_resolution_time]  # Complaint types (e.g., Access Denied, Password Reset Issue)
    resolutiontime = [float(row[1]) for row in avg_resolution_time]  # Average resolution time (in hours)

    # Print out the processed data for debugging purposes
    print(f"Category : {categories2}")
    print(f"Average Resolution Time: {resolutiontime} hours")

    # Process data for departments (Department counts)
    departments = [row[0] for row in department_data]  # Department names
    department_counts = [row[1] for row in department_data]  # Total complaints per department

    # Process data for statuses (Complaint statuses)
    statuses = [item[0] for item in status_data]  # Complaint statuses
    status_counts = [item[1] for item in status_data]  # Total complaints per status

    # Fetch statistics based on user_id (Total complaints by user)
    cur.execute("SELECT COUNT(*) FROM complaint WHERE user_id = %s", [user_id])
    total_complaints = cur.fetchone()[0]

    # Print out the total complaints for the user (for debugging)
    print(f"Total Complaints for user with ID :{user_id} is {total_complaints}")

    cur.execute("""
    SELECT COUNT(*) 
    FROM complaint 
    WHERE user_id = %s 
    AND (status = 'work in progress' OR status = 'overdue')
    AND (complain_status = 'overdue' OR complain_status = 'done')
""", [user_id])

    total_in_progress = cur.fetchone()[0]

    cur.execute("""
    SELECT COUNT(*) 
    FROM complaint 
    WHERE user_id = %s 
    AND status = 'resolved' 
    AND complain_status = 'done'
""", [user_id])
    total_resolved = cur.fetchone()[0]

    # Count complaints where status is 'resolved' and complain_status is 'overdue'
    cur.execute("""
        SELECT COUNT(*) 
        FROM complaint 
        WHERE user_id = %s 
        AND status = 'resolved' 
        AND complain_status = 'overdue'
    """, [user_id])
    resolved_overdue = cur.fetchone()[0]

    cur.execute("""
    SELECT COUNT(*) 
    FROM complaint 
    WHERE user_id = %s 
    AND complain_status = 'done' 
    AND status = 'pending'
""", [user_id])
    total_overdue = cur.fetchone()[0]
    
    # Calculate the date 7 days ago from today
    seven_days_ago = datetime.now() - timedelta(days=7)
    seven_days_ago_str = seven_days_ago.strftime("%Y-%m-%d")

    # Fetch complaints from the last 7 days
    cur.execute("""
        SELECT * FROM complaint
        WHERE user_id = %s AND issue_raise_date >= %s
        ORDER BY issue_raise_date DESC
    """, [user_id, seven_days_ago_str])
    complaints = cur.fetchall()

    # Fetch logged-in user's details
    cur.execute("SELECT * FROM user WHERE id = %s", [user_id])
    user = cur.fetchone()
    
    print(f'The user records are : {user}')

    user_records = []
    
    # Get 'fromdate' and 'todate' filters from the request
    fromdate_str = request.GET.get('fromdate', '')
    todate_str = request.GET.get('todate', '')
    
    # Convert 'fromdate' and 'todate' to datetime objects
    try:
        if fromdate_str:
            fromdate = datetime.strptime(fromdate_str, "%Y-%m-%d")
            # Set the time to the beginning of the day (00:00:00)
            fromdate = fromdate.replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            fromdate = None
        
        if todate_str:
            todate = datetime.strptime(todate_str, "%Y-%m-%d")
            # Set the time to the end of the day (23:59:59)
            todate = todate.replace(hour=23, minute=59, second=59, microsecond=999999)
        else:
            todate = None
    except ValueError:
        messages.error(request, 'Invalid date format. Please use YYYY-MM-DD.')
        fromdate = todate = None
    
    # Filter complaints based on 'fromdate' and 'todate'
    if user:
        for complaint in complaints:
            issue_raise_date = complaint[7]
            
            # Ensure issue_raise_date is a datetime object
            if isinstance(issue_raise_date, str):
                issue_raise_date = datetime.strptime(issue_raise_date, "%Y-%m-%d %H:%M:%S")
            
            # Check if the issue_raise_date is within the selected date range
            if fromdate and issue_raise_date < fromdate:
                continue  # Skip this complaint if it's before 'fromdate'
            if todate and issue_raise_date > todate:
                continue  # Skip this complaint if it's after 'todate'
            
            updated_at = complaint[11] if complaint[13].lower() == "resolved" else None
            
            # Ensure `updated_at` is a datetime object if resolved
            if updated_at and isinstance(updated_at, str):
                updated_at = datetime.strptime(updated_at, "%Y-%m-%d %H:%M:%S")
            
            # Calculate resolution time if resolved
            resolution_time = (
                format_resolution_time(updated_at - issue_raise_date) if updated_at else "Not Resolved Yet"
            )
            
            # Add the updated_due_date and reason fields from complaint[17] and complaint[18]
            updated_due_date = complaint[16] if complaint[16] else "Not Updated"
            reason = complaint[17] if complaint[17] else "No Reason Provided"

            user_records.append({
                'id': user_id,
                'name': user[1],  
                'department': complaint[3], 
                'location': complaint[4].capitalize(), 
                'complaint_type': complaint[5], 
                'status': complaint[13].capitalize(), 
                'resolution' : complaint[14],
                'issue_raise_date': issue_raise_date.strftime("%Y-%m-%d %H:%M:%S"),
                'due_date': complaint[9].strftime("%Y-%m-%d %H:%M:%S"),
                'description' : complaint[6],
                'updated_at': updated_at.strftime("%Y-%m-%d %H:%M:%S") if updated_at else "Not Resolved Yet",
                'resolution_time': resolution_time,
                'updated_due_date': updated_due_date.strftime("%Y-%m-%d %H:%M:%S") if isinstance(updated_due_date, datetime) else updated_due_date,  # Ensure proper formatting
                'reason': reason,
                 
            })

    # Close the database connection
    db.close() 

    # Define the context for the template
    context = {
        'total_complaints': total_complaints,
        'resolved_count': total_resolved,
        'in_progress_count': total_in_progress,
        'overdue_count': total_overdue,
        'resolved_overdue': resolved_overdue,
        'user_records': user_records,
        'user_name': user_name,
        'user_role': user_role,
        'fromdate': fromdate_str,
        'todate': todate_str,
        'statuses': statuses,
        'status_counts': status_counts,
        'departments': departments,
        'department_counts': department_counts,
        'categories1': categories1,
        'category_counts1': category_counts1,
        'priorities1': priorities1,
        'avg_resolution_time' :avg_resolution_time,
        'categories2': categories2,
        'resolutiontime': resolutiontime,

    }

    return render(request, 'userdashboard.html', context)

# def user(request):
#     return render(request,'user.html')

def user(request):
    # Get the logged-in user's ID, name, and role from the session
    user_id = request.session.get('user_id')
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')
    
    # If user_id is not found in session, redirect to login page or show an error
    if not user_id:
        return redirect('login')  # Adjust to your actual login URL

    # Establish the connection and get the cursor
    cur, db = connection()

    # Join user and complaint tables to get the desired data, including user_id, filtered by logged-in user
    query = """
    SELECT 
        c.id, 
        u.name, 
        c.department, 
        c.location, 
        c.issue_raise_date, 
        c.status, 
        c.description, 
        c.due_date, 
        c.complaint_type, 
        c.updated_at,
        c.updated_due_date,   -- Fetch updated_due_date column
        c.reason              -- Fetch reason column
    FROM 
        user u
    INNER JOIN 
        complaint c ON u.id = c.user_id
    WHERE
        u.id = %s  -- Filter complaints for the logged-in user
    ORDER BY 
        c.issue_raise_date DESC
    """
    
    cur.execute(query, (user_id,))
    complaints = cur.fetchall()

    complaint_data = []
    
    for complaint_details in complaints:
        issue_raise_date = complaint_details[4]  # Fetching issue_raise_date
        updated_at = complaint_details[9]  # Fetching updated_at
        status = complaint_details[5].lower()  # Get the status and convert to lowercase for comparison
        updated_due_date = complaint_details[10]  # Fetching updated_due_date
        reason = complaint_details[11]  # Fetching reason

        # Initialize TAT as 'Not resolved yet' by default
        tat_formatted = 'Not resolved yet'

        if status == "resolved" and updated_at:
            # Calculate TAT as updated_at - issue_raise_date only if resolved
            tat_duration = updated_at - issue_raise_date

            # Extract days, hours, minutes, and seconds from TAT duration
            total_seconds = tat_duration.total_seconds()
            days = int(total_seconds // 86400)  # Calculate days
            hours = int((total_seconds % 86400) // 3600)  # Calculate hours
            minutes = int((total_seconds % 3600) // 60)  # Calculate minutes
            seconds = int(total_seconds % 60)  # Calculate seconds

            # Format TAT
            tat_formatted = ""
            if days > 0:
                tat_formatted += f"{days}d "
            if hours > 0 or days > 0:  # Always show hours if there are days
                tat_formatted += f"{hours}h "
            tat_formatted += f"{minutes}m {seconds}s"  # Add minutes and seconds

        # Add the formatted data to complaint_data
        complaint_data.append({
            'complaint_id': complaint_details[0],
            'user_name': complaint_details[1],
            'department': complaint_details[2],
            'location': complaint_details[3].capitalize(),
            'issue_raise_date': complaint_details[4].strftime("%Y-%m-%d %H:%M:%S"),
            'status': complaint_details[5].capitalize(),
            'description': complaint_details[6],
            'due_date': complaint_details[7].strftime("%Y-%m-%d %H:%M:%S"),
            'complaint_type': complaint_details[8],
            'resolved_at': updated_at.strftime("%Y-%m-%d %H:%M:%S") if updated_at else 'Not resolved yet',
            'tat': tat_formatted,  # Add the formatted TAT to the complaint data
            'updated_due_date': updated_due_date.strftime("%Y-%m-%d %H:%M:%S") if updated_due_date else 'No updated due date',  # Add updated_due_date
            'reason': reason if reason else 'No reason provided'  # Add reason
        })

    # Close the cursor and database connection
    cur.close()
    db.close()

    # Return the template with the complaint data
    return render(request, 'user.html', {'complaints': complaint_data, 'user_name': user_name, 'user_role': user_role})

def update_due_date_with_remaining_tat(issue_raise_date, tat_hours, complaint_id, due_date_obj,user_id, country='IN'):
    # Default work start and end times
    DEFAULT_WORK_START = 9
    DEFAULT_WORK_END = 18

    # Fetch resolver's work schedule and public holidays
    cur, db = connection()
    try:
        # Fetch resolver's work schedule based on location and department
        cur.execute("""
            SELECT 
                EXTRACT(HOUR FROM ot.work_start_time) AS work_start_hour,
                EXTRACT(HOUR FROM ot.work_end_time) AS work_end_hour,
                ot.non_working_days 
            FROM user u
            JOIN office_timings ot
                ON u.location = ot.location
                AND u.department = ot.department
            WHERE u.id = %s
        """, [user_id])
        user_schedule = cur.fetchone()
        
        print(f'The Work Start Time of Resolver is : {user_schedule[0]}')
        print(f'The Work End Time of Resolver is : {user_schedule[1]}')
        print(f'The Non-Working Days of Resolver is : {user_schedule[2]}')

        if user_schedule:
            work_start_time = int(user_schedule[0]) if user_schedule[0] else DEFAULT_WORK_START
            work_end_time = int(user_schedule[1]) if user_schedule[1] else DEFAULT_WORK_END
            non_working_days = user_schedule[2] or 'None'

            # Interpret non_working_days based on the enum value
            if non_working_days == 'Saturday':
                non_working_days_set = {'Saturday'}
            elif non_working_days == 'Sunday':
                non_working_days_set = {'Sunday'}
            elif non_working_days == 'Both':
                non_working_days_set = {'Saturday', 'Sunday'}
            else:  # 'None'
                non_working_days_set = set()
        else:
            raise ValueError(f"No schedule found for the user with ID {user_id}")
    finally:
        cur.close()
        db.close()

    # Fetch the updated_at value from the database for the specific complaint
    cur, db = connection()
    try:
        cur.execute("SELECT updated_at FROM complaint WHERE id = %s", [complaint_id])
        updated_at_row = cur.fetchone()
        if updated_at_row:
            updated_at = updated_at_row[0]  # Assuming updated_at is the first column in the result
        else:
            raise ValueError(f"Complaint with ID {complaint_id} not found")
    finally:
        cur.close()
        db.close()

    # Calculate the time already passed (elapsed time)
    elapsed_time = updated_at - issue_raise_date
    elapsed_hours = elapsed_time.total_seconds() / 3600  # Convert to hours
    print(f"Elapsed time: {elapsed_time}, Elapsed hours: {elapsed_hours}")

    # Calculate the remaining TAT
    remaining_tat_hours = tat_hours - elapsed_hours
    if remaining_tat_hours < 0:
        remaining_tat_hours = 0  # If the remaining TAT is negative, set it to 0
    print(f"Remaining TAT hours: {remaining_tat_hours}")

    # Initialize updated_due_date with the current updated_due_date
    updated_due_date = due_date_obj
    print(f"Initial updated due date: {updated_due_date}")

    # Get the public holidays from the database
    cur, db = connection()
    try:
        cur.execute("SELECT holiday_date FROM public_holidays")
        db_holidays = cur.fetchall()

        # Convert db_holidays to a set of dates
        db_holiday_dates = set([holiday[0] for holiday in db_holidays])  # Assuming holiday_date is in a date format
        all_holidays = db_holiday_dates
        print(f"Fetched holidays from DB: {all_holidays}")
    finally:
        cur.close()
        db.close()

    while remaining_tat_hours > 0:
        print(f"Checking if the updated_due_date is a weekend, non-working day, or holiday: {updated_due_date.date()}")

        # Check if the current day is a weekend, holiday, or non-working day for the user
        if (
            updated_due_date.date() in all_holidays or 
            updated_due_date.strftime('%A') in non_working_days_set
        ):
            # Skip to the next workday
            print(f"Weekend, holiday, or non-working day detected. Skipping to next workday.")
            updated_due_date += timedelta(days=1)
            updated_due_date = updated_due_date.replace(hour=work_start_time, minute=0, second=0)
            continue

        # Calculate remaining working hours in the current day
        current_hour = updated_due_date.hour
        current_minute = updated_due_date.minute
        print(f"Current time: {updated_due_date}, Current hour: {current_hour}, Current minute: {current_minute}")

        if current_hour < work_start_time:
            # If before working hours, start at work start time
            print(f"Before work hours: {current_hour}. Setting start time to {work_start_time}.")
            current_hour = work_start_time
            updated_due_date = updated_due_date.replace(hour=work_start_time, minute=0, second=0)

        # Calculate the remaining work minutes today
        remaining_work_minutes_today = (work_end_time * 60) - (current_hour * 60 + current_minute)
        remaining_work_hours_today = remaining_work_minutes_today / 60
        print(f"Remaining work hours today: {remaining_work_hours_today}, Current remaining TAT hours: {remaining_tat_hours}")

        if remaining_tat_hours <= remaining_work_hours_today:
            # If remaining TAT fits within the remaining hours of the current day
            updated_due_date += timedelta(hours=remaining_tat_hours)
            print(f"Remaining TAT fits within today: New updated due date: {updated_due_date}")
            remaining_tat_hours = 0
        else:
            # Use up the remaining hours of the day and move to the next workday
            remaining_tat_hours -= remaining_work_hours_today
            updated_due_date += timedelta(days=1)
            updated_due_date = updated_due_date.replace(hour=work_start_time, minute=0, second=0)
            print(f"Not enough hours today, moving to next workday. Remaining TAT hours: {remaining_tat_hours}, New updated due date: {updated_due_date}")

    return updated_due_date

# def resolver(request, complaint_id):
#     user_name = request.session.get('user_name')
#     user_role = request.session.get('user_role')  # Retrieve the user's role from the session
#     user_id = request.session.get('user_id')

#     # Establish the connection and get the cursor
#     cur, db = connection()  # Ensure `connection()` is correctly defined in your project
#     complaint_details = None  # Initialize the complaint_details variable
#     try:
#         if request.method == "POST":
#             # Get the resolution text, status, updated_due_date, and reason from the form submission
#             resolution = request.POST.get('resolution')
#             status = request.POST.get('status')  # Get the status selected from the dropdown
#             updated_due_date = request.POST.get('updated_due_date')  # Get the updated due date
#             reason = request.POST.get('reason')  # Get the reason for the update

#             # Convert the updated_due_date to a datetime object if it exists
#             due_date_obj = None
#             if updated_due_date:
#                 try:
#                     due_date_obj = datetime.strptime(updated_due_date, "%Y-%m-%dT%H:%M")
#                     print(f'Value of Updated Due Date : {due_date_obj}')
#                 except ValueError:
#                     messages.error(request, "Invalid date format. Please use the correct format.")
#                     print(f'New updated due date is :{due_date_obj}')
#                     return redirect(f'/resolver/{complaint_id}')

#             # Retrieve complaint details to get issue_raise_date and tat_hours
#             query = """
#             SELECT issue_raise_date, tat,updated_at
#             FROM complaint
#             WHERE id = %s
#             """
#             cur.execute(query, (complaint_id,))
#             complaint_data = cur.fetchone()

#             if complaint_data:
#                 issue_raise_date = complaint_data[0]
#                 tat_hours = complaint_data[1]

#                 # Now, update the complaint with the new resolution, status, updated_due_date, and reason
#                 update_query = """
#                 UPDATE complaint 
#                 SET resolution = %s, status = %s, updated_due_date = %s, reason = %s
#                 WHERE id = %s
#                 """
#                 cur.execute(update_query, (resolution, status,due_date_obj,reason,complaint_id))
#                 db.commit()
                
#                 # Now, update the due date based on the issue raise date, updated_at, and remaining TAT
#                 updated_due_date = update_due_date_with_remaining_tat(issue_raise_date, tat_hours, complaint_id,due_date_obj,user_id, country='IN')
                
#                 print(f'Updated Due Date is : {updated_due_date}')
                
#                 new_due_date_query = """
#                 UPDATE complaint
#                 SET updated_due_date = %s WHERE id = %s
#                 """
#                 cur.execute(new_due_date_query,(updated_due_date,complaint_id))
#                 db.commit()
                
#                 # Insert a record into the complaint_history table to log the changes
#                 history_query = """
#                 INSERT INTO complaint_history 
#                 (complaint_id, user_id, resolved_by_id, department, location, complaint_type, 
#                 description, issue_raise_date, complain_status, due_date, tat, updated_at, name, 
#                 status, resolution, assigned_to, updated_due_date, reason, change_timestamp, change_action)
#                 SELECT c.id, c.user_id, c.resolved_by_id, c.department, c.location, c.complaint_type, 
#                 c.description, c.issue_raise_date, c.complain_status, c.due_date, c.tat, c.updated_at, 
#                 c.name, c.status, c.resolution, c.assigned_to, c.updated_due_date, c.reason, 
#                 NOW(), 'UPDATE' 
#                 FROM complaint c WHERE c.id = %s
#                 """
#                 cur.execute(history_query, (complaint_id,))
#                 db.commit()

#                 messages.success(request, "Resolution, status, updated due date, and reason updated successfully.")
#                 return redirect('resolverdashboard')

#         # Query to fetch the specific complaint details using complaint_id
#         query = """
#         SELECT 
#             c.name AS user_name, 
#             u.department, 
#             c.complaint_type, 
#             c.description, 
#             c.location, 
#             c.issue_raise_date, 
#             c.due_date, 
#             c.updated_due_date,  
#             c.status,
#             c.resolution,
#             c.updated_at,
#             c.reason  -- Include reason here
#         FROM 
#             complaint c 
#         JOIN 
#             user u ON c.user_id = u.id 
#         WHERE 
#             c.id = %s
#         """
#         cur.execute(query, (complaint_id,))
#         complaint_details = cur.fetchone()

#         # Fetch the distinct statuses from the status_master table
#         status_query = "SELECT DISTINCT status FROM status_master"
#         cur.execute(status_query)
#         statuses = [row[0] for row in cur.fetchall()]

#         # Prepare context for the template
#         context = {
#             'user_name': user_name,
#             'is_admin': user_role == 'admin',
#             'complaint': {
#                 'user_name': complaint_details[0],
#                 'department': complaint_details[1],
#                 'complaint_type': complaint_details[2],
#                 'description': complaint_details[3],
#                 'location': complaint_details[4],
#                 'issue_raise_date': complaint_details[5].strftime("%Y-%m-%d %H:%M:%S"),
#                 'due_date': complaint_details[6].strftime("%Y-%m-%d %H:%M:%S") if complaint_details[6] else '',
#                 'updated_due_date': complaint_details[7].strftime("%Y-%m-%d %H:%M:%S") if complaint_details[7] else '',
#                 'status': complaint_details[8],
#                 'resolution': complaint_details[9],
#                 'updated_at': complaint_details[10].strftime("%Y-%m-%d %H:%M:%S"),
#                 'reason': complaint_details[11]  # Pass the reason to the context
#             },
#             'statuses': statuses
#         }

#         return render(request, 'resolver.html', context)

#     except Exception as e:
#         db.rollback()
#         messages.error(request, f"Error resolving complaint: {str(e)}")
#         return render(request, 'resolver.html', {'complaint': complaint_details})

#     finally:
#         cur.close()
#         db.close()

def resolver(request, complaint_id):
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')  # Retrieve the user's role from the session
    user_id = request.session.get('user_id')

    # Establish the connection and get the cursor
    cur, db = connection()  # Ensure `connection()` is correctly defined in your project
    complaint_details = None  # Initialize the complaint_details variable

    try:
        if request.method == "POST":
            # Get the resolution text, status, updated_due_date, and reason from the form submission
            resolution = request.POST.get('resolution')
            status = request.POST.get('status')  # Get the status selected from the dropdown
            updated_due_date = request.POST.get('updated_due_date')  # Get the updated due date
            reason = request.POST.get('reason')  # Get the reason for the update

            # Initialize `due_date_obj`
            due_date_obj = None

            # Check if the status is "Overdue" and update the due date accordingly
            if status == "tat extend" and updated_due_date:
                try:
                    # Convert the updated_due_date to a datetime object if it exists
                    due_date_obj = datetime.strptime(updated_due_date, "%Y-%m-%dT%H:%M")
                except ValueError:
                    messages.error(request, "Invalid date format. Please use the correct format.")
                    return redirect(f'/resolver/{complaint_id}')
            elif status == "Resolved":
                # Retain the existing updated_due_date if status is "Resolved"
                cur.execute("SELECT updated_due_date FROM complaint WHERE id = %s", (complaint_id,))
                due_date_obj = cur.fetchone()[0]  # Fetch the existing value

            # Retrieve complaint details to get issue_raise_date and tat_hours
            query = """
            SELECT issue_raise_date, tat, updated_at
            FROM complaint
            WHERE id = %s
            """
            cur.execute(query, (complaint_id,))
            complaint_data = cur.fetchone()

            if complaint_data:
                issue_raise_date = complaint_data[0]
                tat_hours = complaint_data[1]

                # Construct the update query
                update_query = """
                UPDATE complaint 
                SET resolution = %s, status = %s, reason = %s {updated_due_date_clause}
                WHERE id = %s
                """
                # Add `updated_due_date` to the query only if it's provided
                if due_date_obj:
                    updated_due_date_clause = ", updated_due_date = %s"
                    params = (resolution, status, reason, due_date_obj, complaint_id)
                else:
                    updated_due_date_clause = ""
                    params = (resolution, status, reason, complaint_id)

                cur.execute(update_query.format(updated_due_date_clause=updated_due_date_clause), params)
                db.commit()

                # Update the due date based on the issue raise date, updated_at, and remaining TAT
                if status == "tat extend" and due_date_obj:
                    updated_due_date = update_due_date_with_remaining_tat(issue_raise_date, tat_hours, complaint_id, due_date_obj, user_id, country='IN')
                    cur.execute("UPDATE complaint SET updated_due_date = %s WHERE id = %s", (updated_due_date, complaint_id))
                    db.commit()

                # Insert a record into the complaint_history table to log the changes
                history_query = """
                INSERT INTO complaint_history 
                (complaint_id, user_id, resolved_by_id, department, location, complaint_type, 
                description, issue_raise_date, complain_status, due_date, tat, updated_at, name, 
                status, resolution, assigned_to, updated_due_date, reason, change_timestamp, change_action)
                SELECT c.id, c.user_id, c.resolved_by_id, c.department, c.location, c.complaint_type, 
                c.description, c.issue_raise_date, c.complain_status, c.due_date, c.tat, c.updated_at, 
                c.name, c.status, c.resolution, c.assigned_to, c.updated_due_date, c.reason, 
                NOW(), 'UPDATE' 
                FROM complaint c WHERE c.id = %s
                """
                cur.execute(history_query, (complaint_id,))
                db.commit()

                messages.success(request, "Resolution, status, updated due date, and reason updated successfully.")
                return redirect('resolverdashboard')

        # Query to fetch the specific complaint details using complaint_id
        query = """
        SELECT 
            c.name AS user_name, 
            u.department, 
            c.complaint_type, 
            c.description, 
            c.location, 
            c.issue_raise_date, 
            c.due_date, 
            c.updated_due_date,  
            c.status,
            c.resolution,
            c.updated_at,
            c.reason
        FROM 
            complaint c 
        JOIN 
            user u ON c.user_id = u.id 
        WHERE 
            c.id = %s
        """
        cur.execute(query, (complaint_id,))
        complaint_details = cur.fetchone()

        # Fetch the distinct statuses from the status_master table
        status_query = "SELECT DISTINCT status FROM status_master"
        cur.execute(status_query)
        statuses = [row[0].lower() for row in cur.fetchall()]

        # Prepare context for the template
        context = {
            'user_name': user_name,
            'is_admin': user_role == 'admin',
            'complaint': {
                'user_name': complaint_details[0],
                'department': complaint_details[1],
                'complaint_type': complaint_details[2],
                'description': complaint_details[3],
                'location': complaint_details[4],
                'issue_raise_date': complaint_details[5].strftime("%Y-%m-%d %H:%M:%S"),
                'due_date': complaint_details[6].strftime("%Y-%m-%d %H:%M:%S") if complaint_details[6] else '',
                'updated_due_date': complaint_details[7].strftime("%Y-%m-%d %H:%M:%S") if complaint_details[7] else '',
                'status': complaint_details[8].lower(),
                'resolution': complaint_details[9],
                'updated_at': complaint_details[10].strftime("%Y-%m-%d %H:%M:%S"),
                'reason': complaint_details[11]
            },
            'statuses': statuses
        }
        
        print(context)

        return render(request, 'resolver.html', context)

    except Exception as e:
        db.rollback()
        messages.error(request, f"Error resolving complaint: {str(e)}")
        return render(request, 'resolver.html', {'complaint': complaint_details})

    finally:
        cur.close()
        db.close()

# def fetch_holidays(resolver_id):
#     """
#     Fetch public holidays for a given resolver from the database.

#     :param resolver_id: ID of the resolver.
#     :return: Set of holiday dates as datetime.date objects.
#     """
#     holidays = set()
#     cur, db = connection()
#     try:
#         query = """
#         SELECT ph.holiday_date
#         FROM public_holidays ph
#         JOIN office_timings ot ON ot.holiday_id = ph.id
#         WHERE ot.resolver_id = 7;
#         """
#         cur.execute(query, (resolver_id,))
#         rows = cur.fetchall()

#         for row in rows:
#             holidays.add(row[0])

#     except Exception as e:
#         print(f"Error fetching holidays: {e}")
#     finally:
#         cur.close()
#         db.close()

#     return holidays

def add_tat_excluding_weekends_and_holidays(issue_raise_date, tat_hours, assigned_admin_id):
    WORK_START = 9  # Default work starts at 9 AM
    WORK_END = 18   # Default work ends at 6 PM
    WORK_HOURS = WORK_END - WORK_START  # Total work hours in a day

    # Fetch the admin's work schedule and public holidays from the database
    cur, db = connection()
    try:
        # Fetch admin's work schedule based on location and department
        cur.execute("""
            SELECT 
                EXTRACT(HOUR FROM ot.work_start_time) AS work_start_hour,
                EXTRACT(HOUR FROM ot.work_end_time) AS work_end_hour,
                ot.non_working_days 
            FROM user u
            JOIN office_timings ot
                ON u.location = ot.location
                AND u.department = ot.department
            WHERE u.id = %s
        """, [assigned_admin_id])
        admin_schedule = cur.fetchone()
        
        print(f'Admin Work Start Time is : {admin_schedule[0]}')
        print(f'Admin Work End Time is : {admin_schedule[1]}')
        print(f'Admin Non Working Day is : {admin_schedule[2]}')

        if admin_schedule:
            work_start_time = int(admin_schedule[0]) if admin_schedule[0] else WORK_START
            work_end_time = int(admin_schedule[1]) if admin_schedule[1] else WORK_END
            non_working_days = admin_schedule[2] or 'None'  # Default to 'None' if not set

            # Interpret non_working_days based on the enum value
            if non_working_days == 'Saturday':
                non_working_days_set = {'Saturday'}
            elif non_working_days == 'Sunday':
                non_working_days_set = {'Sunday'}
            elif non_working_days == 'Both':
                non_working_days_set = {'Saturday', 'Sunday'}
            else:  # 'None'
                non_working_days_set = set()
        else:
            print("No schedule found for the admin. Using default work hours.")
            work_start_time = WORK_START
            work_end_time = WORK_END
            non_working_days_set = set()

        # Fetch public holidays
        cur.execute("SELECT holiday_date FROM public_holidays")
        db_holidays = cur.fetchall()

        # Convert db_holidays to a set of dates
        db_holiday_dates = set([holiday[0] for holiday in db_holidays])  # Assuming holiday_date is in a date format
        print(f"Fetched holidays from DB: {db_holiday_dates}")

    finally:
        cur.close()
        db.close()

    # Initialize due_date with the issue raise date
    due_date = issue_raise_date
    print(f"Initial issue raise date: {due_date}, TAT Hours: {tat_hours}")

    while tat_hours > 0:
        print(f"Checking if the due_date is a weekend, non-working day, or holiday: {due_date.date()}")

        # Check if the current day is a weekend, holiday, or non-working day for the admin
        if due_date.date() in db_holiday_dates or due_date.strftime('%A') in non_working_days_set:
            # Move to the next workday
            print(f"Weekend, holiday, or non-working day detected. Skipping to next workday.")
            due_date += timedelta(days=1)
            due_date = due_date.replace(hour=work_start_time, minute=0, second=0)
            continue

        # Calculate remaining working hours in the current day
        current_hour = due_date.hour
        current_minute = due_date.minute

        if current_hour < work_start_time:
            # If before working hours, start at work start time
            print(f"Before work hours: {current_hour}. Setting start time to {work_start_time}.")
            current_hour = work_start_time
            due_date = due_date.replace(hour=work_start_time, minute=0, second=0)

        # Calculate the remaining minutes today, considering both hour and minute difference
        remaining_work_minutes_today = (work_end_time * 60) - (current_hour * 60 + current_minute)
        remaining_work_hours_today = remaining_work_minutes_today / 60
        print(f"Remaining work hours today: {remaining_work_hours_today}, Current TAT hours: {tat_hours}")

        if tat_hours <= remaining_work_hours_today:
            # If TAT fits within the remaining hours of the current day
            due_date += timedelta(hours=tat_hours)
            print(f"TAT fits within today: New due date: {due_date}")
            tat_hours = 0
        else:
            # Use up the remaining hours of the day and move to the next workday
            tat_hours -= remaining_work_hours_today
            due_date += timedelta(days=1)
            due_date = due_date.replace(hour=work_start_time, minute=0, second=0)
            print(f"Not enough hours today, moving to next workday. Remaining TAT hours: {tat_hours}, New due date: {due_date}")

    return due_date

def complain(request):
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')
    user_location = request.session.get('user_location')  # Get location from session
    user_department = request.session.get('user_department')  # Get department from session
    cur, db = connection()

    # Fetch distinct values for location, department, and complaint_type
    cur.execute("SELECT DISTINCT location FROM location_master;")
    record = [{'location': row[0]} for row in cur.fetchall()]
    
    cur.execute("SELECT DISTINCT department FROM department_master;")
    record1 = [{'department': row[0]} for row in cur.fetchall()]
    
    cur.execute("SELECT DISTINCT complaint_type FROM complaint_type_master;")
    record2 = [{'complaint_type': row[0]} for row in cur.fetchall()]
    
    cur.execute("SELECT DISTINCT department FROM complaint_type_master;")
    record3 = [{'department': row[0]} for row in cur.fetchall()]
    
    # Determine if we're on the 'raisecomplaint' page and update context accordingly
    context = {
        'user_name': user_name,
        'record': record,
        'record1': record1,
        'record2': record2,
        'record3': record3,
        'user_location': user_location,  # Pass session location to template
        'user_department': user_department,  # Pass session department to template
    }

    # If on 'raisecomplaint' page, hide 'All Complaints' section in the sidebar
    if request.path == '/raisecomplaint/':
        context['show_all_complaints'] = False
    else:
        context['show_all_complaints'] = True  # Show 'All Complaints' if not on 'raisecomplaint' page

    if user_role != 'user':
        context['error'] = 'You do not have permission to raise a complaint.'
        return redirect('user')

    if request.method == 'POST':
        user_id = request.session.get('user_id')
        if user_id is None:
            context['error'] = 'User not found. Please log in.'
            return render(request, 'complain.html', context)

        # Gather form data
        name = user_name
        custom_location = request.POST.get('custom_location')
        custom_department = request.POST.get('custom_department')
        complaint_type = request.POST.get('complaint-type')
        custom_complaint_type = request.POST.get('custom_complaint_type')
        description = request.POST.get('description')
        
        # Use location and department from session if they exist
        location = user_location
        department = user_department

        # Handle custom input for location
        if location == 'other' and custom_location:
            location = custom_location.strip()
            cur.execute("SELECT location FROM location_master WHERE location = %s;", (location,))
            if cur.fetchone() is None:
                cur.execute("INSERT INTO location_master (location) VALUES (%s);", (location,))
                db.commit()
                
        # Handle custom input for department
        if department == 'other' and custom_department:
            department = custom_department.strip()
            cur.execute("SELECT department FROM department_master WHERE department = %s;", (department,))
            if cur.fetchone() is None:
                cur.execute("INSERT INTO department_master (department) VALUES (%s);", (department,))
                db.commit()
        
        # Handle custom input for complaint type
        if complaint_type == 'other' and custom_complaint_type:
            complaint_type = custom_complaint_type.strip()
            cur.execute("SELECT complaint_type FROM complaint_type_master WHERE complaint_type = %s;", (complaint_type,))
            if cur.fetchone() is None:
                cur.execute("INSERT INTO complaint_type_master (complaint_type) VALUES (%s);", (complaint_type,))
                db.commit()

        status = 'pending'

        # Get current time in IST
        india_tz = pytz.timezone('Asia/Kolkata')
        issue_raise_date = datetime.now(india_tz)
        
        # Query to find admins who match the location and department
        cur.execute("""
            SELECT id, name, created_at 
            FROM user
            WHERE role = 'admin' 
            AND location LIKE %s 
            AND department LIKE %s
        """, (f"%{location.strip()}%", f"%{department.strip()}%"))

        admins = cur.fetchall()

        if not admins:
            context['error'] = 'No admin available for the specified location and department.'
            return render(request, 'complain.html', context)

        # Find the admin with the least pending complaints
        min_pending_complaints = None
        selected_admin = None

        for admin in admins:
            admin_id = admin[0]
            created_at = admin[2]

            # Query to count pending complaints for each admin
            cur.execute("""
                SELECT COUNT(*) 
                FROM complaint 
                WHERE assigned_to = %s AND status = 'pending'
            """, (admin_id,))
            pending_count = cur.fetchone()[0]

            # Select the admin with the least pending complaints, and if equal, choose the recent one
            if min_pending_complaints is None or pending_count < min_pending_complaints:
                min_pending_complaints = pending_count
                selected_admin = admin
            elif pending_count == min_pending_complaints and created_at > selected_admin[2]:
                selected_admin = admin

        # Assign the complaint to the selected admin
        assigned_admin_id = selected_admin[0]  # admin's ID
        print(f'Complaint is assigned to Admin with ID : {assigned_admin_id}')

        # Fetch the TAT for the selected complaint type
        cur.execute("SELECT TAT FROM complaint_type_master WHERE complaint_type = %s;", (complaint_type,))
        tat_result = cur.fetchone()

        if tat_result:
            tat_hours = tat_result[0]  # TAT is in hours
        else:
            tat_hours = 24  # Default TAT if no specific complaint type is found

        # Call the updated function to add TAT excluding weekends
        due_date = add_tat_excluding_weekends_and_holidays(issue_raise_date, tat_hours, assigned_admin_id)
        
        # Insert the complaint into the complaint table
        cur.execute(""" 
            INSERT INTO complaint (user_id, name, department, location, complaint_type, description, issue_raise_date, due_date, tat, status, assigned_to)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (user_id, name, department, location, complaint_type, description, issue_raise_date, due_date, tat_hours, status, assigned_admin_id))
        db.commit()

        # Get the complaint_id of the newly inserted complaint
        cur.execute("SELECT LAST_INSERT_ID()")
        complaint_id = cur.fetchone()[0]
        print(f"Complaint ID is :{complaint_id}")

        # Insert the complaint into the complaint_history table
        cur.execute(""" 
            INSERT INTO complaint_history (
                complaint_id, user_id, resolved_by_id, department, location, complaint_type, 
                description, issue_raise_date, complain_status, due_date, tat, updated_at, 
                name, status, resolution, assigned_to, updated_due_date, reason, 
                change_timestamp, change_action
            )
            VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, 
                %s, %s, NULL, %s, NULL, NULL, CURRENT_TIMESTAMP, 'INSERT'
            )
        """, (
            complaint_id, user_id, assigned_admin_id, department, location, complaint_type, 
            description, issue_raise_date, 'done', due_date, tat_hours, name, status, assigned_admin_id
        ))
        db.commit()

        # Close the cursor and the database connection
        cur.close()
        db.close()

        return redirect('userdashboard')

    return render(request, 'complain.html', context)

def complainview(request, complaint_id):
    user_name = request.session.get('user_name')
    # Establish the connection and get the cursor
    cur, db = connection()  # Ensure `connection()` is correctly defined in your project

    try:
        # Query to fetch the specific complaint details using complaint_id
        query = """
        SELECT 
            u.name AS user_name, 
            u.department, 
            c.complaint_type, 
            c.description, 
            u.location, 
            c.issue_raise_date, 
            c.due_date, 
            c.status,
            c.updated_at,  -- Include updated_at in the query
            c.resolution    -- Add resolution to the query
        FROM 
            complaint c 
        JOIN 
            user u ON c.user_id = u.id 
        WHERE 
            c.id = %s
        """
        cur.execute(query, (complaint_id,))
        complaint_details = cur.fetchone()

        if complaint_details:
            # Calculate TAT as updated_at - issue_raise_date
            issue_raise_date = complaint_details[5]  # Fetching issue_raise_date
            updated_at = complaint_details[8]  # Fetching updated_at
            
            # Use updated_at for TAT calculation if it exists; otherwise, use the current time
            tat_duration = updated_at - issue_raise_date if updated_at else datetime.now() - issue_raise_date
            
            # Extract days, hours, minutes, and seconds from TAT duration
            total_seconds = tat_duration.total_seconds()
            days = int(total_seconds // 86400)  # Calculate days
            hours = int((total_seconds % 86400) // 3600)  # Calculate hours
            minutes = int((total_seconds % 3600) // 60)  # Calculate minutes
            seconds = int(total_seconds % 60)  # Calculate seconds

            # Format TAT
            tat_formatted = ""
            if days > 0:
                tat_formatted += f"{days} Days "
            if hours > 0 or days > 0:  # Always show hours if there are days
                tat_formatted += f"{hours} Hours "
            tat_formatted += f"{minutes} Minutes {seconds} Seconds"  # Add seconds

            # Prepare context for the template
            context = {
                'user_name' : user_name,
                'complaint': {
                    'user_name': complaint_details[0],
                    'department': complaint_details[1],
                    'complaint_type': complaint_details[2],
                    'description': complaint_details[3],
                    'location': complaint_details[4],
                    'issue_raise_date': complaint_details[5].strftime("%Y-%m-%d %H:%M:%S"),
                    'due_date': complaint_details[6].strftime("%Y-%m-%d %H:%M:%S"),
                    'status': complaint_details[7].capitalize(),
                    'resolved_at': updated_at.strftime("%Y-%m-%d %H:%M:%S") if updated_at else 'Not resolved yet',  # Add resolved_at
                    'resolution': complaint_details[9] if complaint_details[9] else 'No resolution provided',  # Add resolution
                    'tat': tat_formatted,  # Add the formatted TAT to the context
                }
            }
        else:
            context = {'error': 'Complaint not found'}

    finally:
        # Close the cursor and database connection
        cur.close()
        db.close()

    return render(request, 'complainview.html', context)

# def adduser(request):
#     return render(request,'adduser.html')

def adduser(request):
    user_name = request.session.get('user_name')
    cur, db = connection()
    
    try:
        # Fetch locations and departments from the database
        cur.execute("SELECT distinct location FROM location_master;")
        locations = [{'location': row[0]} for row in cur.fetchall()]

        cur.execute("SELECT distinct department FROM department_master;")
        departments = [{'department': row[0]} for row in cur.fetchall()]
        
        cur.execute("SELECT  distinct level FROM escalationmatrix1;")
        levels = [{'level': row[0]} for row in cur.fetchall()]
          
        # Calculate current IST date and time
        ist = pytz.timezone('Asia/Kolkata')
        current_ist_datetime = datetime.now(ist).strftime('%Y-%m-%dT%H:%M')

        if request.method == "POST":
            name = request.POST.get('name')
            location = request.POST.get('location')
            department = request.POST.get('department')
            date = request.POST.get('date')
            role = request.POST.get('role')
            level = request.POST.get('level')  # Capture the selected level
            email = request.POST.get('email')  # Capture email
            password = request.POST.get('password')  # Capture password

            # Insert new user data into the database
            cur.execute("""
                INSERT INTO user (name, location, department, date, role, levels, email, password) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (name, location, department, date, role, level, email, password))

            db.commit()
            messages.success(request, 'User added successfully!')
            return redirect('superadmin')

        # Fetch all users from the database
        cur.execute("SELECT * FROM user")
        users = cur.fetchall()

        # Return the render with the correct context
        return render(request, 'adduser.html', {
            'users': users,
            'user_name': user_name,
            'locations': locations,
            'departments': departments,
            'current_ist_datetime': current_ist_datetime,# Pass the current IST datetime
            'levels' : levels
        })

    except Exception as e:
        db.rollback()
        messages.error(request, f'Error adding user: {str(e)}')
        return render(request, 'adduser.html', {'users': [], 'user_name': user_name, 'locations': locations, 'departments': departments,'current_ist_datetime':''})

    finally:
        cur.close()
        db.close()
        
def adduser(request):
    user_name = request.session.get('user_name')
    cur, db = connection()

    try:
        # Fetch locations, departments, and levels
        cur.execute("SELECT distinct location FROM location_master;")
        locations = [{'location': row[0]} for row in cur.fetchall()]

        cur.execute("SELECT distinct department FROM department_master;")
        departments = [{'department': row[0]} for row in cur.fetchall()]

        cur.execute("SELECT distinct level FROM escalationmatrix1;")
        levels = [{'level': row[0]} for row in cur.fetchall()]

        # Calculate current IST date and time
        ist = pytz.timezone('Asia/Kolkata')
        current_ist_datetime = datetime.now(ist).strftime('%Y-%m-%dT%H:%M')

        if request.method == "POST":
            name = request.POST.get('name')
            email = request.POST.get('email')
            password = request.POST.get('password')
            location = request.POST.get('location')
            department = request.POST.get('department')
            date = request.POST.get('date')
            role = request.POST.get('role')
            level = request.POST.get('level')

            # Check if username or email already exists
            cur.execute("SELECT id FROM user WHERE name = %s OR email = %s", (name, email))
            if cur.fetchone():
                messages.error(request, 'Username or email already exists!')
                return redirect('adduser')

            # Insert new user
            cur.execute("""
                INSERT INTO user (name, location, department, date, role, levels, email, password) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (name, location, department, date, role, level, email, password))

            db.commit()
            messages.success(request, 'User added successfully!')
            return redirect('superadmin')

        # Fetch all users
        cur.execute("SELECT * FROM user")
        users = cur.fetchall()

        return render(request, 'adduser.html', {
            'users': users,
            'user_name': user_name,
            'locations': locations,
            'departments': departments,
            'current_ist_datetime': current_ist_datetime,
            'levels': levels,
        })

    except Exception as e:
        db.rollback()
        messages.error(request, f'Error adding user: {str(e)}')
        return render(request, 'adduser.html', {
            'users': [],
            'user_name': user_name,
            'locations': locations,
            'departments': departments,
            'current_ist_datetime': current_ist_datetime,
            'levels': levels,
        })

    finally:
        cur.close()
        db.close()

def check_username(request):
    if request.method == "GET":
        name = request.GET.get('name')
        cur, db = connection()
        try:
            cur.execute("SELECT COUNT(*) FROM user WHERE name = %s", (name,))
            exists = cur.fetchone()[0] > 0
            return JsonResponse({'exists': exists})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        finally:
            cur.close()
            db.close()
        
# def edit(request):
#     return render(request,'edit.html')


# def edit(request, user_id):
#     cur, db = connection()

#     if request.method == 'POST':
#         name = request.POST.get('name')
#         location = request.POST.get('location')
#         department = request.POST.get('department')
#         date = request.POST.get('date')
#         role = request.POST.get('role')
#         emp_level = request.POST.get('emp_level')

#         try:
#             cur.execute("""
#                 UPDATE user SET name = %s, location = %s, department = %s, date = %s, role = %s, emp_level = %s
#                 WHERE id = %s
#             """, (name, location, department, date, role, emp_level, user_id))
#             db.commit()
#             messages.success(request, 'User updated successfully!')
#             return redirect('superadmin')
#         except Exception as e:
#             db.rollback()
#             messages.error(request, f'Error updating user: {str(e)}')
#         finally:
#             cur.close()
#             db.close()

#     try:
#         cur.execute("SELECT * FROM user WHERE id = %s", (user_id,))
#         user = cur.fetchone()
#         if user:
#             return render(request, 'edit.html', {'user': user})
#         else:
#             messages.error(request, 'User not found.')
#             return redirect('superadmin')
#     finally:
#         cur.close()
#         db.close()

def edit_user(request, user_id):
    # Fetch session details for the logged-in user
    user_role = request.session.get('user_role')
    user_name = request.session.get('user_name')

    # Establish the database connection
    cur, db = connection()
    try:
        # Fetch locations and departments from their respective tables
        cur.execute("SELECT distinct location FROM location_master;")
        locations = [row[0] for row in cur.fetchall()]

        cur.execute("SELECT distinct department FROM department_master;")
        departments = [row[0] for row in cur.fetchall()]
        
        cur.execute("SELECT distinct level FROM escalationmatrix1;")
        levels = [{'level': row[0]} for row in cur.fetchall()]  # Keep the structure consistent for the dropdown

        # Fetch the user details for pre-filling the form
        cur.execute(
            """
            SELECT id, name, role, department, location, levels, email 
            FROM user 
            WHERE id = %s
            """, 
            (user_id,)
        )
        user = cur.fetchone()
        print(f"User Details : {user}")

        if user is None:
            return HttpResponse("User not found.", status=404)

        # If the form is submitted (POST request)
        if request.method == 'POST':
            # Retrieve data from the submitted form
            name = request.POST.get('name')
            location = request.POST.get('location')
            department = request.POST.get('department')
            role = request.POST.get('role')
            level = request.POST.get('level')
            email = request.POST.get('email')

            # Validate the form data
            if not (name and location and department and role and level and email):
                return HttpResponse("All fields are required.", status=400)

            # Update the user's details in the database
            cur.execute(
                """
                UPDATE user 
                SET name = %s, location = %s, department = %s, role = %s, levels = %s, email = %s
                WHERE id = %s
                """,
                (name, location, department, role, level, email, user_id)
            )
            db.commit()

            # Redirect to the superadmin page after a successful update
            return redirect('superadmin')

        # Render the edit form with the user details and additional data
        context = {
            'user': {
                'id': user[0],
                'name': user[1],
                'role': user[2],
                'department': user[3],
                'location': user[4],
                'level': user[5],  # Pass the current level to the template
                'email': user[6],
            },
            'locations': locations,
            'departments': departments,
            'user_name': user_name,
            'user_role': user_role,
            'levels': levels  # Pass the levels to the template for the dropdown
        }
        return render(request, 'edit_user.html', context)

    finally:
        # Close the database connection
        cur.close()
        db.close()   
           
def reports(request):
    user_name = request.session.get('user_name')

    # Retrieve filter parameters from GET request
    fromdate = request.GET.get('fromdate', '')
    todate = request.GET.get('todate', '')
    status = request.GET.get('status', '')
    location_type = request.GET.get('location', '')
    department = request.GET.get('department', '')
    complaint_type = request.GET.get('complaint_type', '')

    # Prepare filter conditions
    filter_conditions = []
    query_params = []

    # Date filtering for fromdate and todate
    if fromdate and todate:
        filter_conditions.append("c.issue_raise_date BETWEEN %s AND %s")
        query_params.extend([f"{fromdate} 00:00:00", f"{todate} 23:59:59"])
    elif fromdate:
        filter_conditions.append("c.issue_raise_date >= %s")
        query_params.append(f"{fromdate} 00:00:00")
    elif todate:
        filter_conditions.append("c.issue_raise_date <= %s")
        query_params.append(f"{todate} 23:59:59")

    # Apply location filter using 'location' from the 'complaint' table
    if location_type:
        filter_conditions.append("c.location = %s")
        query_params.append(location_type)

    # Additional filters
    if department:
        filter_conditions.append("u.department = %s")
        query_params.append(department)
    if complaint_type:
        filter_conditions.append("c.complaint_type = %s")
        query_params.append(complaint_type)
    if status:
        filter_conditions.append("LOWER(c.status) = LOWER(%s)")
        query_params.append(status)

    # Build the WHERE clause if there are any filter conditions
    where_clause = f"WHERE {' AND '.join(filter_conditions)}" if filter_conditions else ""

    # Final query construction
    query = f"""
    SELECT 
        u.name AS complainant_name, 
        u.department, 
        c.location, 
        COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
        c.issue_raise_date AS date,
        c.status,
        c.resolution,
        c.description,
        c.complaint_type,  
        CASE 
            WHEN LOWER(c.status) = 'resolved' THEN c.updated_at
            ELSE NULL
        END AS resolved_at
    FROM 
        complaint AS c
    JOIN 
        user AS u ON c.user_id = u.id
    LEFT JOIN 
        user AS u2 ON c.resolved_by_id = u2.id AND u2.role = 'admin'
    {where_clause}
    ORDER BY 
        c.issue_raise_date DESC;
    """

    # Execute the query with parameters
    cur.execute(query, query_params)
    resolution_reports = cur.fetchall()

    # Process the reports into a list of dictionaries
    report_data = []
    for report in resolution_reports:
        complainant_name = report[0]
        department = report[1]
        location = report[2].capitalize()
        resolved_by = report[3]
        date = report[4]
        status = report[5].capitalize()
        resolution = report[6] if report[6] else "Not Resolved Yet"
        description = report[7]
        complaint_type = report[8]
        resolved_at = report[9] if report[9] else "Not Resolved Yet"

        report_data.append({
            'complainant_name': complainant_name,
            'department': department,
            'location': location,
            'resolved_by': resolved_by,
            'date': date,
            'status': status,
            'resolution': resolution,
            'description': description,
            'complaint_type': complaint_type,
            'resolved_at': resolved_at
        })

    # If user requests download, create Excel file
    if request.GET.get('download') == 'true':
        return generate_xlsx(report_data)

    # Fetch distinct values for dropdown filters
    location_query = "SELECT DISTINCT location FROM complaint"
    department_query = "SELECT DISTINCT department FROM department_master"
    complaint_type_query = "SELECT DISTINCT complaint_type FROM complaint_type_master"
    status_query = "SELECT DISTINCT status FROM status_master"  # Retrieve both lowercase and uppercase values

    cur.execute(location_query)
    distinct_locations = [row[0] for row in cur.fetchall()]
    
    cur.execute(department_query)
    distinct_departments = [row[0] for row in cur.fetchall()]

    cur.execute(complaint_type_query)
    distinct_complaint_types = [row[0] for row in cur.fetchall()]
    
    cur.execute(status_query)
    statuses = [row[0] for row in cur.fetchall()]

    return render(request, 'report.html', {
        'report_data': report_data,
        'distinct_locations': distinct_locations,
        'distinct_departments': distinct_departments,
        'distinct_complaint_types': distinct_complaint_types,
        'from_date': fromdate,
        'to_date': todate,
        'user_name': user_name,
        'statuses': statuses
    })
    
def generate_xlsx(report_data):
    """
    Generate an Excel file from the report data and return it as an HTTP response.
    """
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Resolution Reports"

    # Define the column headers in the required order
    headers = [
        "Name", "Dept", "Location", "Type", "Description", "Issue Date",
        "Status", "Resolved By", "Resolved At", "Resolution"
    ]
    
    # Define header styles
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Grey background
    header_font = Font(bold=True, color="000000",size=12)  # White text
    
    # Apply header styles and set column headers
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f'{col_letter}1']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Populate the worksheet with report data
    for row_num, report in enumerate(report_data, 2):
        ws[f'A{row_num}'] = report['complainant_name']
        ws[f'B{row_num}'] = report['department']
        ws[f'C{row_num}'] = report['location']
        ws[f'D{row_num}'] = report['complaint_type']
        ws[f'E{row_num}'] = report['description']
        ws[f'F{row_num}'] = report['date']
        ws[f'G{row_num}'] = report['status']
        ws[f'H{row_num}'] = report['resolved_by']
        ws[f'I{row_num}'] = report['resolved_at']
        ws[f'J{row_num}'] = report['resolution']

        # Format date columns for readability
        ws[f'E{row_num}'].number_format = 'DD-MM-YYYY HH:MM:SS'
        ws[f'J{row_num}'].number_format = 'DD-MM-YYYY HH:MM:SS'

        # Enable text wrapping for "Description" and "Resolution" columns
        ws[f'E{row_num}'].alignment = Alignment(wrap_text=True)
        ws[f'J{row_num}'].alignment = Alignment(wrap_text=True)

    # Adjust column widths for better readability
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2  # Add extra padding
        ws.column_dimensions[col_letter].width = adjusted_width

    # Set fixed widths for description and resolution columns (E and J)
    ws.column_dimensions['E'].width = 27  # Adjust width for Description column
    ws.column_dimensions['J'].width = 19  # Adjust width for Resolution column
    ws.column_dimensions['I'].width = 19  # Adjust width for Resolved at column

    # Save the workbook to a BytesIO stream
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Create the response with the Excel file as an attachment
    response = HttpResponse(stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=resolution_reports.xlsx'
    return response

def some_view(request):
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')

    return render(request, 'superadmindashboared', {
        'user_name': user_name,
        'user_role': user_role,
    })

# def editcomplain(request,user_id):
#     return render(request,'editcomplain.html')

def editcomplain(request, complaint_id):
    # Establish a custom database connection
    user_name = request.session.get('user_name')
    cur, db = connection()
    
    try:
        # Fetch the complaint details by complaint_id (not user_id) to pre-fill the form
        cur.execute("SELECT * FROM complaint WHERE id = %s LIMIT 1", [complaint_id])
        complaint = cur.fetchone()

        if not complaint:
            return redirect('user')  # Redirect if complaint not found

        if request.method == 'POST':
            # Get form data for updating the complaint
            name = request.POST.get('name')
            location = request.POST.get('location')
            department = request.POST.get('department')
            complaint_type = request.POST.get('complaint-type')
            issue_raise_date = request.POST.get('issue-date')
            description = request.POST.get('description')

            # Execute the update query using the custom connection
            cur.execute(""" 
                UPDATE complaint
                SET 
                    name = %s,
                    location = %s,
                    department = %s,
                    complaint_type = %s,
                    issue_raise_date = %s,
                    description = %s,
                    updated_at = %s
                WHERE id = %s
            """, [name, location, department, complaint_type, issue_raise_date, description, datetime.now(), complaint_id])

            # Commit the changes
            db.commit()

            return redirect('user')  # Redirect to the user page after updating

        # Pass complaint details to the template for rendering
        context = {
            'user_name' : user_name,
            'complaint': {
                'id': complaint[0],
                'user_id': complaint[1],
                'resolved_by_id': complaint[2],
                'department': complaint[3],
                'location': complaint[4],
                'complaint_type': complaint[5],
                'description': complaint[6],
                'issue_raise_date': complaint[7],
                'complain_status': complaint[8],
                'due_date': complaint[9],
                'tat': complaint[10],
                'updated_at': complaint[11],
                'name': complaint[12],
                'status': complaint[13],
                'resolution': complaint[14],
            }
        }
        return render(request, 'editcomplain.html', context)

    finally:
        # Close the cursor and database connection
        cur.close()
        db.close()

def update_complaint_status(request, complaint_id):
    # Check if the user is logged in
    if 'user_id' not in request.session:
        return redirect('login')  # Redirect to the login page if not logged in

    # Establish the connection and get the cursor
    cur, db = connection()  # Ensure `connection()` is correctly defined in your project

    try:
        # Check the current status and complain_status of the complaint
        status_check_query = "SELECT complain_status, status FROM complaint WHERE id = %s"
        cur.execute(status_check_query, [complaint_id])
        result = cur.fetchone()

        # Ensure the complaint exists before proceeding
        if result:
            complain_status, current_status = result

            # If the current_status is "resolved", do not update
            if current_status == 'resolved':
                # No change needed, simply return without making updates
                return redirect('resolver', complaint_id=complaint_id)  # Adjust 'resolver' to match your URL name

            # If complain_status is "done" and status is "resolved", do not update
            if complain_status == 'done' and current_status == 'resolved':
                return redirect('resolver', complaint_id=complaint_id)
            
              # If complain_status is "overdue" and status is "resolved", do not update
            if complain_status == 'overdue' and current_status == 'resolved':
                return redirect('resolver', complaint_id=complaint_id)

            # If complain_status is "overdue", update the status to "work in progress"
            if complain_status == 'overdue':
                update_query = "UPDATE complaint SET status = 'work in progress' WHERE id = %s"
                cur.execute(update_query, [complaint_id])

            # If complain_status is "pending" and status is not "overdue", update to "work in progress"
            elif current_status == 'pending' and complain_status != 'overdue':
                update_query = "UPDATE complaint SET status = 'work in progress' WHERE id = %s"
                cur.execute(update_query, [complaint_id])

            # Commit the changes to the database
            db.commit()

        # Redirect to the resolver view after checking/updating
        return redirect('resolver', complaint_id=complaint_id)  # Adjust 'resolver' to match your URL name

    except Exception as e:
        print(f"Error updating complaint status: {e}")
        return HttpResponse("Error updating complaint status.", status=500)
    finally:
        # Close the cursor and database connection
        cur.close()
        db.close()
        
def delete_complaint(request, complaint_id):
    # Establish a database connection
    cur, db = connection()

    # Check if the complaint exists
    check_query = "SELECT COUNT(*) FROM complaint WHERE id = %s"
    cur.execute(check_query, (complaint_id,))
    exists = cur.fetchone()[0] > 0

    if exists:
        # Delete the complaint if it exists
        delete_query = "DELETE FROM complaint WHERE id = %s"
        cur.execute(delete_query, (complaint_id,))
        db.commit()
        
        cur.close()
        db.close()
        
        # Redirect to the complaints list page after deletion
        return redirect(reverse('user'))
    else:
        cur.close()
        db.close()
        # Return a 404 response if the complaint does not exist
        return HttpResponseNotFound("Complaint not found.")

def download_resolution_reports_xlsx(request):
    # Check if the user is logged in
    if 'user_id' not in request.session:
        return redirect('login')  # Redirect to the login page if not logged in

    # Get date range from request parameters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    # Validate and parse dates
    date_filter = ""
    params = []
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d")
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d")
            date_filter = "AND c.issue_raise_date BETWEEN %s AND %s"
            params = [from_date, to_date]
        except ValueError:
            return HttpResponse("Invalid date format. Please use YYYY-MM-DD.", status=400)

    # Establish database connection
    cur, db = connection()

    # Create an Excel workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Resolution Reports'

    # Header row with a gray background
    headers = ['Name', 'Department', 'Location', 'Resolved By', 'Date', 'Status', 'Resolution', 'Resolved At']
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.fill = gray_fill

    try:
        # Fetch resolution reports with date filter
        query = f"""
            SELECT 
                c.name AS complainant_name, 
                c.department, 
                c.location, 
                COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
                c.issue_raise_date AS date,
                c.status,
                c.resolution,
                c.updated_at
            FROM 
                complaint AS c
            JOIN 
                user AS u ON c.user_id = u.id
            LEFT JOIN 
                user AS u2 ON c.resolved_by_id = u2.id AND (u2.role = 'admin' OR u2.role = 'superadmin')
            WHERE 
                c.status = 'resolved' {date_filter}
            ORDER BY 
                c.issue_raise_date DESC;
        """
        cur.execute(query, params)
        resolution_reports = cur.fetchall()

        # Write data rows to the worksheet
        for row_num, report in enumerate(resolution_reports, 2):  # Start from row 2
            for col_num, value in enumerate(report, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)

    except Exception as e:
        print(f"Error: {e}")
    finally:
        cur.close()  # Ensure cursor is closed
        db.close()   # Ensure database connection is closed

    # Prepare the HTTP response to send the XLSX file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="resolution_reports.xlsx"'
    workbook.save(response)

    return response

def report(request):
    user_name = request.session.get('user_name')
    return render(request,'report.html',{'user_name' : user_name})

@never_cache
def download_filtered_resolution_reports_xlsx(request):
    # Get date range from request parameters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    date_filter = ""
    params = []

    if from_date and to_date:
        try:
            # Parse from_date and to_date
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d")
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(microseconds=1)

            # Create date filter with inclusive dates and add parameters
            date_filter = "AND c.issue_raise_date BETWEEN %s AND %s"
            params.extend([from_date_obj, to_date_obj])

        except ValueError:
            return HttpResponse("Invalid date format. Please use YYYY-MM-DD.", status=400)

    # Establish database connection
    cur, db = connection()

    try:
        # Fetch resolution reports with date filter
        query = f"""
            SELECT 
                u.name AS complainant_name, 
                c.department, 
                c.location, 
                COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
                c.issue_raise_date AS date,
                c.status,
                c.resolution,
                c.description,
                CASE 
                    WHEN c.status = 'resolved' THEN c.updated_at
                    ELSE NULL
                END AS resolved_at
            FROM 
                complaint AS c
            JOIN 
                user AS u ON c.user_id = u.id
            LEFT JOIN 
                user AS u2 ON c.resolved_by_id = u2.id AND (u2.role = 'admin' OR u2.role = 'superadmin')
            WHERE 
                (c.status IN ('resolved', 'work in progress', 'pending') OR c.complain_status = 'overdue') {date_filter}
            ORDER BY 
                c.issue_raise_date DESC;
        """
        
        cur.execute(query, params)
        resolution_reports = cur.fetchall()

        # Create an Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resolution Reports"

        # Define column headers
        headers = [
            'Complainant Name', 'Department', 'Location', 'Resolved By', 'Date', 
            'Status', 'Resolution', 'Description', 'Resolved At'
        ]
        ws.append(headers)

        # Add data to the sheet
        for report in resolution_reports:
            ws.append([
                report[0],  # complainant_name
                report[1],  # department
                report[2].capitalize(),  # location
                report[3],  # resolved_by_name
                report[4].strftime("%Y-%m-%d %H:%M:%S"),  # date
                report[5].capitalize(),  # status
                report[6],  # resolution
                report[7],  # description
                report[8].strftime("%Y-%m-%d %H:%M:%S") if report[8] else 'Not Resolved',  # resolved_at
            ])

        # Prepare the response to send the file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=resolution_reports.xlsx'
        
        # Save the workbook to the response
        wb.save(response)

        return response

    except Exception as e:
        print(f"Error: {e}")
        return HttpResponse(f"An error occurred: {e}", status=500)

    finally:
        cur.close()
        db.close()

# Add a new location
def location_list(request):
    user_name = request.session.get('user_name')
    cur, db = connection()

    try:
        if request.method == "POST":
            # Check if the request is an AJAX request (Content-Type: application/json)
            if request.headers.get('Content-Type') == 'application/json':
                data = json.loads(request.body)
                location_name = data.get('location_name')    
                # If the location name is provided
                if location_name:
                    # Insert the new location into the database
                    cur.execute(
                        "INSERT INTO location_master (location) VALUES (%s)",
                        (location_name,)
                    )
                    db.commit()

                    # Get the last inserted location's ID
                    cur.execute("SELECT LAST_INSERT_ID()")
                    location_id = cur.fetchone()[0]

                    # Return a JSON response with the new location's ID and name
                    response_data = {
                        'success': True,
                        'location': {'id': location_id, 'location': location_name}
                    }
                    return JsonResponse(response_data)

                # If no location name was provided, return an error response
                return JsonResponse({'success': False, 'message': 'Location name is required'}, status=400)

        # Fetch all locations to display in the template
        cur.execute("""
        SELECT MIN(id) AS id, location
        FROM location_master
        GROUP BY location
        """)
        locations = cur.fetchall()
        location_data = [{'id': loc[0], 'location': loc[1]} for loc in locations]

    except Exception as e:
        db.rollback()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

    finally:
        cur.close()
        db.close()

    # Render the page with the list of locations (for initial page load)
    return render(request, 'location_list.html', {
        'user_name': user_name,
        'locations': location_data
    })

# Edit an existing location
def edit_location(request, location_id):
    if request.method == "POST":
        # Retrieve JSON data
        data = json.loads(request.body)
        new_location_name = data.get('location_name')

        if new_location_name:
            cur, db = connection()  # Establish connection to the database
            try:
                # Update location in the database
                cur.execute(
                    "UPDATE location_master SET location = %s WHERE id = %s",
                    (new_location_name, location_id)
                )
                db.commit()

                # Check if any row was updated
                if cur.rowcount > 0:
                    return JsonResponse({'success': True, 'message': 'Location updated successfully'})
                else:
                    return JsonResponse({'success': False, 'message': 'Location not found or not modified'}, status=400)

            except Exception as e:
                db.rollback()
                return JsonResponse({'success': False, 'message': str(e)}, status=500)

            finally:
                cur.close()
                db.close()

        else:
            return JsonResponse({'success': False, 'message': 'Location name is required'}, status=400)

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

# Delete a location
def delete_location(request, location_id):
    if request.method == "POST":
        cur, db = connection()

        try:
            # Delete location from the database
            cur.execute("DELETE FROM location_master WHERE id = %s", (location_id,))
            db.commit()

            # Check if the deletion was successful
            if cur.rowcount > 0:
                return JsonResponse({"success": True})
            else:
                return JsonResponse({"success": False, "message": "Location not found"}, status=400)

        except Exception as e:
            db.rollback()
            return JsonResponse({"success": False, "message": str(e)}, status=500)

        finally:
            cur.close()
            db.close()

    return JsonResponse({"success": False, "message": "Invalid request method"}, status=400)

def department_list(request):
    user_name = request.session.get('user_name')
    cur, db = connection()

    try:
        if request.method == "POST":
            # Check if the request is an AJAX request (Content-Type: application/json)
            if request.headers.get('Content-Type') == 'application/json':
                data = json.loads(request.body)
                department_name = data.get('department_name')

                # If the department name is provided
                if department_name:
                    # Insert the new department into the database
                    cur.execute(
                        "INSERT INTO department_master (department) VALUES (%s)",
                        (department_name,)
                    )
                    db.commit()

                    # Get the last inserted department's ID
                    cur.execute("SELECT LAST_INSERT_ID()")
                    department_id = cur.fetchone()[0]

                    # Return a JSON response with the new department's ID and name
                    response_data = {
                        'success': True,
                        'department': {'id': department_id, 'department': department_name}
                    }
                    return JsonResponse(response_data)

                # If no department name was provided, return an error response
                return JsonResponse({'success': False, 'message': 'Department name is required'}, status=400)

        # Fetch all departments to display in the template (unique departments)
        cur.execute("""
        SELECT MIN(id) AS id, department
        FROM department_master
        GROUP BY department
        """)
        departments = cur.fetchall()
        department_data = [{'id': dept[0], 'department': dept[1]} for dept in departments]

    except Exception as e:
        db.rollback()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

    finally:
        cur.close()
        db.close()

    # Render the page with the list of departments (for initial page load)
    return render(request, 'department_list.html', {
        'user_name': user_name,
        'departments': department_data
    })

# Edit an existing department
def edit_department(request, department_id):
    if request.method == "POST":
        # Retrieve JSON data
        data = json.loads(request.body)
        new_department_name = data.get('department_name')

        if new_department_name:
            cur, db = connection()  # Establish connection to the database
            try:
                # Update department in the database
                cur.execute(
                    "UPDATE department_master SET department = %s WHERE id = %s",
                    (new_department_name, department_id)
                )
                db.commit()

                # Check if any row was updated
                if cur.rowcount > 0:
                    return JsonResponse({'success': True, 'message': 'Department updated successfully'})
                else:
                    return JsonResponse({'success': False, 'message': 'Department not found or not modified'}, status=400)

            except Exception as e:
                db.rollback()
                return JsonResponse({'success': False, 'message': str(e)}, status=500)

            finally:
                cur.close()
                db.close()

        else:
            return JsonResponse({'success': False, 'message': 'Department name is required'}, status=400)

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

# Delete a department
def delete_department(request, department_id):
    if request.method == "POST":
        cur, db = connection()

        try:
            # Delete department from the database
            cur.execute("DELETE FROM department_master WHERE id = %s", (department_id,))
            db.commit()

            # Check if the deletion was successful
            if cur.rowcount > 0:
                return JsonResponse({"success": True})
            else:
                return JsonResponse({"success": False, "message": "Department not found"}, status=400)

        except Exception as e:
            db.rollback()
            return JsonResponse({"success": False, "message": str(e)}, status=500)

        finally:
            cur.close()
            db.close()

    return JsonResponse({"success": False, "message": "Invalid request method"}, status=400)

def status_list(request):
    user_name = request.session.get('user_name')
    cur, db = connection()

    try:
        if request.method == "POST":
            # Check if the request is an AJAX request (Content-Type: application/json)
            if request.headers.get('Content-Type') == 'application/json':
                data = json.loads(request.body)
                status_name = data.get('status_name')

                # If the status name is provided
                if status_name:
                    # Insert the new status into the database
                    cur.execute(
                        "INSERT INTO status_master (status) VALUES (%s)",
                        (status_name,)
                    )
                    db.commit()

                    # Get the last inserted status's ID
                    cur.execute("SELECT LAST_INSERT_ID()")
                    status_id = cur.fetchone()[0]

                    # Return a JSON response with the new status's ID and name
                    response_data = {
                        'success': True,
                        'status': {'id': status_id, 'status': status_name}
                    }
                    return JsonResponse(response_data)

                # If no status name was provided, return an error response
                return JsonResponse({'success': False, 'message': 'Status name is required'}, status=400)

        # Fetch all statuses to display in the template (unique statuses)
        cur.execute("""
        SELECT id, status
        FROM status_master
        """)
        statuses = cur.fetchall()
        status_data = [{'id': stat[0], 'status': stat[1]} for stat in statuses]

    except Exception as e:
        db.rollback()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

    finally:
        cur.close()
        db.close()

    # Render the page with the list of statuses (for initial page load)
    return render(request, 'status_list.html', {
        'user_name': user_name,
        'statuses': status_data
    })

def edit_status(request, status_id):
    if request.method == "POST":
        # Retrieve JSON data
        data = json.loads(request.body)
        new_status_name = data.get('status_name')

        if new_status_name:
            cur, db = connection()  # Establish connection to the database
            try:
                # Update status in the database
                cur.execute(
                    "UPDATE status_master SET status = %s WHERE id = %s",
                    (new_status_name, status_id)
                )
                db.commit()

                # Check if any row was updated
                if cur.rowcount > 0:
                    return JsonResponse({'success': True, 'message': 'Status updated successfully'})
                else:
                    return JsonResponse({'success': False, 'message': 'Status not found or not modified'}, status=400)

            except Exception as e:
                db.rollback()
                return JsonResponse({'success': False, 'message': str(e)}, status=500)

            finally:
                cur.close()
                db.close()

        else:
            return JsonResponse({'success': False, 'message': 'Status name is required'}, status=400)

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

def delete_status(request, status_id):
    if request.method == "POST":
        cur, db = connection()

        try:
            # Delete status from the database
            cur.execute("DELETE FROM status_master WHERE id = %s", (status_id,))
            db.commit()

            # Check if the deletion was successful
            if cur.rowcount > 0:
                return JsonResponse({"success": True})
            else:
                return JsonResponse({"success": False, "message": "Status not found"}, status=400)

        except Exception as e:
            db.rollback()
            return JsonResponse({"success": False, "message": str(e)}, status=500)

        finally:
            cur.close()
            db.close()

    return JsonResponse({"success": False, "message": "Invalid request method"}, status=400)

def complaint_type_list(request):
    user_name = request.session.get('user_name')
    cur, db = connection()

    try:
        if request.method == "POST":
            if request.headers.get('Content-Type') == 'application/json':
                data = json.loads(request.body)
                complaint_type_name = data.get('complaint_type_name')
                days = data.get('TAT_days', 0)
                hours = data.get('TAT_hours', '0')  # Get the hours input as a string
                department = data.get('department')

                # Handle the case where hours is a time string like '17:00:00'
                if isinstance(hours, str) and ':' in hours:
                    hour_part = hours.split(':')[0]  # Extract hours from 'HH:MM:SS'
                    hours = int(hour_part)  # Convert to integer hours

                # Calculate total TAT hours
                total_tat_hours = (int(days) * 24) + hours
                print(total_tat_hours)

                if complaint_type_name and department:
                    # Insert into the database with calculated total_tat_hours
                    cur.execute(
                        "INSERT INTO complaint_type_master (complaint_type, TAT, department) VALUES (%s, %s, %s)",
                        (complaint_type_name, total_tat_hours, department)
                    )
                    db.commit()

                    # Get the last inserted complaint type's ID
                    cur.execute("SELECT LAST_INSERT_ID()")
                    complaint_type_id = cur.fetchone()[0]

                    # Return response with the new complaint type details
                    response_data = {
                        'success': True,
                        'complaint_type': {
                            'id': complaint_type_id,
                            'complaint_type': complaint_type_name,
                            'TAT': total_tat_hours,
                            'department': department
                        }
                    }
                    return JsonResponse(response_data)

                return JsonResponse({'success': False, 'message': 'Complaint type name and department are required'}, status=400)

        # Fetch complaint types from the database
        cur.execute("""
            SELECT MIN(id) AS id, complaint_type, TAT, department
            FROM complaint_type_master
            GROUP BY complaint_type, department
        """)
        complaint_types = cur.fetchall()

        complaint_type_data = []
        for ctype in complaint_types:
            total_tat_hours = ctype[2]
            
            # Check if total_tat_hours is None, then set to 0 if it is
            if total_tat_hours is None:
                total_tat_hours = 0

            # Convert total TAT hours to days and hours
            days = total_tat_hours // 24
            hours = total_tat_hours % 24

            complaint_type_data.append({
                'id': ctype[0],
                'complaint_type': ctype[1],
                'TAT_days': days,
                'TAT_hours': hours,
                'department': ctype[3]
            })

    except Exception as e:
        db.rollback()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

    finally:
        cur.close()
        db.close()

    # Render the page with the list of complaint types
    return render(request, 'complaint_type_list.html', {
        'user_name': user_name,
        'complaint_types': complaint_type_data
    })

@csrf_exempt
def edit_complaint_type(request, complaint_type_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            print(f"Received data: {data}")  # Debugging log
            
            complaint_type_name = data.get('complaint_type_name')
            tat_days = int(data.get('TAT_days', 0))  # Get the days from the input
            tat_hours = data.get('TAT_hours', '0')  # Get the hours input in HH:MM format
            department = data.get('department')

            # Just split the hours in HH:MM format but don't perform any calculation
            try:
                hours, minutes = map(int, tat_hours.split(':'))  # Extract hours and minutes
            except ValueError:
                raise ValueError(f"Invalid time format for TAT_hours: {tat_hours}. Expected format is HH:MM.")
            
            # If you only care about hours and want to ignore minutes, you can just use:
            hours = hours  # No need to store or calculate minutes if they're not needed
            
            # Calculate the total TAT hours in the same way as in the complaint_type_list view
            total_tat_hours = (tat_days * 24) + hours  # Convert days to hours and add the hours provided by the user
            print(f"Total TAT hours: {total_tat_hours}")
            
            # Connect to DB
            cur, db = connection()

            # Update the record by only changing the days and hours (no TAT calculation, but setting the calculated TAT)
            query = """
            UPDATE complaint_type_master
            SET complaint_type = %s, department = %s, TAT = %s
            WHERE id = %s
            """

            # Execute the query to update the complaint type with the calculated TAT in hours
            cur.execute(query, (complaint_type_name, department, total_tat_hours, complaint_type_id))
            db.commit()

            # Directly return the updated TAT as provided by the user (without any conversion)
            return JsonResponse({
                'success': True,
                'message': 'Complaint type updated successfully',
                'complaint_type': {
                    'id': complaint_type_id,
                    'complaint_type': complaint_type_name,
                    'TAT_days': tat_days,  # Return the days as provided by the user
                    'TAT_hours': hours,  # Return the hours as provided by the user
                    'department': department
                }
            })

        except Exception as e:
            print(f"Error: {e}")  # Debugging log
            try:
                db.rollback()  # Ensure rollback happens if DB transaction fails
            except:
                pass  # In case db is not initialized
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
        finally:
            try:
                cur.close()
                db.close()
            except:
                pass  # In case cur/db was never initialized

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

def delete_complaint_type(request, complaint_type_id):
    if request.method == "POST":
        cur, db = connection()


        try:
            # Delete complaint type from the database
            cur.execute("DELETE FROM complaint_type_master WHERE id = %s", (complaint_type_id,))
            db.commit()


            # Check if the deletion was successful
            if cur.rowcount > 0:
                return JsonResponse({"success": True})
            else:
                return JsonResponse({"success": False, "message": "Complaint type not found"}, status=400)


        except Exception as e:
            db.rollback()
            return JsonResponse({"success": False, "message": str(e)}, status=500)


        finally:
            cur.close()
            db.close()

    return JsonResponse({"success": False, "message": "Invalid request method"}, status=400)

# def complaint_type_list(request):
#     user_name = request.session.get('user_name')
#     cur, db = connection()

#     try:
#         if request.method == "POST":
#             # Check if the request is an AJAX request (Content-Type: application/json)
#             if request.headers.get('Content-Type') == 'application/json':
#                 data = json.loads(request.body)
#                 complaint_type_name = data.get('complaint_type_name')
#                 days = int(data.get('days', 0))  # Default days to 0 if not provided
#                 hours = int(data.get('hours', 0))  # Default hours to 0 if not provided
#                 department = data.get('department')  # Capture department

#                 # Debugging: Print the received values for days, hours, and complaint_type_name
#                 print(f"Received data: complaint_type_name={complaint_type_name}, days={days}, hours={hours}, department={department}")

#                 # Validate required fields
#                 if not complaint_type_name or not department:
#                     return JsonResponse({'success': False, 'message': 'Complaint type name and department are required'}, status=400)

#                 # Calculate TAT in hours
#                 tat = (days * 24) + hours

#                 # Debugging: Print the calculated TAT value
#                 print(f"Calculated TAT: {tat} hours")

#                 # Insert the new complaint type, TAT, and department into the database
#                 cur.execute(
#                     "INSERT INTO complaint_type_master (complaint_type, TAT, department) VALUES (%s, %s, %s)",
#                     (complaint_type_name, tat, department)
#                 )
#                 db.commit()

#                 # Get the last inserted complaint type's ID
#                 cur.execute("SELECT LAST_INSERT_ID()")
#                 complaint_type_id = cur.fetchone()[0]

#                 # Return a JSON response with the new complaint type's details
#                 response_data = {
#                     'success': True,
#                     'complaint_type': {
#                         'id': complaint_type_id,
#                         'complaint_type': complaint_type_name,
#                         'TAT': tat,
#                         'department': department
#                     }
#                 }
#                 return JsonResponse(response_data)

#         # Fetch all unique complaint types from the database
#         cur.execute("""
#         SELECT MIN(id) AS id, complaint_type, TAT, department
#         FROM complaint_type_master
#         GROUP BY complaint_type, department
#         """)
#         complaint_types = cur.fetchall()
#         complaint_type_data = [
#             {'id': ctype[0], 'complaint_type': ctype[1], 'TAT': ctype[2], 'department': ctype[3]} 
#             for ctype in complaint_types
#         ]

#     except Exception as e:
#         db.rollback()
#         return JsonResponse({'success': False, 'message': str(e)}, status=500)

#     finally:
#         cur.close()
#         db.close()

#     # Render the page with the list of complaint types (for initial page load)
#     return render(request, 'complaint_type_list.html', {
#         'user_name': user_name,
#         'complaint_types': complaint_type_data
#     })


# @csrf_exempt
# def edit_complaint_type(request, complaint_type_id):
#     if request.method == 'POST':
#         try:
#             data = json.loads(request.body)
#             complaint_type_name = data.get('complaint_type_name')
#             days = int(data.get('days', 0))  # Default to 0 if not provided
#             hours = int(data.get('hours', 0))  # Default to 0 if not provided
#             department = data.get('department')  # Capture department

#             # Debugging: Print the received values for days, hours, and complaint_type_name
#             print(f"Received data: complaint_type_name={complaint_type_name}, days={days}, hours={hours}, department={department}")

#             # Validate required fields
#             if not complaint_type_name or not department:
#                 return JsonResponse({'success': False, 'message': 'Complaint type name and department are required'}, status=400)

#             # Calculate TAT in hours
#             tat = (days * 24) + hours

#             # Debugging: Print the calculated TAT value
#             print(f"Calculated TAT: {tat} hours")

#             cur, db = connection()

#             # Update complaint type in the database
#             query = """
#             UPDATE complaint_type_master
#             SET complaint_type = %s, TAT = %s, department = %s
#             WHERE id = %s
#             """
#             cur.execute(query, (complaint_type_name, tat, department, complaint_type_id))
#             db.commit()

#             return JsonResponse({'success': True, 'message': 'Complaint type updated successfully'})

#         except Exception as e:
#             db.rollback()
#             return JsonResponse({'success': False, 'message': str(e)}, status=500)

#         finally:
#             cur.close()
#             db.close()

#     return JsonResponse({"success": False, "message": "Invalid request method"}, status=400)

# @csrf_exempt
# def delete_complaint_type(request, complaint_type_id):
#     if request.method == "POST":
#         cur, db = connection()

#         try:
#             # Delete complaint type from the database
#             cur.execute("DELETE FROM complaint_type_master WHERE id = %s", (complaint_type_id,))
#             db.commit()

#             # Check if the deletion was successful
#             if cur.rowcount > 0:
#                 return JsonResponse({"success": True})
#             else:
#                 return JsonResponse({"success": False, "message": "Complaint type not found"}, status=400)

#         except Exception as e:
#             db.rollback()
#             return JsonResponse({"success": False, "message": str(e)}, status=500)

#         finally:
#             cur.close()
#             db.close()

#     return JsonResponse({"success": False, "message": "Invalid request method"}, status=400)

def filtered_user_reports(request):
    if 'user_id' not in request.session:
        return redirect('login')

    user_id = request.session.get('user_id')
    cur, db = connection()

    # Fetch user complaints
    cur.execute("SELECT * FROM complaint WHERE user_id = %s", [user_id])
    complaints = cur.fetchall()

    # Fetch logged-in user's details
    cur.execute("SELECT * FROM user WHERE id = %s", [user_id])
    user = cur.fetchone()

    user_reports = []
    if user:
        for complaint in complaints:
            issue_raise_date = complaint[7]
            updated_at = complaint[11] if complaint[13].lower() == "resolved" else None

            # Convert date strings to datetime objects if needed
            if isinstance(issue_raise_date, str):
                issue_raise_date = datetime.strptime(issue_raise_date, "%Y-%m-%d %H:%M:%S")
            if updated_at and isinstance(updated_at, str):
                updated_at = datetime.strptime(updated_at, "%Y-%m-%d %H:%M:%S")

            resolution_time = (
                format_resolution_time(updated_at - issue_raise_date) if updated_at else "Not Resolved Yet"
            )

            user_reports.append({
                'id': user_id,
                'name': user[1],  
                'department': complaint[3], 
                'location': complaint[4].capitalize(), 
                'complaint_type': complaint[5], 
                'status': complaint[13].capitalize(), 
                'resolution': complaint[14] if complaint[13].lower() == 'resolved' else "No Resolution Yet",
                'issue_raise_date': issue_raise_date.strftime("%Y-%m-%d %H:%M:%S"),
                'due_date': complaint[9].strftime("%Y-%m-%d %H:%M:%S"),
                'description': complaint[6],
                'updated_at': updated_at.strftime("%Y-%m-%d %H:%M:%S") if updated_at else "Not Resolved Yet",
                'resolution_time': resolution_time 
            })

    # Apply filtering based on query parameters
    status_filter = request.GET.get('status')
    department_filter = request.GET.get('department')
    complaint_type_filter = request.GET.get('complaint_type')
    location_filter = request.GET.get('location_type')
    from_date_filter = request.GET.get('from_date')
    to_date_filter = request.GET.get('to_date')

    # Get distinct status values for the specified user
    cur.execute("SELECT DISTINCT status FROM complaint WHERE user_id = %s", [user_id])
    statuses = [row[0] for row in cur.fetchall()]

    # Get distinct department values for the specified user
    cur.execute("SELECT DISTINCT department FROM complaint WHERE user_id = %s", [user_id])
    departments = [row[0] for row in cur.fetchall()]

    # Get distinct location values for the specified user
    cur.execute("SELECT DISTINCT location FROM complaint WHERE user_id = %s", [user_id])
    locations = [row[0] for row in cur.fetchall()]

    # Get distinct complaint type values for the specified user
    cur.execute("SELECT DISTINCT complaint_type FROM complaint WHERE user_id = %s", [user_id])
    complaint_types = [row[0] for row in cur.fetchall()]

    # Apply date filters if provided (full day range)
    if from_date_filter:
        from_date = datetime.strptime(from_date_filter, "%Y-%m-%d")
        from_date = from_date.replace(hour=0, minute=0, second=0, microsecond=0)  # Start of the day
        user_reports = [report for report in user_reports if datetime.strptime(report['issue_raise_date'], "%Y-%m-%d %H:%M:%S") >= from_date]

    if to_date_filter:
        to_date = datetime.strptime(to_date_filter, "%Y-%m-%d")
        to_date = to_date.replace(hour=23, minute=59, second=59, microsecond=999999)  # End of the day
        user_reports = [report for report in user_reports if datetime.strptime(report['issue_raise_date'], "%Y-%m-%d %H:%M:%S") <= to_date]

    # Apply other filters
    if status_filter:
        user_reports = [report for report in user_reports if report['status'].lower() == status_filter.lower()]

    # Apply department filter
    if department_filter:
        if department_filter in departments:
            user_reports = [report for report in user_reports if report['department'] == department_filter]

    # Apply complaint type filter
    if complaint_type_filter:
        if complaint_type_filter in complaint_types:
            user_reports = [report for report in user_reports if report['complaint_type'] == complaint_type_filter]

    # Apply location filter
    if location_filter:
        if location_filter in locations:
            user_reports = [report for report in user_reports if report['location'] == location_filter]

    # Check if download is requested as an Excel file
    if 'download' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="user_reports_{user_id}.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "User Reports"

        # Add headers with the desired order
        headers = [
            'ID', 'Name', 'Dept', 'Location', 'Type', 'Description', 
            'Issue Date', 'Due Date', 'Status', 'Resolution', 'Resolved At', 'Resolution Time'
        ]

        # Define header styles
        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Grey background
        header_font = Font(bold=True, color="000000", size=12)  # Black text, bold, size 12
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Apply header styles and set column headers
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[f'{col_letter}1']
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add data rows
        for row_num, report in enumerate(user_reports, 2):
            sheet[f'A{row_num}'] = report['id']
            sheet[f'B{row_num}'] = report['name']
            sheet[f'C{row_num}'] = report['department']
            sheet[f'D{row_num}'] = report['location']
            sheet[f'E{row_num}'] = report['complaint_type']
            sheet[f'F{row_num}'] = report['description']
            sheet[f'G{row_num}'] = report['issue_raise_date']
            sheet[f'H{row_num}'] = report['due_date']
            sheet[f'I{row_num}'] = report['status']
            sheet[f'J{row_num}'] = report['resolution']
            sheet[f'K{row_num}'] = report['updated_at']
            sheet[f'L{row_num}'] = report['resolution_time']

            # Apply text wrapping for Description and Resolution columns
            for col in ['F', 'J']:
                cell = sheet[f'{col}{row_num}']
                cell.alignment = Alignment(wrap_text=True)

        # Set fixed widths for Description and Resolution columns (F and J) to prevent stretching
        sheet.column_dimensions['F'].width = 20  # Set width for Description column
        sheet.column_dimensions['J'].width = 20  # Set width for Resolution column
        sheet.column_dimensions['G'].width = 18  # Set width for Issue Date column
        sheet.column_dimensions['H'].width = 18  # Set width for Due Date column
        sheet.column_dimensions['I'].width = 16  # Set width for Status column
        sheet.column_dimensions['L'].width = 16.50  # Set width for Resolution Time column
        sheet.column_dimensions['E'].width = 20  # Set width for Type column

        # Adjust other column widths for better readability
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 2  # Add extra padding
            if col_letter not in ['F','J','G','H','I','L','E']:  # Skip the fixed width columns
                sheet.column_dimensions[col_letter].width = adjusted_width

        # Save the workbook to the response
        workbook.save(response)
        return response

    # Pass the filtered user reports to the template
    context = {
        'user_reports': user_reports,
        'user_name': request.session.get('user_name'),
        'user_role': request.session.get('user_role'),
        'departments': departments,
        'locations': locations,
        'complaint_types': complaint_types,
        'statuses': statuses
    }

    return render(request, 'user_reports.html', context)

def filtered_resolver_reports(request):
    # Get the logged-in user's name and role from the session (optional)
    # Get the logged-in user's ID (assuming it's stored in the session)
    user_id = request.session.get('user_id')
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')

    # Initialize query parameters
    filters = []
    query_conditions = []

    # Check for filters in the GET request
    location_filter = request.GET.get('location')
    complaint_type_filter = request.GET.get('complaint_type')
    department_filter = request.GET.get('department')
    from_date_filter = request.GET.get('from_date')
    to_date_filter = request.GET.get('to_date')
    status_filter = request.GET.get('status')

    # Establish the connection and get the cursor
    cur, db = connection()  # Ensure `connection()` is correctly defined in your project
    
# Fetch resolver's location and department from the database
    cur.execute("SELECT location, department FROM user WHERE id = %s", [user_id])
    resolver_info = cur.fetchone()
    if resolver_info:
        resolver_location = resolver_info[0]  # Extract the location value
        resolver_department = resolver_info[1]  # Extract the department value
    else:
        resolver_location = None  # Default to None if no location found
        resolver_department = None  # Default to None if no department found

    # Build the SQL query dynamically based on filters
    query = """
    SELECT 
        c.id AS complaint_id, 
        c.name AS user_name, 
        c.department AS user_department, 
        c.complaint_type, 
        c.description, 
        c.location, 
        c.issue_raise_date, 
        c.status,
        c.resolution,
        c.updated_at,
        c.due_date 
    FROM 
        complaint c 
    JOIN 
        user u ON c.user_id = u.id
    WHERE
        c.assigned_to = %s  -- Filter for complaints assigned to the logged-in resolver
    """
    
    filters = [user_id]  # Add the resolver's user_id as the first filter
    
    # Apply resolver's location as a filter
    if resolver_location:
        query_conditions.append("c.location = %s")
        filters.append(resolver_location)
        
    if resolver_department:
        query_conditions.append("c.department = %s")
    filters.append(resolver_department)

    # Apply filters to the query
    if location_filter:
        query_conditions.append("c.location = %s")
        filters.append(location_filter)
    
    if complaint_type_filter:
        query_conditions.append("c.complaint_type = %s")
        filters.append(complaint_type_filter)

    if department_filter:
        query_conditions.append("c.department = %s")
        filters.append(department_filter)

    if from_date_filter:
        # Assuming from_date is inclusive, and the time is set to '00:00:00' for full date inclusion
        from_date = datetime.strptime(from_date_filter, '%Y-%m-%d')
        query_conditions.append("c.issue_raise_date >= %s")
        filters.append(from_date)

    if to_date_filter:
        # Assuming to_date is inclusive, and the time is set to '23:59:59' for full date inclusion
        to_date = datetime.strptime(to_date_filter, '%Y-%m-%d')
        to_date = to_date.replace(hour=23, minute=59, second=59)
        query_conditions.append("c.issue_raise_date <= %s")
        filters.append(to_date)

    if status_filter:
        query_conditions.append("c.status = %s")
        filters.append(status_filter)

    # Combine the conditions to the base query
    if query_conditions:
        query += " AND " + " AND ".join(query_conditions)
    
    # Order the result by issue raise date in descending order
    query += " ORDER BY c.issue_raise_date DESC;"

    # Execute the query with filters
    cur.execute(query, filters)
    filtered_data = cur.fetchall()  # Fetch all results

    # Process the fetched data if needed (e.g., formatting date)
    formatted_data = []
    for complaint in filtered_data:
        issue_raise_date = complaint[6]
        updated_at = complaint[9] if complaint[7].lower() == "resolved" else None

        # Convert date strings to datetime objects if needed
        if isinstance(issue_raise_date, str):
            issue_raise_date = datetime.strptime(issue_raise_date, "%Y-%m-%d %H:%M:%S")
        if updated_at and isinstance(updated_at, str):
            updated_at = datetime.strptime(updated_at, "%Y-%m-%d %H:%M:%S")

        resolution_time = (
            format_resolution_time(updated_at - issue_raise_date) if updated_at else "Not Resolved Yet"
        )

        formatted_data.append({
            'complaint_id': complaint[0],
            'user_name': complaint[1],  
            'user_department': complaint[2],
            'complaint_type': complaint[3],
            'description': complaint[4],
            'location': complaint[5].capitalize(),
            'issue_raise_date': issue_raise_date.strftime("%Y-%m-%d %H:%M:%S"),
            'status': complaint[7].capitalize(),
            'resolution': complaint[8] if complaint[7].lower() == 'resolved' else "No Resolution Yet",
            'updated_at': updated_at.strftime("%Y-%m-%d %H:%M:%S") if updated_at else "Not Resolved Yet",
            'resolution_time': resolution_time
        })

    # Fetch distinct location, department, and complaint type for filter options
    location_query = "SELECT DISTINCT location FROM location_master"
    department_query = "SELECT DISTINCT department FROM department_master"
    complaint_type_query = "SELECT DISTINCT complaint_type FROM complaint_type_master"
    status_query = "SELECT DISTINCT status FROM status_master"

    cur.execute(location_query)
    locations = cur.fetchall()

    cur.execute(department_query)
    departments = cur.fetchall()

    cur.execute(complaint_type_query)
    complaint_types = cur.fetchall()
    
    cur.execute(status_query)
    statuses = cur.fetchall()

    # Close the database connection
    db.close()

    # Check if download is requested as an Excel file
    if 'download' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="filtered_resolver_reports_{user_name}.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Resolver Reports"

        # Add headers with bold, black background, and size 12
        headers = [
            'ID', 'Name', 'Dept', 'Type', 'Description', 'Location', 
            'Issue Date', 'Status', 'Resolution', 'Resolved At', 'Resolution Time'
        ]

        # Define header styles
        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Grey background
        header_font = Font(bold=True, color="000000", size=12)  # Black text, bold, size 12
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Apply header styles and set column headers
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[f'{col_letter}1']
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add data rows
        for row_num, report in enumerate(formatted_data, 2):
            sheet[f'A{row_num}'] = report['complaint_id']
            sheet[f'B{row_num}'] = report['user_name']
            sheet[f'C{row_num}'] = report['user_department']
            sheet[f'D{row_num}'] = report['complaint_type']
            sheet[f'E{row_num}'] = report['description']
            sheet[f'F{row_num}'] = report['location']
            sheet[f'G{row_num}'] = report['issue_raise_date']
            sheet[f'H{row_num}'] = report['status']
            sheet[f'I{row_num}'] = report['resolution']
            sheet[f'J{row_num}'] = report['updated_at']
            sheet[f'K{row_num}'] = report['resolution_time']

            # Apply text wrapping (alt+enter) for Description and Resolution columns
            for col in ['E', 'I']:
                cell = sheet[f'{col}{row_num}']
                cell.alignment = Alignment(wrap_text=True)

        # Set fixed widths for description and resolution columns (E and I)
        sheet.column_dimensions['E'].width = 25  # Set width for Description column (adjust as needed)
        sheet.column_dimensions['I'].width = 20  # Set width for Resolution column (adjust as needed)

        # Adjust other column widths for better readability
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 2  # Add extra padding
            if col_letter not in ['E', 'I']:  # Skip the fixed width columns
                sheet.column_dimensions[col_letter].width = adjusted_width

        # Save the workbook to the response
        workbook.save(response)
        return response


    # Pass the filtered data and filter options to the template
    context = {
        'user_name': user_name,
        'user_role': user_role,
        'user_id':user_id,
        'filtered_data': formatted_data,
        'locations': locations,
        'resolver_location': resolver_location,
        'departments': departments,
        'complaint_types': complaint_types,
        'statuses': statuses,
        'status_options': ['Pending', 'Work in Progress', 'Resolved', 'Overdue'],
    }

    return render(request, 'resolver_reports.html', context)

def pending_complaints_view(request):   # Template to show just table information for status = pending in superadmin dashboard
    # Connect to the database
    cur, db = connection()
    
    try:
        # Fetch pending complaints
        query = """
            SELECT 
                u.name AS complainant_name, 
                u.department, 
                u.location, 
                COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
                c.issue_raise_date AS date,
                c.due_date,
                c.status,
                c.resolution,
                c.description,
                c.complaint_type
            FROM 
                complaint AS c
            JOIN 
                user AS u ON c.user_id = u.id
            LEFT JOIN 
                user AS u2 ON c.resolved_by_id = u2.id AND (u2.role = 'admin' OR u2.role = 'superadmin')
            WHERE 
                c.status = 'pending'
            ORDER BY 
                c.issue_raise_date DESC;
        """
        cur.execute(query)
        pending_complaints = cur.fetchall()
        
        # Prepare data for template rendering
        columns = [col[0] for col in cur.description]  # Extract column names
        pending_complaints = [dict(zip(columns, row)) for row in pending_complaints]
        
    except Exception as e:
        pending_complaints = []
        print(f"Error fetching pending complaints: {e}")
    
    finally:
        cur.close()
        db.close()
    
    # Render the pending complaints in a template
    return render(request, 'pending_complaints.html', {'pending_complaints': pending_complaints})

def in_progress_complaints_view(request):  # Template to show just table information for status = 'work in progress' in superadmin dashboard
    # Connect to the database
    cur, db = connection()
    
    try:
        # Fetch complaints in progress
        query = """
            SELECT 
                u.name AS complainant_name, 
                u.department, 
                u.location, 
                COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
                c.issue_raise_date AS date,
                c.due_date,
                c.status,
                c.resolution,
                c.description,
                c.complaint_type
            FROM 
                complaint AS c
            JOIN 
                user AS u ON c.user_id = u.id
            LEFT JOIN 
                user AS u2 ON c.resolved_by_id = u2.id AND (u2.role = 'admin' OR u2.role = 'superadmin')
            WHERE 
                c.status = 'work in progress'
            ORDER BY 
                c.issue_raise_date DESC;
        """
        cur.execute(query)
        in_progress_complaints = cur.fetchall()
        
        # Prepare data for template rendering
        columns = [col[0] for col in cur.description]  # Extract column names
        in_progress_complaints = [dict(zip(columns, row)) for row in in_progress_complaints]
        
    except Exception as e:
        in_progress_complaints = []
        print(f"Error fetching in-progress complaints: {e}")
    
    finally:
        cur.close()
        db.close()
    
    # Render the in-progress complaints in a template
    return render(request, 'in_progress_complaints.html', {'in_progress_complaints': in_progress_complaints})

def resolved_overtime_complaints_view(request):  # Template to show just table information for status = 'resolved_overtime' in superadmin dashboard
    # Connect to the database
    cur, db = connection()

    try:
        # Fetch complaints resolved overtime
        query = """
            SELECT 
                u.name AS complainant_name, 
                u.department, 
                u.location, 
                COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
                c.issue_raise_date AS date,
                c.due_date,
                c.status,
                c.resolution,
                c.description,
                c.complaint_type,
                CASE 
                    WHEN c.status = 'resolved' THEN c.updated_at
                    ELSE NULL
                END AS resolved_at
            FROM 
                complaint AS c
            JOIN 
                user AS u ON c.user_id = u.id
            LEFT JOIN 
                user AS u2 ON c.resolved_by_id = u2.id AND (u2.role = 'admin' OR u2.role = 'superadmin')
            WHERE 
                c.status = 'resolved' AND c.complain_status = 'overdue'
            ORDER BY 
                c.issue_raise_date DESC;
        """
        cur.execute(query)
        resolved_overtime_complaints = cur.fetchall()

        # Prepare data for template rendering with TAT calculation
        columns = [col[0] for col in cur.description]  # Extract column names
        complaints_data = []

        for row in resolved_overtime_complaints:
            # Extract the data for each complaint
            complainant_name = row[0]
            department = row[1]
            location = row[2].capitalize()
            resolved_by = row[3]
            date = row[4]
            due_date = row[5]
            status = row[6].capitalize()
            resolution = row[7]
            description = row[8]
            complaint_type = row[9]
            resolved_at = row[10]

            # Calculate TAT if resolved_at exists
            if resolved_at:
                tat_duration = resolved_at - date
            else:
                tat_duration = datetime.now() - date

            total_seconds = tat_duration.total_seconds()
            days = int(total_seconds // 86400)
            hours = int((total_seconds % 86400) // 3600)
            minutes = int((total_seconds % 3600) // 60)
            seconds = int(total_seconds % 60)

            # Format TAT based on the conditions
            if days > 0:
                tat_formatted = f"{days}d {hours}h {minutes}m {seconds}s"
            elif hours > 0:
                tat_formatted = f"{hours}h {minutes}m {seconds}s"
            else:
                tat_formatted = f"{minutes}m {seconds}s"

            complaints_data.append({
                'name': complainant_name,
                'department': department,
                'location': location,
                'resolved_by': resolved_by,
                'date': date.strftime("%Y-%m-%d %H:%M:%S"),
                'due_date': due_date.strftime("%Y-%m-%d %H:%M:%S"),
                'status': status,
                'resolution': resolution,
                'description': description,
                'complaint_type': complaint_type,
                'resolved_at': resolved_at.strftime("%Y-%m-%d %H:%M:%S") if resolved_at else 'None',
                'tat': tat_formatted if resolved_at else 'None'  # Add TAT to the report data
            })

        # Pass the complaint data to the template
        resolved_overtime_complaints = complaints_data

    except Exception as e:
        resolved_overtime_complaints = []
        print(f"Error fetching resolved overtime complaints: {e}")

    finally:
        cur.close()
        db.close()

    # Render the resolved overtime complaints in a template
    return render(request, 'resolved_overtime_complaints.html', {'resolved_overtime_complaints': resolved_overtime_complaints})

def resolved_within_time_complaints_view(request):
    # Connect to the database
    cur, db = connection()

    try:
        # Fetch complaints resolved within time
        query = """
            SELECT 
                u.name AS complainant_name, 
                u.department, 
                u.location, 
                COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
                c.issue_raise_date AS date,
                c.due_date,
                c.status,
                c.resolution,
                c.description,
                c.complaint_type,
                c.updated_at AS resolved_at  -- Include resolved_at field
            FROM 
                complaint AS c
            JOIN 
                user AS u ON c.user_id = u.id
            LEFT JOIN 
                user AS u2 ON c.resolved_by_id = u2.id AND (u2.role = 'admin' OR u2.role = 'superadmin')
            WHERE 
                c.status = 'resolved' AND c.complain_status = 'done'
            ORDER BY 
                c.issue_raise_date DESC;
        """
        cur.execute(query)
        resolved_within_time_complaints = cur.fetchall()

        # Prepare data for template rendering
        columns = [col[0] for col in cur.description]  # Extract column names
        complaints_data = []

        for row in resolved_within_time_complaints:
            # Extract the data for each complaint
            complainant_name = row[0]
            department = row[1]
            location = row[2].capitalize()
            resolved_by = row[3]
            date = row[4]
            due_date = row[5]
            status = row[6].capitalize()
            resolution = row[7]
            description = row[8]
            complaint_type = row[9]
            resolved_at = row[10]

            # Calculate TAT if resolved_at exists
            if resolved_at:
                tat_duration = resolved_at - date
            else:
                tat_duration = datetime.now() - date

            total_seconds = tat_duration.total_seconds()
            days = int(total_seconds // 86400)
            hours = int((total_seconds % 86400) // 3600)
            minutes = int((total_seconds % 3600) // 60)
            seconds = int(total_seconds % 60)

            # Format TAT based on the conditions
            if days > 0:
                tat_formatted = f"{days}d {hours}h {minutes}m {seconds}s"
            elif hours > 0:
                tat_formatted = f"{hours}h {minutes}m {seconds}s"
            else:
                tat_formatted = f"{minutes}m {seconds}s"

            complaints_data.append({
                'name': complainant_name,
                'department': department,
                'location': location,
                'resolved_by': resolved_by,
                'date': date.strftime("%Y-%m-%d %H:%M:%S"),
                'due_date': due_date.strftime("%Y-%m-%d %H:%M:%S"),
                'status': status,
                'resolution': resolution,
                'description': description,
                'complaint_type': complaint_type,
                'resolved_at': resolved_at.strftime("%Y-%m-%d %H:%M:%S") if resolved_at else 'None',
                'tat': tat_formatted if resolved_at else 'None'  # Add TAT to the report data
            })

        # Pass the complaint data to the template
        resolved_within_time_complaints = complaints_data

    except Exception as e:
        resolved_within_time_complaints = []
        print(f"Error fetching resolved within time complaints: {e}")

    finally:
        cur.close()
        db.close()

    # Render the resolved within time complaints in a template
    return render(request, 'resolved_within_time_complaints.html', {'resolved_within_time_complaints': resolved_within_time_complaints})

def all_complaints_view(request):  # Template to show all complaints in superadmin dashboard
    # Connect to the database
    cur, db = connection()

    try:
        # Fetch all complaints (no status filter)
        query = """
            SELECT 
                u.name AS complainant_name, 
                u.department, 
                u.location, 
                COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
                c.issue_raise_date AS date,
                c.due_date,
                c.status,
                c.resolution,
                c.description,
                c.complaint_type
            FROM 
                complaint AS c
            JOIN 
                user AS u ON c.user_id = u.id
            LEFT JOIN 
                user AS u2 ON c.resolved_by_id = u2.id AND (u2.role = 'admin' OR u2.role = 'superadmin')
            ORDER BY 
                c.issue_raise_date DESC;
        """
        cur.execute(query)
        all_complaints = cur.fetchall()

        # Prepare data for template rendering
        columns = [col[0] for col in cur.description]  # Extract column names
        all_complaints = [dict(zip(columns, row)) for row in all_complaints]

    except Exception as e:
        all_complaints = []
        print(f"Error fetching all complaints: {e}")

    finally:
        cur.close()
        db.close()

    # Render the all complaints in a template
    return render(request, 'all_complaints.html', {'all_complaints': all_complaints})

@never_cache
def resolver_complaints_view(request):  # Template to show complaints for a specific resolver
    # Check if the user is logged in
    if 'user_id' not in request.session:
        return redirect('login')  # Redirect to the login page if not logged in

    # Get the logged-in user's ID, name, role, and location from the session
    user_id = request.session.get('user_id')
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')

    # Establish the connection and get the cursor
    cur, db = connection()

    try:
        # Fetch resolver's location from the database
        cur.execute("SELECT location FROM user WHERE id = %s", [user_id])
        resolver_location = cur.fetchone()
        if resolver_location:
            resolver_location = resolver_location[0]  # Extract the location value
        else:
            resolver_location = None  # Default to None if no location found

        # Fetch all complaints for the resolver's location
        query = """
            SELECT 
                u.name AS complainant_name, 
                u.department, 
                u.location, 
                COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
                c.issue_raise_date AS date,
                c.due_date,
                c.status,
                c.resolution,
                c.description,
                c.complaint_type,
                c.updated_at
            FROM 
                complaint AS c
            JOIN 
                user AS u ON c.user_id = u.id
            LEFT JOIN 
                user AS u2 ON c.resolved_by_id = u2.id AND (u2.role = 'admin' OR u2.role = 'superadmin')
            WHERE
                c.location = %s
            ORDER BY 
                c.issue_raise_date DESC;
        """

        # Execute the query with the resolver's location filter
        cur.execute(query, [resolver_location])
        resolver_complaints = cur.fetchall()

        # Prepare data for template rendering
        columns = [col[0] for col in cur.description]  # Extract column names
        resolver_complaints = [dict(zip(columns, row)) for row in resolver_complaints]

    except Exception as e:
        resolver_complaints = []
        print(f"Error fetching resolver complaints: {e}")

    finally:
        cur.close()
        db.close()

    # Render the complaints in a template
    return render(request, 'resolver_complaints.html', {
        'resolver_complaints': resolver_complaints,
        'user_name': user_name,
        'user_role': user_role,
    })

@never_cache
def resolved_within_time_view(request):  # Template to show complaints resolved within time
    # Check if the user is logged in
    if 'user_id' not in request.session:
        return redirect('login')  # Redirect to the login page if not logged in

    # Get the logged-in user's ID, name, role, and location from the session
    user_id = request.session.get('user_id')
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')

    # Establish the connection and get the cursor
    cur, db = connection()

    try:
        # Fetch resolver's location from the database
        cur.execute("SELECT location FROM user WHERE id = %s", [user_id])
        resolver_location = cur.fetchone()
        if resolver_location:
            resolver_location = resolver_location[0]  # Extract the location value
        else:
            resolver_location = None  # Default to None if no location found

        # Fetch all complaints resolved within time for the resolver's location
        query = """
            SELECT 
                u.name AS complainant_name, 
                u.department, 
                u.location, 
                COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
                c.issue_raise_date AS date,
                c.due_date,
                c.status,
                c.resolution,
                c.description,
                c.complaint_type,
                c.updated_at
            FROM 
                complaint AS c
            JOIN 
                user AS u ON c.user_id = u.id
            LEFT JOIN 
                user AS u2 ON c.resolved_by_id = u2.id AND (u2.role = 'admin' OR u2.role = 'superadmin')
            WHERE
                c.location = %s
                AND c.status = 'resolved'
                AND c.complain_status = 'done'
            ORDER BY 
                c.issue_raise_date DESC;
        """

        # Execute the query with the resolver's location filter
        cur.execute(query, [resolver_location])
        resolved_within_time = cur.fetchall()

        # Prepare data for template rendering
        columns = [col[0] for col in cur.description]  # Extract column names
        resolved_within_time = [dict(zip(columns, row)) for row in resolved_within_time]

    except Exception as e:
        resolved_within_time = []
        print(f"Error fetching resolved within time complaints: {e}")

    finally:
        cur.close()
        db.close()

    # Render the resolved complaints in a template
    return render(request, 'resolved_within_time.html', {
        'resolved_within_time': resolved_within_time,
        'user_name': user_name,
        'user_role': user_role,
    })
    
@never_cache
def resolved_over_time_view(request):  # Template to show complaints resolved over time
    # Check if the user is logged in
    if 'user_id' not in request.session:
        return redirect('login')  # Redirect to the login page if not logged in

    # Get the logged-in user's ID, name, role, and location from the session
    user_id = request.session.get('user_id')
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')

    # Establish the connection and get the cursor
    cur, db = connection()

    try:
        # Fetch resolver's location from the database
        cur.execute("SELECT location FROM user WHERE id = %s", [user_id])
        resolver_location = cur.fetchone()
        if resolver_location:
            resolver_location = resolver_location[0]  # Extract the location value
        else:
            resolver_location = None  # Default to None if no location found

        # Fetch all complaints resolved over time for the resolver's location
        query = """
            SELECT 
                u.name AS complainant_name, 
                u.department, 
                u.location, 
                COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
                c.issue_raise_date AS date,
                c.due_date,
                c.status,
                c.resolution,
                c.description,
                c.complaint_type,
                c.updated_at
            FROM 
                complaint AS c
            JOIN 
                user AS u ON c.user_id = u.id
            LEFT JOIN 
                user AS u2 ON c.resolved_by_id = u2.id AND (u2.role = 'admin' OR u2.role = 'superadmin')
            WHERE
                c.location = %s
                AND c.status = 'resolved'
                AND c.updated_at > c.due_date
                AND c.complain_status = 'overdue'
            ORDER BY 
                c.issue_raise_date DESC;
        """

        # Execute the query with the resolver's location filter
        cur.execute(query, [resolver_location])
        resolved_over_time = cur.fetchall()

        # Prepare data for template rendering
        columns = [col[0] for col in cur.description]  # Extract column names
        resolved_over_time = [dict(zip(columns, row)) for row in resolved_over_time]

    except Exception as e:
        resolved_over_time = []
        print(f"Error fetching resolved over time complaints: {e}")

    finally:
        cur.close()
        db.close()

    # Render the resolved complaints in a template
    return render(request, 'resolved_over_time.html', {
        'resolved_over_time': resolved_over_time,
        'user_name': user_name,
        'user_role': user_role,
    })
    
@never_cache
def pending_resolver_complaint(request):  # Template to show pending complaints for the resolver
    # Check if the user is logged in
    if 'user_id' not in request.session:
        return redirect('login')  # Redirect to the login page if not logged in

    # Get the logged-in user's ID, name, role, and location from the session
    user_id = request.session.get('user_id')
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')

    # Establish the connection and get the cursor
    cur, db = connection()

    try:
        # Fetch resolver's location from the database
        cur.execute("SELECT location FROM user WHERE id = %s", [user_id])
        resolver_location = cur.fetchone()
        if resolver_location:
            resolver_location = resolver_location[0]  # Extract the location value
        else:
            resolver_location = None  # Default to None if no location found

        # Fetch all pending complaints for the resolver's location
        query = """
            SELECT 
                u.name AS complainant_name, 
                u.department, 
                u.location, 
                COALESCE(u2.name, 'Not Assigned') AS resolved_by_name,  
                c.issue_raise_date AS date,
                c.due_date,
                c.status,
                c.resolution,
                c.description,
                c.complaint_type,
                c.updated_at
            FROM 
                complaint AS c
            JOIN 
                user AS u ON c.user_id = u.id
            LEFT JOIN 
                user AS u2 ON c.resolved_by_id = u2.id AND (u2.role = 'admin' OR u2.role = 'superadmin')
            WHERE
                c.location = %s
                AND c.status = 'pending'
            ORDER BY 
                c.issue_raise_date DESC;
        """

        # Execute the query with the resolver's location filter
        cur.execute(query, [resolver_location])
        pending_complaints = cur.fetchall()

        # Prepare data for template rendering
        columns = [col[0] for col in cur.description]  # Extract column names
        pending_complaints = [dict(zip(columns, row)) for row in pending_complaints]

    except Exception as e:
        pending_complaints = []
        print(f"Error fetching pending complaints: {e}")

    finally:
        cur.close()
        db.close()

    # Render the pending complaints in a template
    return render(request, 'pendingresolvercomplaints.html', {
        'pending_complaints': pending_complaints,
        'user_name': user_name,
        'user_role': user_role,
    })
    
@never_cache
def inprogress_resolver_complaint(request):  # Template to show 'work in progress' complaints for the resolver
    # Check if the user is logged in
    if 'user_id' not in request.session:
        return redirect('login')  # Redirect to the login page if not logged in

    # Get the logged-in user's ID, name, role, and location from the session
    user_id = request.session.get('user_id')
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')

    # Establish the connection and get the cursor
    cur, db = connection()

    try:
        # Fetch resolver's location from the database
        cur.execute("SELECT location FROM user WHERE id = %s", [user_id])
        resolver_location = cur.fetchone()
        if resolver_location:
            resolver_location = resolver_location[0]  # Extract the location value
        else:
            resolver_location = None  # Default to None if no location found

        # Fetch all 'work in progress' complaints for the resolver's location
        query = """
            SELECT 
                u.name AS complainant_name, 
                u.department, 
                u.location, 
                COALESCE(u2.name, 'Not Assigned') AS resolved_by_name,  
                c.issue_raise_date AS date,
                c.due_date,
                c.status,
                c.resolution,
                c.description,
                c.complaint_type,
                c.updated_at
            FROM 
                complaint AS c
            JOIN 
                user AS u ON c.user_id = u.id
            LEFT JOIN 
                user AS u2 ON c.resolved_by_id = u2.id AND (u2.role = 'admin' OR u2.role = 'superadmin')
            WHERE
                c.location = %s
                AND c.status = 'work in progress'
            ORDER BY 
                c.issue_raise_date DESC;
        """

        # Execute the query with the resolver's location filter
        cur.execute(query, [resolver_location])
        inprogress_complaints = cur.fetchall()

        # Prepare data for template rendering
        columns = [col[0] for col in cur.description]  # Extract column names
        inprogress_complaints = [dict(zip(columns, row)) for row in inprogress_complaints]

    except Exception as e:
        inprogress_complaints = []
        print(f"Error fetching 'work in progress' complaints: {e}")

    finally:
        cur.close()
        db.close()

    # Render the 'work in progress' complaints in a template
    return render(request, 'inprogressresolvercomplaints.html', {
        'inprogress_complaints': inprogress_complaints,
        'user_name': user_name,
        'user_role': user_role,
    })
    
def user_complaints(request):
    # Get the logged-in user's ID, name, and role from the session
    user_id = request.session.get('user_id')
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')
    
    # If user_id is not found in session, redirect to login page
    if not user_id:
        return redirect('login')  # Adjust to your actual login URL

    # Establish the connection and get the cursor
    cur, db = connection()

    # Join user and complaint tables to get the desired data, filtered by the logged-in user
    query = """
    SELECT 
        c.id, 
        u.name, 
        c.department, 
        c.location, 
        c.issue_raise_date, 
        c.status, 
        c.description, 
        c.due_date, 
        c.complaint_type, 
        c.updated_at
    FROM 
        user u
    INNER JOIN 
        complaint c ON u.id = c.user_id
    WHERE
        u.id = %s  -- Filter complaints for the logged-in user
    ORDER BY 
        c.issue_raise_date DESC
    """
    
    cur.execute(query, (user_id,))
    complaints = cur.fetchall()

    complaint_data = []
    
    for complaint_details in complaints:
        issue_raise_date = complaint_details[4]  # Fetching issue_raise_date
        updated_at = complaint_details[9]  # Fetching updated_at
        status = complaint_details[5].lower()  # Get the status and convert to lowercase for comparison

        # Initialize TAT as 'Not resolved yet' by default
        tat_formatted = 'Not resolved yet'

        if status == "resolved" and updated_at:
            # Calculate TAT as updated_at - issue_raise_date only if resolved
            tat_duration = updated_at - issue_raise_date

            # Extract days, hours, minutes, and seconds from TAT duration
            total_seconds = tat_duration.total_seconds()
            days = int(total_seconds // 86400)  # Calculate days
            hours = int((total_seconds % 86400) // 3600)  # Calculate hours
            minutes = int((total_seconds % 3600) // 60)  # Calculate minutes
            seconds = int(total_seconds % 60)  # Calculate seconds

            # Format TAT
            tat_formatted = ""
            if days > 0:
                tat_formatted += f"{days}d "
            if hours > 0 or days > 0:  # Always show hours if there are days
                tat_formatted += f"{hours}h "
            tat_formatted += f"{minutes}m {seconds}s"  # Add minutes and seconds

        # Add the formatted data to complaint_data
        complaint_data.append({
            'complaint_id': complaint_details[0],
            'user_name': complaint_details[1],
            'department': complaint_details[2],
            'location': complaint_details[3].capitalize(),
            'issue_raise_date': complaint_details[4].strftime("%Y-%m-%d %H:%M:%S"),
            'status': complaint_details[5].capitalize(),
            'description': complaint_details[6],
            'due_date': complaint_details[7].strftime("%Y-%m-%d %H:%M:%S"),
            'complaint_type': complaint_details[8],
            'resolved_at': updated_at.strftime("%Y-%m-%d %H:%M:%S") if updated_at else 'Not resolved yet',
            'tat': tat_formatted,  # Add the formatted TAT to the complaint data
        })

    # Close the cursor and database connection
    cur.close()
    db.close()

    # Return the template with the complaint data
    return render(request, 'user_complaints.html', {
        'complaints': complaint_data,
        'user_name': user_name,
        'user_role': user_role,
    })
    
def pending_user_complaints(request):
    # Get the logged-in user's ID, name, and role from the session
    user_id = request.session.get('user_id')
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')
    
    # If user_id is not found in session, redirect to login page
    if not user_id:
        return redirect('login')  # Adjust to your actual login URL

    # Establish the connection and get the cursor
    cur, db = connection()

    # Modify the query to filter only pending complaints
    query = """
    SELECT 
        c.id, 
        u.name, 
        c.department, 
        c.location, 
        c.issue_raise_date, 
        c.status, 
        c.description, 
        c.due_date, 
        c.complaint_type, 
        c.updated_at
    FROM 
        user u
    INNER JOIN 
        complaint c ON u.id = c.user_id
    WHERE
        u.id = %s  -- Filter complaints for the logged-in user
        AND c.status = 'pending'  -- Only get pending complaints
    ORDER BY 
        c.issue_raise_date DESC
    """
    
    cur.execute(query, (user_id,))
    complaints = cur.fetchall()

    complaint_data = []
    
    for complaint_details in complaints:
        issue_raise_date = complaint_details[4]  # Fetching issue_raise_date
        updated_at = complaint_details[9]  # Fetching updated_at
        status = complaint_details[5].lower()  # Get the status and convert to lowercase for comparison

        # Initialize TAT as 'Not resolved yet' by default
        tat_formatted = 'Not resolved yet'

        if status == "resolved" and updated_at:
            # Calculate TAT as updated_at - issue_raise_date only if resolved
            tat_duration = updated_at - issue_raise_date

            # Extract days, hours, minutes, and seconds from TAT duration
            total_seconds = tat_duration.total_seconds()
            days = int(total_seconds // 86400)  # Calculate days
            hours = int((total_seconds % 86400) // 3600)  # Calculate hours
            minutes = int((total_seconds % 3600) // 60)  # Calculate minutes
            seconds = int(total_seconds % 60)  # Calculate seconds

            # Format TAT
            tat_formatted = ""
            if days > 0:
                tat_formatted += f"{days}d "
            if hours > 0 or days > 0:  # Always show hours if there are days
                tat_formatted += f"{hours}h "
            tat_formatted += f"{minutes}m {seconds}s"  # Add minutes and seconds

        # Add the formatted data to complaint_data
        complaint_data.append({
            'complaint_id': complaint_details[0],
            'user_name': complaint_details[1],
            'department': complaint_details[2],
            'location': complaint_details[3].capitalize(),
            'issue_raise_date': complaint_details[4].strftime("%Y-%m-%d %H:%M:%S"),
            'status': complaint_details[5].capitalize(),
            'description': complaint_details[6],
            'due_date': complaint_details[7].strftime("%Y-%m-%d %H:%M:%S"),
            'complaint_type': complaint_details[8],
            'resolved_at': updated_at.strftime("%Y-%m-%d %H:%M:%S") if updated_at else 'Not resolved yet',
            'tat': tat_formatted,  # Add the formatted TAT to the complaint data
        })

    # Close the cursor and database connection
    cur.close()
    db.close()

    # Return the template with the complaint data
    return render(request, 'pendingusercomplaints.html', {
        'complaints': complaint_data,
        'user_name': user_name,
        'user_role': user_role,
    })
    
def inprogress_user_complaints(request):
    # Get the logged-in user's ID, name, and role from the session
    user_id = request.session.get('user_id')
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')
    
    # If user_id is not found in session, redirect to login page
    if not user_id:
        return redirect('login')  # Adjust to your actual login URL

    # Establish the connection and get the cursor
    cur, db = connection()

    # Modify the query to filter only pending complaints
    query = """
    SELECT 
        c.id, 
        u.name, 
        c.department, 
        c.location, 
        c.issue_raise_date, 
        c.status, 
        c.description, 
        c.due_date, 
        c.complaint_type, 
        c.updated_at
    FROM 
        user u
    INNER JOIN 
        complaint c ON u.id = c.user_id
    WHERE
        u.id = %s  -- Filter complaints for the logged-in user
        AND c.status = 'work in progress'  -- Only get pending complaints
    ORDER BY 
        c.issue_raise_date DESC
    """
    
    cur.execute(query, (user_id,))
    complaints = cur.fetchall()

    complaint_data = []
    
    for complaint_details in complaints:
        issue_raise_date = complaint_details[4]  # Fetching issue_raise_date
        updated_at = complaint_details[9]  # Fetching updated_at
        status = complaint_details[5].lower()  # Get the status and convert to lowercase for comparison

        # Initialize TAT as 'Not resolved yet' by default
        tat_formatted = 'Not resolved yet'

        if status == "resolved" and updated_at:
            # Calculate TAT as updated_at - issue_raise_date only if resolved
            tat_duration = updated_at - issue_raise_date

            # Extract days, hours, minutes, and seconds from TAT duration
            total_seconds = tat_duration.total_seconds()
            days = int(total_seconds // 86400)  # Calculate days
            hours = int((total_seconds % 86400) // 3600)  # Calculate hours
            minutes = int((total_seconds % 3600) // 60)  # Calculate minutes
            seconds = int(total_seconds % 60)  # Calculate seconds

            # Format TAT
            tat_formatted = ""
            if days > 0:
                tat_formatted += f"{days}d "
            if hours > 0 or days > 0:  # Always show hours if there are days
                tat_formatted += f"{hours}h "
            tat_formatted += f"{minutes}m {seconds}s"  # Add minutes and seconds

        # Add the formatted data to complaint_data
        complaint_data.append({
            'complaint_id': complaint_details[0],
            'user_name': complaint_details[1],
            'department': complaint_details[2],
            'location': complaint_details[3].capitalize(),
            'issue_raise_date': complaint_details[4].strftime("%Y-%m-%d %H:%M:%S"),
            'status': complaint_details[5].capitalize(),
            'description': complaint_details[6],
            'due_date': complaint_details[7].strftime("%Y-%m-%d %H:%M:%S"),
            'complaint_type': complaint_details[8],
            'resolved_at': updated_at.strftime("%Y-%m-%d %H:%M:%S") if updated_at else 'Not resolved yet',
            'tat': tat_formatted,  # Add the formatted TAT to the complaint data
        })

    # Close the cursor and database connection
    cur.close()
    db.close()

    # Return the template with the complaint data
    return render(request, 'inprogressusercomplaints.html', {
        'complaints': complaint_data,
        'user_name': user_name,
        'user_role': user_role,
    })
    
def resolvedwithintime_user_complaints(request):
    # Get the logged-in user's ID, name, and role from the session
    user_id = request.session.get('user_id')
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')
    
    # If user_id is not found in session, redirect to login page
    if not user_id:
        return redirect('login')  # Adjust to your actual login URL

    # Establish the connection and get the cursor
    cur, db = connection()

    # Modify the query to filter only pending complaints
    query = """
    SELECT 
        c.id, 
        u.name, 
        c.department, 
        c.location, 
        c.issue_raise_date, 
        c.status, 
        c.description, 
        c.due_date, 
        c.complaint_type, 
        c.updated_at
    FROM 
        user u
    INNER JOIN 
        complaint c ON u.id = c.user_id
    WHERE
        u.id = %s  -- Filter complaints for the logged-in user
        AND c.status = 'resolved'  -- Filter complaints with status 'resolved'
        AND c.complain_status = 'done'  -- Filter complaints with complain_status 'done'
    ORDER BY 
        c.issue_raise_date DESC
    """
    
    cur.execute(query, (user_id,))
    complaints = cur.fetchall()

    complaint_data = []
    
    for complaint_details in complaints:
        issue_raise_date = complaint_details[4]  # Fetching issue_raise_date
        updated_at = complaint_details[9]  # Fetching updated_at
        status = complaint_details[5].lower()  # Get the status and convert to lowercase for comparison

        # Initialize TAT as 'Not resolved yet' by default
        tat_formatted = 'Not resolved yet'

        if status == "resolved" and updated_at:
            # Calculate TAT as updated_at - issue_raise_date only if resolved
            tat_duration = updated_at - issue_raise_date

            # Extract days, hours, minutes, and seconds from TAT duration
            total_seconds = tat_duration.total_seconds()
            days = int(total_seconds // 86400)  # Calculate days
            hours = int((total_seconds % 86400) // 3600)  # Calculate hours
            minutes = int((total_seconds % 3600) // 60)  # Calculate minutes
            seconds = int(total_seconds % 60)  # Calculate seconds

            # Format TAT
            tat_formatted = ""
            if days > 0:
                tat_formatted += f"{days}d "
            if hours > 0 or days > 0:  # Always show hours if there are days
                tat_formatted += f"{hours}h "
            tat_formatted += f"{minutes}m {seconds}s"  # Add minutes and seconds

        # Add the formatted data to complaint_data
        complaint_data.append({
            'complaint_id': complaint_details[0],
            'user_name': complaint_details[1],
            'department': complaint_details[2],
            'location': complaint_details[3].capitalize(),
            'issue_raise_date': complaint_details[4].strftime("%Y-%m-%d %H:%M:%S"),
            'status': complaint_details[5].capitalize(),
            'description': complaint_details[6],
            'due_date': complaint_details[7].strftime("%Y-%m-%d %H:%M:%S"),
            'complaint_type': complaint_details[8],
            'resolved_at': updated_at.strftime("%Y-%m-%d %H:%M:%S") if updated_at else 'Not resolved yet',
            'tat': tat_formatted,  # Add the formatted TAT to the complaint data
        })

    # Close the cursor and database connection
    cur.close()
    db.close()

    # Return the template with the complaint data
    return render(request, 'resolvedwithintimeusercomplaints.html', {
        'complaints': complaint_data,
        'user_name': user_name,
        'user_role': user_role,
    })
    
def resolvedovertime_user_complaints(request):
    # Get the logged-in user's ID, name, and role from the session
    user_id = request.session.get('user_id')
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')
    
    # If user_id is not found in session, redirect to login page
    if not user_id:
        return redirect('login')  # Adjust to your actual login URL

    # Establish the connection and get the cursor
    cur, db = connection()

    # Modify the query to filter only pending complaints
    query = """
    SELECT 
        c.id, 
        u.name, 
        c.department, 
        c.location, 
        c.issue_raise_date, 
        c.status, 
        c.description, 
        c.due_date, 
        c.complaint_type, 
        c.updated_at
    FROM 
        user u
    INNER JOIN 
        complaint c ON u.id = c.user_id
    WHERE
        u.id = %s  -- Filter complaints for the logged-in user
        AND c.status = 'resolved'  -- Filter complaints with status 'resolved'
        AND c.complain_status = 'overdue'  -- Filter complaints with complain_status 'done'
    ORDER BY 
        c.issue_raise_date DESC
    """
    
    cur.execute(query, (user_id,))
    complaints = cur.fetchall()

    complaint_data = []
    
    for complaint_details in complaints:
        issue_raise_date = complaint_details[4]  # Fetching issue_raise_date
        updated_at = complaint_details[9]  # Fetching updated_at
        status = complaint_details[5].lower()  # Get the status and convert to lowercase for comparison

        # Initialize TAT as 'Not resolved yet' by default
        tat_formatted = 'Not resolved yet'

        if status == "resolved" and updated_at:
            # Calculate TAT as updated_at - issue_raise_date only if resolved
            tat_duration = updated_at - issue_raise_date

            # Extract days, hours, minutes, and seconds from TAT duration
            total_seconds = tat_duration.total_seconds()
            days = int(total_seconds // 86400)  # Calculate days
            hours = int((total_seconds % 86400) // 3600)  # Calculate hours
            minutes = int((total_seconds % 3600) // 60)  # Calculate minutes
            seconds = int(total_seconds % 60)  # Calculate seconds

            # Format TAT
            tat_formatted = ""
            if days > 0:
                tat_formatted += f"{days}d "
            if hours > 0 or days > 0:  # Always show hours if there are days
                tat_formatted += f"{hours}h "
            tat_formatted += f"{minutes}m {seconds}s"  # Add minutes and seconds

        # Add the formatted data to complaint_data
        complaint_data.append({
            'complaint_id': complaint_details[0],
            'user_name': complaint_details[1],
            'department': complaint_details[2],
            'location': complaint_details[3].capitalize(),
            'issue_raise_date': complaint_details[4].strftime("%Y-%m-%d %H:%M:%S"),
            'status': complaint_details[5].capitalize(),
            'description': complaint_details[6],
            'due_date': complaint_details[7].strftime("%Y-%m-%d %H:%M:%S"),
            'complaint_type': complaint_details[8],
            'resolved_at': updated_at.strftime("%Y-%m-%d %H:%M:%S") if updated_at else 'Not resolved yet',
            'tat': tat_formatted,  # Add the formatted TAT to the complaint data
        })

    # Close the cursor and database connection
    cur.close()
    db.close()

    # Return the template with the complaint data
    return render(request, 'resolvedovertimeusercomplaints.html', {
        'complaints': complaint_data,
        'user_name': user_name,
        'user_role': user_role,
    })
    
def resolvercharts(request):
    # Establish the connection and get the cursor
    cur, db = connection()  # Ensure `connection()` is correctly defined in your project

    # Fetch Graphical additional data
    cur.execute("SELECT location, COUNT(*) FROM complaint GROUP BY location")
    location_data = cur.fetchall()

    cur.execute("SELECT status, COUNT(*) FROM complaint GROUP BY status")
    status_data = cur.fetchall()

    cur.execute("SELECT location, status, COUNT(*) FROM complaint GROUP BY location, status")
    state_complaints_data = cur.fetchall()

    cur.execute("SELECT complaint_type, COUNT(*) FROM complaint GROUP BY complaint_type")
    category_data = cur.fetchall()

    # Fetch performance data for each resolver
    cur.execute("SELECT resolved_by_id, COUNT(*) FROM complaint WHERE status = 'resolved' GROUP BY resolved_by_id")
    resolver_performance_data = cur.fetchall()

    cur.execute("SELECT department, COUNT(*) FROM complaint GROUP BY department")
    department_data = cur.fetchall()

    cur.execute("SELECT location, COUNT(*) FROM complaint WHERE status = 'resolved' GROUP BY location")
    locations_data = cur.fetchall()

    # Get current resolver's data (for example, Alice with resolver ID 1)
    current_resolver_id = 1  # Update this with the actual resolver ID, e.g., from session or user model
    resolver_score = 0
    total_complaints = 0
    for resolver, count in resolver_performance_data:
        if resolver == current_resolver_id:
            resolver_score = count
        total_complaints += count

    # Prepare data for graphs
    locations = [item[0] for item in location_data]
    location_counts = [item[1] for item in location_data]

    statuses = [item[0] for item in status_data]
    status_counts = [item[1] for item in status_data]

    states = [item[0] for item in state_complaints_data]
    states_status = [item[1] for item in state_complaints_data]
    state_counts = [item[2] for item in state_complaints_data]

    categories = [item[0] for item in category_data]
    category_counts = [item[1] for item in category_data]

    resolvers = [item[0] for item in resolver_performance_data]
    resolver_counts = [item[1] for item in resolver_performance_data]

    departments = [row[0] for row in department_data]
    department_counts = [row[1] for row in department_data]

    locations2 = [item[0] for item in locations_data]
    location_counts2 = [item[1] for item in locations_data]

    # Close the cursor and database connection
    cur.close()
    db.close()

    # Prepare the context data for the template
    context = {
        'departments': departments,
        'department_counts': department_counts,
        'locations': locations,
        'location_counts': location_counts,
        'locations2': locations2,
        'location_counts2': location_counts2,
        'statuses': statuses,
        'status_counts': status_counts,
        'states': states,
        'states_status': states_status,
        'state_counts': state_counts,
        'categories': categories,
        'category_counts': category_counts,
        'resolvers': resolvers,
        'resolver_counts': resolver_counts,
        'resolver_score': resolver_score,
        'total_complaints': total_complaints,  # Add total complaints to the context
    }

    return render(request, 'resolverdash.html', context)

def getcomplainttypes(request):
    selectedDept = request.GET.get('dept')  # Get the selected department from the request
    if not selectedDept:
        return HttpResponse(
            json.dumps({'error': 'No department selected'}),
            content_type="application/json",
            status=400
        )
    
    try:
        cur, db = connection()  # Get the cursor and database connection
        query = "SELECT DISTINCT complaint_type FROM complaint_type_master WHERE department = %s;"
        cur.execute(query, [selectedDept])  # Use parameterized query to prevent SQL injection
        
        complainttypes = [{'complaint_type': row[0]} for row in cur.fetchall()]  # Format the results
        
        return HttpResponse(
            json.dumps({'data': complainttypes}),
            content_type="application/json",
            status=200
        )
    except Exception as e:
        return HttpResponse(
            json.dumps({'error': str(e)}),
            content_type="application/json",
            status=500
        )
    finally:
        cur.close()  # Close the cursor
        db.close()   # Close the database connection
        
def generate_superadmin_details_xlsx(complaint_data):
    """
    Generate an Excel file for Complaint History and return it as an HTTP response.
    """
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Complaint History"

    # Define the column headers in the required order
    headers = [
        "ID", "Name", "Department", "Location", "Complaint Type", 
        "Status", "Assigned To", "Issue Date","Updated At", "Resolution Date", "Remarks"
    ]
    
    # Define header styles
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Grey background
    header_font = Font(bold=True, color="000000", size=12)  # Black text

    # Apply header styles and set column headers
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f'{col_letter}1']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Populate the worksheet with complaint data
    for row_num, complaint in enumerate(complaint_data, 2):
        ws[f'A{row_num}'] = complaint['complaint_id']
        ws[f'B{row_num}'] = complaint['name']  # Updated field for complainant name
        ws[f'C{row_num}'] = complaint['department']
        ws[f'D{row_num}'] = complaint['location']
        ws[f'E{row_num}'] = complaint['complaint_type']
        ws[f'F{row_num}'] = complaint['status']  # Updated field for status
        ws[f'G{row_num}'] = complaint['assigned_to']
        ws[f'H{row_num}'] = complaint['issue_raise_date']  # Updated field for issue date
        ws[f'I{row_num}'] = complaint['updated_at']  # Added updated_at column
        ws[f'J{row_num}'] = complaint['updated_due_date']  # Updated field for resolution date
        ws[f'K{row_num}'] = complaint['reason']  # Updated field for remarks

        # Format date columns for readability
        ws[f'H{row_num}'].number_format = 'DD-MM-YYYY HH:MM:SS'
        ws[f'I{row_num}'].number_format = 'DD-MM-YYYY HH:MM:SS'
        ws[f'J{row_num}'].number_format = 'DD-MM-YYYY HH:MM:SS'

        # Enable text wrapping for "Remarks" column
        ws[f'K{row_num}'].alignment = Alignment(wrap_text=True)

    # Adjust column widths for better readability
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2  # Add extra padding
        ws.column_dimensions[col_letter].width = adjusted_width

    # Set fixed widths for specific columns
    ws.column_dimensions['E'].width = 27  # Complaint Type column
    ws.column_dimensions['K'].width = 30  # Remarks column

    # Save the workbook to a BytesIO stream
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Create the response with the Excel file as an attachment
    response = HttpResponse(stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=complaint_history.xlsx'
    return response
        
def detailed_superadmin_reports(request):
    user_name = request.session.get('user_name')

    # Retrieve filter parameters from GET request
    fromdate = request.GET.get('fromdate', '')
    todate = request.GET.get('todate', '')
    complaint_type = request.GET.get('complaint_type', '')
    status = request.GET.get('status', '')
    location = request.GET.get('location', '')
    department = request.GET.get('department', '')

    # Prepare filter conditions
    filter_conditions = []
    query_params = []

    # Date filtering for fromdate and todate
    if fromdate and todate:
        filter_conditions.append("ch.issue_raise_date BETWEEN %s AND %s")
        query_params.extend([f"{fromdate} 00:00:00", f"{todate} 23:59:59"])
    elif fromdate:
        filter_conditions.append("ch.issue_raise_date >= %s")
        query_params.append(f"{fromdate} 00:00:00")
    elif todate:
        filter_conditions.append("ch.issue_raise_date <= %s")
        query_params.append(f"{todate} 23:59:59")

    # Filter by complaint_type
    if complaint_type:
        filter_conditions.append("ch.complaint_type = %s")
        query_params.append(complaint_type)

    # Filter by status
    if status:
        filter_conditions.append("ch.status = %s")
        query_params.append(status)

    # Filter by location
    if location:
        filter_conditions.append("ch.location = %s")
        query_params.append(location)

    # Filter by department
    if department:
        filter_conditions.append("ch.department = %s")
        query_params.append(department)

    # Build the WHERE clause if there are any filter conditions
    where_clause = f"WHERE {' AND '.join(filter_conditions)}" if filter_conditions else ""

    # Final query construction
    query = f"""
    SELECT 
        ch.id,
        ch.complaint_id,
        ch.user_id,
        ch.resolved_by_id,
        ch.department,
        ch.location,
        ch.complaint_type,
        ch.description,
        ch.issue_raise_date,
        ch.complain_status,
        ch.due_date,
        ch.tat,
        ch.updated_at,
        ch.name,
        ch.status,
        ch.resolution,
        ch.assigned_to,
        ch.updated_due_date,
        ch.reason,
        ch.change_timestamp,
        ch.change_action
    FROM 
        complaint_history AS ch
    {where_clause}
    ORDER BY 
        ch.issue_raise_date DESC;
    """

    # Execute the query with parameters
    cur.execute(query, query_params)
    history_reports = cur.fetchall()

    # Process the reports into a list of dictionaries
    report_data = []
    for report in history_reports:
        report_data.append({
            'id': report[0],
            'complaint_id': report[1],
            'user_id': report[2],
            'resolved_by_id': report[3],
            'department': report[4],
            'location': report[5].capitalize(),
            'complaint_type': report[6],
            'description': report[7],
            'issue_raise_date': report[8],
            'complain_status': report[9].capitalize(),
            'due_date': report[10],
            'tat': report[11],
            'updated_at': report[12],
            'name': report[13],
            'status': report[14].capitalize(),
            'resolution': report[15] if report[15] else "Not Resolved Yet",
            'assigned_to': report[16],
            'updated_due_date': report[17] if report[17] else 'No Resolution Date Yet',
            'reason': report[18] if report[18] else "No Reason Provided",
            'change_timestamp': report[19],
            'change_action': report[20]
        })

    # If user requests download, create Excel file
    if request.GET.get('download') == 'true':
        return generate_superadmin_details_xlsx(report_data)

    # Fetch distinct values for dropdown filters
    complaint_type_query = "SELECT DISTINCT complaint_type FROM complaint_history"
    cur.execute(complaint_type_query)
    distinct_complaint_types = [row[0] for row in cur.fetchall()]

    status_query = "SELECT DISTINCT status FROM complaint_history"
    cur.execute(status_query)
    distinct_statuses = [row[0] for row in cur.fetchall()]

    location_query = "SELECT DISTINCT location FROM complaint_history"
    cur.execute(location_query)
    distinct_locations = [row[0] for row in cur.fetchall()]

    department_query = "SELECT DISTINCT department FROM complaint_history"
    cur.execute(department_query)
    distinct_departments = [row[0] for row in cur.fetchall()]

    return render(request, 'detailed_superadmin_reports.html', {
        'report_data': report_data,
        'distinct_complaint_types': distinct_complaint_types,
        'distinct_statuses': distinct_statuses,
        'distinct_locations': distinct_locations,
        'distinct_departments': distinct_departments,
        'from_date': fromdate,
        'to_date': todate,
        'user_name': user_name
    })

def detailed_resolver_reports(request):
    user_name = request.session.get('user_name')

    # Retrieve filter parameters from GET request
    fromdate = request.GET.get('fromdate', '')
    todate = request.GET.get('todate', '')
    complaint_type = request.GET.get('complaint_type', '')
    status = request.GET.get('status', '')  # Retrieve status from GET parameters

    # Retrieve the assigned_to value from the session (user_id)
    assigned_to = request.session.get('user_id', '')

    # If assigned_to is not set in session, return an error message
    if not assigned_to:
        return render(request, 'error.html', {'message': 'Assigned resolver is not set in session.'})

    # Prepare filter conditions
    filter_conditions = []
    query_params = []

    # Filter by assigned_to (taken from session)
    filter_conditions.append("ch.assigned_to = %s")
    query_params.append(assigned_to)

    # Filter by complaint_type (if provided)
    if complaint_type:
        filter_conditions.append("ch.complaint_type = %s")
        query_params.append(complaint_type)

    # Filter by status (if provided)
    if status:
        filter_conditions.append("ch.status = %s")
        query_params.append(status)

    # Date filtering for fromdate and todate
    if fromdate and todate:
        filter_conditions.append("ch.issue_raise_date BETWEEN %s AND %s")
        query_params.extend([f"{fromdate} 00:00:00", f"{todate} 23:59:59"])
    elif fromdate:
        filter_conditions.append("ch.issue_raise_date >= %s")
        query_params.append(f"{fromdate} 00:00:00")
    elif todate:
        filter_conditions.append("ch.issue_raise_date <= %s")
        query_params.append(f"{todate} 23:59:59")

    # Build the WHERE clause if there are any filter conditions
    where_clause = f"WHERE {' AND '.join(filter_conditions)}" if filter_conditions else ""

    # Final query construction
    query = f"""
    SELECT 
        ch.id,
        ch.complaint_id,
        ch.user_id,
        ch.resolved_by_id,
        ch.department,
        ch.location,
        ch.complaint_type,
        ch.description,
        ch.issue_raise_date,
        ch.complain_status,
        ch.due_date,
        ch.tat,
        ch.updated_at,
        ch.name,
        ch.status,
        ch.resolution,
        ch.assigned_to,
        ch.updated_due_date,
        ch.reason,
        ch.change_timestamp,
        ch.change_action
    FROM 
        complaint_history AS ch
    {where_clause}
    ORDER BY 
        ch.issue_raise_date DESC;
    """

    # Execute the query with parameters
    cur.execute(query, query_params)
    history_reports = cur.fetchall()

    # Process the reports into a list of dictionaries
    report_data = []
    for report in history_reports:
        report_data.append({
            'id': report[0],
            'complaint_id': report[1],
            'user_id': report[2],
            'resolved_by_id': report[3],
            'department': report[4],
            'location': report[5].capitalize(),
            'complaint_type': report[6],
            'description': report[7],
            'issue_raise_date': report[8],
            'complain_status': report[9].capitalize(),
            'due_date': report[10],
            'tat': report[11],
            'updated_at': report[12],
            'name': report[13],
            'status': report[14].capitalize(),
            'resolution': report[15] if report[15] else "Not Resolved Yet",
            'assigned_to': report[16],
            'updated_due_date': report[17] if report[17] else 'No Resolution Date Yet',
            'reason': report[18] if report[18] else "No Reason Provided",
            'change_timestamp': report[19],
            'change_action': report[20]
        })

    # If user requests download, create Excel file
    if request.GET.get('download') == 'true':
        return generate_superadmin_details_xlsx(report_data)

    # Fetch distinct values for complaint_type filtered by assigned_to
    complaint_type_query = """
        SELECT DISTINCT complaint_type 
        FROM complaint_history 
        WHERE assigned_to = %s
    """
    cur.execute(complaint_type_query, [assigned_to])
    distinct_complaint_types = [row[0] for row in cur.fetchall()]

    # Fetch distinct values for status filtered by assigned_to
    status_query = """
        SELECT DISTINCT status 
        FROM complaint_history 
        WHERE assigned_to = %s
    """
    cur.execute(status_query, [assigned_to])
    distinct_statuses = [row[0] for row in cur.fetchall()]

    return render(request, 'detailedresolverreports.html', {
        'report_data': report_data,
        'distinct_complaint_types': distinct_complaint_types,
        'distinct_statuses': distinct_statuses,
        'from_date': fromdate,
        'to_date': todate,
        'user_name': user_name
    })

def detailed_userreports(request):
    user_name = request.session.get('user_name')

    # Retrieve filter parameters from GET request
    fromdate = request.GET.get('fromdate', '')
    todate = request.GET.get('todate', '')
    complaint_type = request.GET.get('complaint_type', '')
    status = request.GET.get('status', '')  # Retrieve status from GET parameters

    # Retrieve the user_id value from the session
    user_id = request.session.get('user_id', '')

    # If user_id is not set in session, return an error message
    if not user_id:
        return render(request, 'error.html', {'message': 'User is not set in session.'})

    # Prepare filter conditions
    filter_conditions = []
    query_params = []

    # Filter by user_id (taken from session)
    filter_conditions.append("ch.user_id = %s")
    query_params.append(user_id)

    # Filter by complaint_type (if provided)
    if complaint_type:
        filter_conditions.append("ch.complaint_type = %s")
        query_params.append(complaint_type)

    # Filter by status (if provided)
    if status:
        filter_conditions.append("ch.status = %s")
        query_params.append(status)

    # Date filtering for fromdate and todate
    if fromdate and todate:
        filter_conditions.append("ch.issue_raise_date BETWEEN %s AND %s")
        query_params.extend([f"{fromdate} 00:00:00", f"{todate} 23:59:59"])
    elif fromdate:
        filter_conditions.append("ch.issue_raise_date >= %s")
        query_params.append(f"{fromdate} 00:00:00")
    elif todate:
        filter_conditions.append("ch.issue_raise_date <= %s")
        query_params.append(f"{todate} 23:59:59")

    # Build the WHERE clause if there are any filter conditions
    where_clause = f"WHERE {' AND '.join(filter_conditions)}" if filter_conditions else ""

    # Final query construction
    query = f"""
    SELECT 
        ch.id,
        ch.complaint_id,
        ch.user_id,
        ch.resolved_by_id,
        ch.department,
        ch.location,
        ch.complaint_type,
        ch.description,
        ch.issue_raise_date,
        ch.complain_status,
        ch.due_date,
        ch.tat,
        ch.updated_at,
        ch.name,
        ch.status,
        ch.resolution,
        ch.assigned_to,
        ch.updated_due_date,
        ch.reason,
        ch.change_timestamp,
        ch.change_action
    FROM 
        complaint_history AS ch
    {where_clause}
    ORDER BY 
        ch.issue_raise_date DESC;
    """

    # Execute the query with parameters
    cur.execute(query, query_params)
    history_reports = cur.fetchall()

    # Process the reports into a list of dictionaries
    report_data = []
    for report in history_reports:
        report_data.append({
            'id': report[0],
            'complaint_id': report[1],
            'user_id': report[2],
            'resolved_by_id': report[3],
            'department': report[4],
            'location': report[5].capitalize(),
            'complaint_type': report[6],
            'description': report[7],
            'issue_raise_date': report[8],
            'complain_status': report[9].capitalize(),
            'due_date': report[10],
            'tat': report[11],
            'updated_at': report[12],
            'name': report[13],
            'status': report[14].capitalize(),
            'resolution': report[15] if report[15] else "Not Resolved Yet",
            'assigned_to': report[16],
            'updated_due_date': report[17] if report[17] else 'No Resolution Date Yet',
            'reason': report[18] if report[18] else "No Reason Provided",
            'change_timestamp': report[19],
            'change_action': report[20]
        })

    # If user requests download, create Excel file
    if request.GET.get('download') == 'true':
        return generate_superadmin_details_xlsx(report_data)

    # Fetch distinct values for complaint_type filter
    complaint_type_query = "SELECT DISTINCT complaint_type FROM complaint_history WHERE user_id = %s"
    cur.execute(complaint_type_query, [user_id])
    distinct_complaint_types = [row[0] for row in cur.fetchall()]

    # Fetch distinct values for status filter
    status_query = "SELECT DISTINCT status FROM complaint_history WHERE user_id = %s"
    cur.execute(status_query, [user_id])
    distinct_statuses = [row[0] for row in cur.fetchall()]

    return render(request, 'detaileduserreports.html', {
        'report_data': report_data,
        'distinct_complaint_types': distinct_complaint_types,
        'distinct_statuses': distinct_statuses,
        'from_date': fromdate,
        'to_date': todate,
        'user_name': user_name
    })

# @csrf_exempt
# def manage_office_timings(request):
#     user_name = request.session.get('user_name')
#     cur, db = connection()

#     try:
#         if request.method == "POST":
#             data = json.loads(request.body)
#             non_working_days = data.get('non_working_days')
#             work_start_time = data.get('work_start_time')
#             work_end_time = data.get('work_end_time')

#             # Validate required fields
#             if not (non_working_days and work_start_time and work_end_time):
#                 return JsonResponse({'success': False, 'message': 'All fields are required'}, status=400)

#             # Add new office timing
#             query = """
#                 INSERT INTO office_timings (non_working_days, work_start_time, work_end_time)
#                 VALUES (%s, %s, %s)
#             """
#             cur.execute(query, (non_working_days, work_start_time, work_end_time))
#             db.commit()
#             cur.execute("SELECT LAST_INSERT_ID()")
#             office_timing_id = cur.fetchone()[0]
#             return JsonResponse({
#                 'success': True,
#                 'message': 'Office timing added successfully',
#                 'office_timing': {
#                     'id': office_timing_id,
#                     'non_working_days': non_working_days,
#                     'work_start_time': work_start_time,
#                     'work_end_time': work_end_time
#                 }
#             })

#         elif request.method == "GET":
#             # Get all office timings
#             cur.execute("SELECT id, non_working_days, work_start_time, work_end_time FROM office_timings")
#             office_timings = cur.fetchall()
#             office_timings_data = [
#                 {
#                     'id': office_timing[0],
#                     'non_working_days': office_timing[1],
#                     'work_start_time': str(office_timing[2]),
#                     'work_end_time': str(office_timing[3])
#                 }
#                 for office_timing in office_timings
#             ]
#             return render(request, 'officetiming.html', {
#                 'user_name': user_name,
#                 'office_timings_data': office_timings_data
#             })


#         else:
#             return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

#     except Exception as e:
#         return JsonResponse({'success': False, 'message': str(e)}, status=500)


#     finally:
#         cur.close()
#         db.close()

# @csrf_exempt
# def edit_office_timing(request, office_timing_id):
#     user_name = request.session.get('user_name')
#     cur, db = connection()

#     try:
#         if request.method == "POST":
#             data = json.loads(request.body)
#             non_working_days = data.get('non_working_days')
#             work_start_time = data.get('work_start_time')
#             work_end_time = data.get('work_end_time')

#             # Validate required fields
#             if not (non_working_days and work_start_time and work_end_time):
#                 return JsonResponse({'success': False, 'message': 'All fields are required'}, status=400)

#             # Update office timing
#             query = """
#                 UPDATE office_timings
#                 SET non_working_days = %s, work_start_time = %s, work_end_time = %s
#                 WHERE id = %s
#             """
#             cur.execute(query, (non_working_days, work_start_time, work_end_time, office_timing_id))
#             db.commit()
#             return JsonResponse({'success': True, 'message': 'Office timing updated successfully'})

#         else:
#             return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

#     except Exception as e:
#         db.rollback()
#         return JsonResponse({'success': False, 'message': str(e)}, status=500)

#     finally:
#         cur.close()
#         db.close()

# @csrf_exempt
# def delete_office_timing(request, office_timing_id):
#     cur, db = connection()

#     try:
#         if request.method == "DELETE":
#             # Delete office timing
#             cur.execute("DELETE FROM office_timings WHERE id = %s", [office_timing_id])
#             db.commit()
#             return JsonResponse({'success': True, 'message': 'Office timing deleted successfully'})

#         else:
#             return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

#     except Exception as e:
#         db.rollback()
#         return JsonResponse({'success': False, 'message': str(e)}, status=500)

#     finally:
#         cur.close()
#         db.close()

@csrf_exempt
def manage_office_timings(request):
    user_name = request.session.get('user_name')
    cur, db = connection()

    try:
        if request.method == "POST":
            data = json.loads(request.body)
            non_working_days = data.get('non_working_days')
            work_start_time = data.get('work_start_time')
            work_end_time = data.get('work_end_time')
            location = data.get('location')  # New field
            department = data.get('department')  # New field

            # Validate required fields
            if not (non_working_days and work_start_time and work_end_time and location and department):
                return JsonResponse({'success': False, 'message': 'All fields are required'}, status=400)

            # Add new office timing
            query = """
                INSERT INTO office_timings (non_working_days, work_start_time, work_end_time, location, department)
                VALUES (%s, %s, %s, %s, %s)
            """
            cur.execute(query, (non_working_days, work_start_time, work_end_time, location, department))
            db.commit()
            cur.execute("SELECT LAST_INSERT_ID()")
            office_timing_id = cur.fetchone()[0]
            return JsonResponse({
                'success': True,
                'message': 'Office timing added successfully',
                'office_timing': {
                    'id': office_timing_id,
                    'non_working_days': non_working_days,
                    'work_start_time': work_start_time,
                    'work_end_time': work_end_time,
                    'location': location,  # Include location
                    'department': department  # Include department
                }
            })

        elif request.method == "GET":
            # Get all office timings
            cur.execute("SELECT id, non_working_days, work_start_time, work_end_time, location, department FROM office_timings")
            office_timings = cur.fetchall()
            office_timings_data = [
                {
                    'id': office_timing[0],
                    'non_working_days': office_timing[1],
                    'work_start_time': str(office_timing[2]),
                    'work_end_time': str(office_timing[3]),
                    'location': office_timing[4],  # Include location
                    'department': office_timing[5]  # Include department
                }
                for office_timing in office_timings
            ]
            
            # Get distinct locations and departments
            cur.execute('SELECT DISTINCT location FROM user')
            locations = [row[0] for row in cur.fetchall()]  # Fetch all distinct locations
            
            cur.execute('SELECT DISTINCT department FROM user')
            departments = [row[0] for row in cur.fetchall()]  # Fetch all distinct departments
            
            cur.execute('SELECT DISTINCT non_working_days FROM office_timings')
            timings = [row[0] for row in cur.fetchall()]  # Fetch all distinct departments

            return render(request, 'officetiming1.html', {
                'user_name': user_name,
                'office_timings_data': office_timings_data,
                'locations': locations,  # Add locations to context
                'departments': departments,  # Add departments to context
                'timings': timings
            })

        else:
            return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

    finally:
        cur.close()
        db.close()
        
@csrf_exempt
def edit_office_timing(request, office_timing_id):
    user_name = request.session.get('user_name')
    cur, db = connection()

    try:
        if request.method == "POST":
            data = json.loads(request.body)
            non_working_days = data.get('non_working_days')
            work_start_time = data.get('work_start_time')
            work_end_time = data.get('work_end_time')
            location = data.get('location')  # New field
            department = data.get('department')  # New field

            # Validate required fields
            if not (non_working_days and work_start_time and work_end_time and location and department):
                return JsonResponse({'success': False, 'message': 'All fields are required'}, status=400)

            # Update office timing
            query = """
                UPDATE office_timings
                SET non_working_days = %s, work_start_time = %s, work_end_time = %s, location = %s, department = %s
                WHERE id = %s
            """
            cur.execute(query, (non_working_days, work_start_time, work_end_time, location, department, office_timing_id))
            db.commit()
            return JsonResponse({'success': True, 'message': 'Office timing updated successfully'})

        else:
            return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

    except Exception as e:
        db.rollback()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

    finally:
        cur.close()
        db.close()

@csrf_exempt
def delete_office_timing(request, office_timing_id):
    cur, db = connection()

    try:
        if request.method == "DELETE":
            # Delete office timing
            cur.execute("DELETE FROM office_timings WHERE id = %s", [office_timing_id])
            db.commit()
            return JsonResponse({'success': True, 'message': 'Office timing deleted successfully'})

        else:
            return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

    except Exception as e:
        db.rollback()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

    finally:
        cur.close()
        db.close()

#Add Holiday
@csrf_exempt
def publicholidays(request):
    user_name = request.session.get('user_name')
    cur, db = connection()

    try:
        if request.method == "POST":
            # Check if the request contains a file (Excel)
            if 'excel_file' in request.FILES:
                excel_file = request.FILES['excel_file']
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                added_holidays = []

                # Read the Excel file and extract holidays
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming data starts from row 2
                    holiday_name = row[0]  # Assuming holiday name is in the first column
                    holiday_date = row[1]  # Assuming holiday date is in the second column
                    if holiday_name and holiday_date:
                        # Check if the holiday already exists in the database
                        cur.execute(
                            "SELECT 1 FROM public_holidays WHERE holiday_name = %s AND holiday_date = %s",
                            (holiday_name, holiday_date)
                        )
                        existing_holiday = cur.fetchone()

                        if not existing_holiday:
                            # Insert holiday record into the database if it doesn't exist
                            cur.execute(
                                "INSERT INTO public_holidays (holiday_name, holiday_date) VALUES (%s, %s)",
                                (holiday_name, holiday_date)
                            )
                            db.commit()

                            # Capture the inserted holiday to return to the front-end
                            added_holidays.append({'holiday_name': holiday_name, 'holiday_date': str(holiday_date)})

                return JsonResponse({
                    'success': True,
                    'message': 'File uploaded and holidays added successfully',
                    'holidays': added_holidays  # Return added holidays
                })

            # Regular holiday addition from the form data (not an Excel file)
            elif request.headers.get('Content-Type') == 'application/json':
                data = json.loads(request.body)
                holiday_name = data.get('holiday_name')
                holiday_date = data.get('holiday_date')

                if holiday_name and holiday_date:
                    # Check if the holiday already exists in the database
                    cur.execute(
                        "SELECT 1 FROM public_holidays WHERE holiday_name = %s AND holiday_date = %s",
                        (holiday_name, holiday_date)
                    )
                    existing_holiday = cur.fetchone()

                    if existing_holiday:
                        return JsonResponse({'success': False, 'message': 'Holiday already exists'}, status=400)

                    # Insert holiday record into the database if it doesn't exist
                    cur.execute(
                        "INSERT INTO public_holidays (holiday_name, holiday_date) VALUES (%s, %s)",
                        (holiday_name, holiday_date)
                    )
                    db.commit()

                    cur.execute("SELECT LAST_INSERT_ID()")
                    holiday_id = cur.fetchone()[0]

                    return JsonResponse({
                        'success': True,
                        'holiday': {
                            'id': holiday_id,
                            'holiday_name': holiday_name,
                            'holiday_date': holiday_date
                        }
                    })
                return JsonResponse({'success': False, 'message': 'Holiday name and date are required'}, status=400)

        # Handle GET request to display holidays
        cur.execute("SELECT id, holiday_name, holiday_date FROM public_holidays")
        holidays = cur.fetchall()
        holiday_data = [
            {
                'id': holiday[0],
                'holiday_name': holiday[1],
                'holiday_date': holiday[2]
            }
            for holiday in holidays
        ]

        return render(request, 'publicholidays.html', {'holiday_data': holiday_data})

    except Exception as e:
        db.rollback()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
def edit_public_holiday(request, holiday_id):
    if request.method == 'POST':
        data = json.loads(request.body)
        holiday_name = data.get('holiday_name')
        holiday_date = data.get('holiday_date')

        if not (holiday_name and holiday_date):
            return JsonResponse({'success': False, 'message': 'Holiday name and date are required'}, status=400)

        cur, db = connection()
        try:
            query = """
            UPDATE public_holidays
            SET holiday_name = %s, holiday_date = %s
            WHERE id = %s
            """
            cur.execute(query, (holiday_name, holiday_date, holiday_id))
            db.commit()

            return JsonResponse({'success': True, 'holiday': {'id': holiday_id, 'holiday_name': holiday_name, 'holiday_date': holiday_date}})
        except Exception as e:
            db.rollback()
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
        finally:
            cur.close()
            db.close()

# Delete Holiday
@csrf_exempt
def delete_public_holiday(request, holiday_id):
    cur, db = connection()

    try:
        cur.execute("DELETE FROM public_holidays WHERE id = %s", [holiday_id])
        db.commit()

        return JsonResponse({'success': True})
    except Exception as e:
        db.rollback()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    finally:
        cur.close()
        db.close()

@csrf_exempt
def escalationmatrix(request):
    user_name = request.session.get('user_name')
    cur, db = connection()

    try:
        if request.method == "POST":
            if request.headers.get('Content-Type') == 'application/json':
                data = json.loads(request.body)
                location = data.get('location')
                department = data.get('department')
                levels = data.get('levels')

                if location and department and levels:
                    try:
                        # Insert data into the EscalationMatrix table
                        cur.execute(
                            "INSERT INTO EscalationMatrix1 (location, department, level) VALUES (%s, %s, %s)",
                            (location, department, levels)
                        )
                        db.commit()

                        # Get the last inserted ID
                        cur.execute("SELECT LAST_INSERT_ID()")
                        escalation_id = cur.fetchone()[0]

                        return JsonResponse({
                            'success': True,
                            'escalation_matrix': {
                                'id': escalation_id,
                                'location': location,
                                'department': department,
                                'levels': levels
                            }
                        })
                    except Exception as e:
                        db.rollback()
                        return JsonResponse({'success': False, 'message': str(e)}, status=500)

                return JsonResponse({'success': False, 'message': 'Location, department, and levels are required'}, status=400)

        # Handle GET request to display escalation matrix
        cur.execute("SELECT id, location, department, level FROM EscalationMatrix1")
        escalation_entries = cur.fetchall()
        escalation_data = [
            {'id': entry[0], 'location': entry[1], 'department': entry[2], 'levels': entry[3]}
            for entry in escalation_entries
        ]

    except Exception as e:
        db.rollback()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

    finally:
        cur.close()
        db.close()

    return render(request, 'escalationmatrix.html', {
        'user_name': user_name,
        'escalation_data': escalation_data
    })

@csrf_exempt
def edit_escalationmatrix(request, escalation_id):
    cur, db = connection()

    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            location = data.get('location')
            department = data.get('department')
            levels = data.get('levels')

            if not (location and department and levels):
                return JsonResponse({'success': False, 'message': 'Location, department, and levels are required'}, status=400)

            try:
                # Update the escalation matrix entry
                query = """
                UPDATE EscalationMatrix1
                SET location = %s, department = %s, level = %s
                WHERE id = %s
                """
                cur.execute(query, (location, department, levels, escalation_id))
                db.commit()

                return JsonResponse({'success': True, 'escalation_matrix': {'id': escalation_id, 'location': location, 'department': department, 'levels': levels}})
            except Exception as e:
                db.rollback()
                return JsonResponse({'success': False, 'message': str(e)}, status=500)

    except Exception as e:
        db.rollback()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

    finally:
        cur.close()
        db.close()

# View to handle deleting escalation matrix entry
@csrf_exempt
def delete_escalationmatrix(request, escalation_id):
    cur, db = connection()

    try:
        # Delete the escalation matrix entry
        cur.execute("DELETE FROM EscalationMatrix1 WHERE id = %s", [escalation_id])
        db.commit()

        return JsonResponse({'success': True})

    except Exception as e:
        db.rollback()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

    finally:
        cur.close()
        db.close()

# def send_email_to_user(email, message, user_email):
#     subject = f"Escalation for Complaint ID: {message['complaint_id']}"
#     body = f"""
#     Complaint Type: {message['complaint_type']}
#     Department: {message['department']}
#     Issue Raised: {message['issue_raise_date']}
#     Escalation Level: {message['escalation_level']}
#     Message: {message['message']}
#     """
    
#     send_mail(
#         subject,
#         body,
#         user_email,  # Your email (from)
#         [email],  # To email address
#         fail_silently=False,
#     )

# def messageshoot(request):
#     cur, db = connection()
#     try:
#         # Get the current time
#         current_time = datetime.now()
#         # print(f'Current time: {current_time}')

#         # Fetch complaints along with the user's email and department details
#         cur.execute("""
#             SELECT 
#                 c.id, 
#                 c.complaint_type, 
#                 c.tat, 
#                 c.due_date, 
#                 c.status, 
#                 c.issue_raise_date, 
#                 c.user_id, 
#                 c.location, 
#                 u.levels, 
#                 u.email, 
#                 c.department AS user_department, 
#                 ctm.department AS complaint_department
#             FROM 
#                 complaint c
#             JOIN 
#                 user u ON c.user_id = u.id
#             JOIN 
#                 complaint_type_master ctm ON c.complaint_type = ctm.complaint_type
#             WHERE 
#                 DATE(c.due_date) <= DATE(%s) AND TIME(c.due_date) <= TIME(%s);
#         """, [current_time, current_time])
        
#         complaints = cur.fetchall()
        
#         # print(f'Complaints: {complaints}')  # Debug: Print fetched complaints

#         # List to hold messages to be shot
#         messages_to_shoot = []

#         for complaint in complaints:
#             (complaint_id, complaint_type, tat, due_date, status, issue_raise_date, 
#              user_id, user_location, user_level, user_email, 
#              user_department, complaint_department) = complaint

#             # Calculate the number of TAT intervals passed since the due date
#             time_diff = current_time - due_date
#             num_intervals = time_diff // timedelta(hours=tat)  # Calculate how many full TAT intervals have passed

#             # Debug: Print time_diff and num_intervals
#             # print(f'Time difference: {time_diff}')
#             # print(f'Number of intervals: {num_intervals}')

#             # If the complaint is unresolved and at least one full TAT interval has passed
#             if status in ['work in progress', 'pending'] and num_intervals > 0:
#                 # Fetch message template from the messages table
#                 cur.execute("""
#                     SELECT message_template
#                     FROM messages
#                     WHERE complaint_type = %s AND department = %s;
#                 """, [complaint_type, complaint_department])
#                 message_template = cur.fetchone()
#                 print(f'Query : {message_template}')
#                 print(f'Complaint Type : {complaint_type}')
#                 print(f'Department : {complaint_department}')                
                
#                 # Debug: Print the message template
#                 print(f'Message template: {message_template}')

#                 if message_template:
#                     # Find higher-level authorities for the user's location and department
#                     cur.execute("""
#                         SELECT u.levels, u.email
#                         FROM user u
#                         WHERE u.location = %s AND u.department = %s AND
#                               FIELD(u.levels, 'L1', 'L2', 'L3', 'L4', 'L5', 'L6') > 
#                               FIELD(%s, 'L1', 'L2', 'L3', 'L4', 'L5', 'L6')
#                         ORDER BY FIELD(u.levels, 'L1', 'L2', 'L3', 'L4', 'L5', 'L6') ASC;
#                     """, [user_location, user_department, user_level])
                    
#                     higher_levels = cur.fetchall()
#                     print(f'User Location : {user_location}')
#                     print(f'User Department : {user_department}')

#                     # Debug: Print higher-level authorities
#                     print(f'Higher levels: {higher_levels}')

#                     # Add messages for each higher-level authority
#                     for level in higher_levels:
#                         escalation_level = level[0]  # Current higher level
#                         email = level[1]  # Email of the higher-level authority

#                         # Prepare the message for the higher-level authority
#                         message = {
#                             "complaint_id": complaint_id,
#                             "message": message_template[0],
#                             "complaint_type": complaint_type,
#                             "complaint_department": complaint_department,
#                             "issue_raise_date": issue_raise_date,
#                             "user_id": user_id,
#                             "user_location": user_location,
#                             "escalation_level": escalation_level,  # Current higher level
#                             "email": email  # Email of the higher-level authority
#                         }

#                         # Add the message to the list
#                         messages_to_shoot.append(message)

#                         # Send the email to the higher-level authority
#                         send_email_to_user(email, message, user_email)

#                     # Optionally, send email to the user who raised the complaint
#                     user_message = {
#                         "complaint_id": complaint_id,
#                         "message": message_template[0],
#                         "complaint_type": complaint_type,
#                         "department": complaint_department,
#                         "issue_raise_date": issue_raise_date,
#                         "user_id": user_id,
#                         "user_location": user_location,
#                         "escalation_level": "User",  # The user level
#                         "email": user_email  # Email of the user who raised the complaint
#                     }

#                     # Send the email to the user who raised the complaint
#                     send_email_to_user(user_email, user_message, user_email)

#                     # Continue sending emails at the TAT interval until the complaint is resolved
#                     while status != 'resolved':
#                         # Calculate the next time when the email should be sent
#                         next_email_time = due_date + timedelta(hours=(num_intervals + 1) * tat)

#                         if current_time >= next_email_time:  # If it's time for the next email
#                             # Send the email to the higher-level authority
#                             send_email_to_user(email, message, user_email)
#                             # Update the complaint's due date for the next TAT interval
#                             cur.execute("""
#                                 UPDATE complaint
#                                 SET due_date = %s
#                                 WHERE id = %s
#                             """, [next_email_time, complaint_id])
#                             db.commit()
#                             num_intervals += 1  # Increase the number of intervals passed

#         # Log the messages to be sent (for debugging purposes)
#         print(f"Messages to shoot: {messages_to_shoot}")

#         # Context for rendering the template (if required)
#         context = {
#             'messages_to_shoot': messages_to_shoot
#         }

#         return render(request, 'email.html', context=context)

#     finally:
#         # Ensure the database connection is closed
#         cur.close()
#         db.close()

# Updated send_email view for handling AJAX requests
@csrf_protect
def send_email(request):
 
    user_name = request.session.get('user_name')
    if request.method == "POST":
        try:
            # Parse the incoming JSON data
            data = json.loads(request.body)
            department = data.get('department')
            complaint = data.get('complaint')
            message = data.get('message')
            signature = data.get('signature')  # Get the signature value
            print(message)
            
            # Ensure all required fields are provided
            if not department or not complaint or not message or not signature:
                return JsonResponse({'message': 'Please provide all required fields.'}, status=400)

            # Strip any HTML tags from the email message
            # message = strip_tags(message)

            # Insert the complaint into the database using a connection cursor
            cur, db = connection()  # Get the database cursor and connection
            query = """
                INSERT INTO email_master (department, complaint, message, signature)
                VALUES (%s, %s, %s, %s)
            """
            cur.execute(query, (department, complaint, message, signature))
            insert_id = cur.lastrowid
            db.commit()  # Commit the transaction

            # Close the cursor and the database connection
            cur.close()
            db.close()

            # Return a success message
            return JsonResponse({'message': 'Complaint submitted successfully!','complaint':{'department':department,'complaint':complaint,'message':message,'signature': signature,'id':insert_id} })

        except Exception as e:
            print(f"Error while submitting complaint: {e}")
            return JsonResponse({'message': 'There was an error submitting your complaint.'}, status=500)

    # Fetch department, complaint types, and complaints from the database
    cur, db = connection()
    try:
        # Fetch departments
        cur.execute("SELECT DISTINCT department FROM complaint_type_master")
        departments = cur.fetchall()
        print(f'The Departments are : {departments}')

        # Fetch complaint types
        cur.execute("SELECT * FROM complaint_type_master")
        complaint_types = cur.fetchall()
        
        # Fetch signatures
        cur.execute("SELECT DISTINCT signature FROM signature_master")
        signatures = cur.fetchall()  # Fetch all results
        # Format the results into a list of dictionaries
        signature_list = [{'id': i, 'html': sig[0]} for i, sig in enumerate(signatures, start=1)]

        # Fetch complaints
        cur.execute("SELECT * FROM email_master")
        complaints = cur.fetchall()

        # Return the data to the template
        return render(request, 'sendemail.html', {
            'departments': departments,
            'complaint_types': complaint_types,
            'complaints': complaints,
            'user_name': user_name,  # Pass user_name to the template
            'signatures': signature_list
                       
        })

    except Exception as e:
        print(f"Error fetching data: {e}")
        return JsonResponse({'message': 'There was an error fetching data from the database.'}, status=500)
    finally:
        # Ensure to close the cursor and database connection
        cur.close()
        db.close()

@csrf_protect
def edit_email(request, id):
    
    print(request)
    cur, db = connection()
    print(000)
    if request.method == 'POST':
        print(000)
               
        # Retrieve updated data from the request
        data = json.loads(request.body)
        
        print(data)
        department = data.get('department')
        complaint = data.get('complaint')
        message = data.get('message')
        signature = data.get('signature')  # Get the signature value
        
        print(message)
        if not department or not complaint or not message or not signature:
            return JsonResponse({'message': 'Please provide all required fieldsssssss.'}, status=400)
        print(222)
        try:
            data=cur.execute("""
                UPDATE email_master 
                SET department = %s, complaint = %s, message = %s, signature = %s
                WHERE id = %s
            """, (department, complaint, message, signature, id))
            
            quey1="select * from email_master where id=%s"
            test=cur.execute(quey1,id)
            print(test)
            print(333)
            print(data)  
            db.commit()
            
            return JsonResponse({'message': 'Complaint updated successfully!', 'id': id, 'department':department,'complaint':complaint,'email_message':message,'signature': signature} )

        except Exception as e:
            print(f"Error while updating complaint: {e}")
            return JsonResponse({'message': 'There was an error updating the complaint.'}, status=500)

@csrf_protect
def delete_email(request, id):
    cur, db = connection()

    if request.method == 'POST':
        try:
            # Delete the email entry by id
            cur.execute("DELETE FROM email_master WHERE id = %s", (id,))
            db.commit()

            return JsonResponse({'message': 'Complaint deleted successfully!', 'id': id})

        except Exception as e:
            print(f"Error while deleting complaint: {e}")
            return JsonResponse({'message': 'There was an error deleting the complaint.'}, status=500)

# # Create a logger instance
logger = logging.getLogger(__name__)

# def send_email_to_higher_authority(email, message):
#     """
#     Send an email to higher authorities about the complaint escalation.
#     """
#     try:
#         # Prepare the email subject and body
#         subject = f"Escalation: Complaint {message['complaint_id']} - {message['complaint_type']}"
#         body = f"""
#         Dear {message['escalation_level']},

#         A complaint has been raised regarding the following issue:
        
#         Complaint ID: {message['complaint_id']}
#         Complaint Type: {message['complaint_type']}
#         Complaint Department: {message['complaint_department']}
#         Issue Raise Date: {message['issue_raise_date']}
#         User ID: {message['user_id']}
#         User Location: {message['user_location']}
        
#         Please review the complaint and take necessary action.

#         Regards,
#         Complaint Management System
#         """

#         # Send the email
#         send_mail(
#             subject,
#             body,
#             settings.DEFAULT_FROM_EMAIL,  # Sender email from settings
#             [email],  # Recipient email (mohamed.anees@1point1.com)
#             fail_silently=False,  # Fail loudly if the email can't be sent
#         )

#         logger.info(f"Email sent to {email}")
#     except Exception as e:
#         logger.error(f"Error sending email: {e}")
        
### Modified send_email_to_higher_authority view ###############################################################################

def send_email_to_higher_authority(email, message):
    """
    Send a professional escalation email to higher authorities about a complaint.
    """
    try:
        # Prepare the email subject and body
        subject = f"Escalation: Immediate Attention Required for Complaint ID {message['complaint_id']}"

        body = f"""
        Dear {message['escalation_level']},

        I am reaching out to escalate a critical issue regarding Complaint ID {message['complaint_id']}, which has not been resolved and is now overdue. The initial complaint was submitted on {message['issue_raise_date']}, with a resolution expected by {message.get('due_date', 'N/A')}. 

        Unfortunately, despite multiple follow-ups, there has been no progress, and I am concerned about the implications of this unresolved issue:
        - Complaint Type:{message['complaint_type']}
        - Complaint Department: {message['complaint_department']}
        - User ID: {message['user_id']}
        - User Location: {message['user_location']}

        I urge you to review this situation and take the necessary steps to resolve it as soon as possible. Your prompt attention to this matter would be greatly appreciated.

        Thank you for your cooperation.

        Sincerely,

        Complaint Management System
        """
        
        # Send the email
        send_mail(
            subject,
            body,
            settings.DEFAULT_FROM_EMAIL,  # Sender email from settings
            [email],  # Recipient email
            fail_silently=False,  # Fail loudly if the email can't be sent
        )

        # Log and print success message
        logger.info(f"Escalation email successfully sent to {email}")
        print(f"Escalation email successfully sent to {email}")

    except Exception as e:
        # Log and print error message
        logger.error(f"Error sending escalation email to {email}: {e}")
        print(f"Error sending escalation email to {email}: {e}")
        
#-----------------------------------------------------------------------------------------------------------------------------#

    # except Exception as e:
    #     print(f"Error sending email: {e}")

# def messageshoot(request):

#-----------------------------------------------------------------------------------------------------------------------------#
# def messageshoot():  ## Original Message shoot function which doesnt consider shooting emails after every TAT interval
#     """
#     Fetch complaints nearing or exceeding their TAT, find higher-level authorities, and send escalation emails.
#     """
#     cur, db = connection()
#     try:
#         current_time = datetime.now()

#         # Fetch complaints nearing or exceeding their due date
#         cur.execute("""
#             SELECT 
#                 c.id, 
#                 c.complaint_type, 
#                 c.tat, 
#                 c.due_date, 
#                 c.status, 
#                 c.issue_raise_date, 
#                 c.user_id, 
#                 c.location, 
#                 u.levels, 
#                 u.email, 
#                 c.department AS user_department, 
#                 ctm.department AS complaint_department
#             FROM 
#                 complaint c
#             JOIN 
#                 user u ON c.user_id = u.id
#             JOIN 
#                 complaint_type_master ctm ON c.complaint_type = ctm.complaint_type
#             WHERE 
#                 DATE(c.due_date) <= DATE(%s) AND TIME(c.due_date) <= TIME(%s);
#         """, [current_time, current_time])

#         complaints = cur.fetchall()
#         print(f"Complaints fetched: {complaints}")

#         for complaint in complaints:
#             (complaint_id, complaint_type, tat, due_date, status, issue_raise_date,
#              user_id, user_location, user_level, user_email,
#              user_department, complaint_department) = complaint

#             print(f"Processing Complaint ID: {complaint_id}")

#             # Calculate if the complaint is overdue based on TAT
#             time_diff = current_time - due_date
#             num_intervals = time_diff // timedelta(hours=tat)

#             if status in ['work in progress', 'pending'] and num_intervals > 0:
#                 cur.execute("""
#                     SELECT message_template
#                     FROM messages
#                     WHERE complaint_type = %s AND department = %s;
#                 """, [complaint_type, complaint_department])
                
#                 message_template = cur.fetchone()
                
#                 if message_template:
#                     cur.execute("""
#                         SELECT u.levels, u.email
#                         FROM user u
#                         WHERE u.location = %s AND u.department = %s AND
#                             FIELD(u.levels, 'L1', 'L2', 'L3', 'L4', 'L5', 'L6') > 
#                             FIELD(%s, 'L1', 'L2', 'L3', 'L4', 'L5', 'L6')
#                         ORDER BY FIELD(u.levels, 'L1', 'L2', 'L3', 'L4', 'L5', 'L6') ASC;
#                     """, [user_location, user_department, user_level])

#                     higher_levels = cur.fetchall()
                    
#                     for level in higher_levels:
#                         escalation_level = level[0]
#                         email = level[1]

#                         # Prepare the message dictionary
#                         message = {
#                             "complaint_id": complaint_id,
#                             "message": message_template[0],
#                             "complaint_type": complaint_type,
#                             "complaint_department": complaint_department,
#                             "issue_raise_date": issue_raise_date,
#                             "user_id": user_id,
#                             "user_location": user_location,
#                             "escalation_level": escalation_level,
#                             "email": email
#                         }

#                         print(f"Sending email to {email} regarding Complaint ID: {complaint_id}")
                        
#                         send_email_to_higher_authority(email, message)

#         return HttpResponse("Emails sent successfully.")

#     except Exception as e:
#         logger.error(f"Error in messageshoot: {e}")
#         return HttpResponse("An error occurred while processing requests.")
    
#     finally:
#         cur.close()
#         db.close()
        
#-----------------------------------------------------------------------------------------------------------------------------#        
        
def messageshoot():   ### Modified Message shoot function to shoot email after every TAT interval 
    """
    Fetch complaints nearing or exceeding their TAT, find higher-level authorities, and send escalation emails
    based on TAT intervals.
    """
    cur, db = connection()
    try:
        current_time = datetime.now()

        # Fetch complaints nearing or exceeding their due date
        cur.execute("""
            SELECT 
                c.id, 
                c.complaint_type, 
                c.tat, 
                c.due_date, 
                c.status, 
                c.issue_raise_date, 
                c.user_id, 
                c.location, 
                u.levels, 
                u.email, 
                c.department AS user_department, 
                ctm.department AS complaint_department,
                c.last_email_sent
            FROM 
                complaint c
            JOIN 
                user u ON c.user_id = u.id
            JOIN 
                complaint_type_master ctm ON c.complaint_type = ctm.complaint_type
            WHERE 
                DATE(c.due_date) <= DATE(%s) AND TIME(c.due_date) <= TIME(%s);
        """, [current_time, current_time])

        complaints = cur.fetchall()
        print(f"Complaints fetched: {complaints}")

        for complaint in complaints:
            (complaint_id, complaint_type, tat, due_date, status, issue_raise_date,
             user_id, user_location, user_level, user_email,
             user_department, complaint_department, last_email_sent) = complaint

            print(f"Processing Complaint ID: {complaint_id}")

            # Calculate if the complaint is overdue based on TAT
            time_diff = current_time - due_date
            num_intervals = time_diff // timedelta(hours=tat)

            if status in ['work in progress', 'pending'] and num_intervals > 0:
                # Check if enough time has passed since the last email
                if last_email_sent:
                    last_email_diff = current_time - last_email_sent
                    intervals_since_last_email = last_email_diff // timedelta(hours=tat)
                else:
                    intervals_since_last_email = num_intervals

                if intervals_since_last_email > 0:
                    # Fetch the escalation message template
                    cur.execute("""
                        SELECT message
                        FROM email_master                                                          
                        WHERE complaint = %s AND department = %s;
                    """, [complaint_type, complaint_department])
                    
                    # Table name use : instead of from messages use email_master and for select use message and for complaint_type use complaint and for department use department only
                    message_template = cur.fetchone()
                    
                    if message_template:
                        cur.execute("""
                            SELECT u.levels, u.email
                            FROM user u
                            WHERE u.location = %s AND u.department = %s AND
                                FIELD(u.levels, 'L1', 'L2', 'L3', 'L4', 'L5', 'L6') > 
                                FIELD(%s, 'L1', 'L2', 'L3', 'L4', 'L5', 'L6')
                            ORDER BY FIELD(u.levels, 'L1', 'L2', 'L3', 'L4', 'L5', 'L6') ASC;
                        """, [user_location, user_department, user_level])

                        higher_levels = cur.fetchall()
                        
                        for level in higher_levels:
                            escalation_level = level[0]
                            email = level[1]

                            # Prepare the message dictionary
                            message = {
                                "complaint_id": complaint_id,
                                "message": message_template[0],
                                "complaint_type": complaint_type,
                                "complaint_department": complaint_department,
                                "issue_raise_date": issue_raise_date,
                                "user_id": user_id,
                                "user_location": user_location,
                                "escalation_level": escalation_level,
                                "email": email,
                                "due_date":due_date,
                            }

                            print(f"Sending email to {email} regarding Complaint ID: {complaint_id}")
                            
                            send_email_to_higher_authority(email, message)

                            # Update last_email_sent timestamp
                            cur.execute("""
                                UPDATE complaint
                                SET last_email_sent = %s
                                WHERE id = %s;
                            """, [current_time, complaint_id])
                            db.commit()
                            

        return HttpResponse("Emails sent successfully.")

    except Exception as e:
        logger.error(f"Error in messageshoot: {e}")
        return HttpResponse("An error occurred while processing requests.")
    
    finally:
        cur.close()
        db.close()
    
#-----------------------------------------------------------------------------------------------------------------------------#    
    
def complaint_search(request):
    # Check if the user is logged in
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')

    # Get the logged-in user's name and role from the session
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')

    # Establish database connection
    cur, db = connection()

    # Fetch distinct values for dropdown filters
    location_query = "SELECT DISTINCT location FROM complaint"
    department_query = "SELECT DISTINCT department FROM complaint"
    complaint_type_query = "SELECT DISTINCT complaint_type FROM complaint"
    status_query = "SELECT DISTINCT status FROM complaint"
    id_query = 'SELECT id FROM complaint ORDER BY id ASC'
    name_query = 'select distinct name from complaint'
    
    cur.execute(location_query)
    distinct_locations = [row[0] for row in cur.fetchall()]

    cur.execute(department_query)
    distinct_departments = [row[0] for row in cur.fetchall()]

    cur.execute(complaint_type_query)
    distinct_complaint_types = [row[0] for row in cur.fetchall()]

    cur.execute(status_query)
    statuses = [row[0] for row in cur.fetchall()]
    
    cur.execute(id_query)
    ids = [row[0] for row in cur.fetchall()]
    
    cur.execute(name_query)
    names = [row[0] for row in cur.fetchall()]

    # Retrieve filter parameters from GET request
    filter_params = {
        'fromdate': request.GET.get('fromdate', ''),
        'todate': request.GET.get('todate', ''),
        'location': request.GET.get('location', ''),
        'department': request.GET.get('department', ''),
        'complaint_type': request.GET.get('complaint_type', ''),
        'status': request.GET.get('status', ''),
        'complaint_id': request.GET.get('complaint_id', ''),
        'complaint_name': request.GET.get('complaint_name', ''),
    }

    # Prepare filter conditions
    filter_conditions = []
    query_params = []

    if filter_params['fromdate'] and filter_params['todate']:
        filter_conditions.append("c.issue_raise_date BETWEEN %s AND %s")
        query_params.extend([f"{filter_params['fromdate']} 00:00:00", f"{filter_params['todate']} 23:59:59"])
    elif filter_params['fromdate']:
        filter_conditions.append("c.issue_raise_date >= %s")
        query_params.append(f"{filter_params['fromdate']} 00:00:00")
    elif filter_params['todate']:
        filter_conditions.append("c.issue_raise_date <= %s")
        query_params.append(f"{filter_params['todate']} 23:59:59")

    if filter_params['location']:
        filter_conditions.append("c.location = %s")
        query_params.append(filter_params['location'])

    if filter_params['department']:
        filter_conditions.append("u.department = %s")
        query_params.append(filter_params['department'])

    if filter_params['complaint_type']:
        filter_conditions.append("c.complaint_type = %s")
        query_params.append(filter_params['complaint_type'])

    if filter_params['status']:
        filter_conditions.append("LOWER(c.status) = LOWER(%s)")
        query_params.append(filter_params['status'])

    if filter_params['complaint_id']:
        complaint_ids = filter_params['complaint_id'].split(',') # Split the complaint ID values by comma seperated values
        filter_conditions.append("c.id IN (%s)" % ','.join(['%s'] * len(complaint_ids))) # join converts to CS strings.
        query_params.extend(complaint_ids)

    if filter_params['complaint_name']:
        filter_conditions.append("LOWER(c.name) LIKE %s")
        query_params.append(f"%{filter_params['complaint_name'].lower()}%")

    # Build the WHERE clause if there are any filter conditions
    where_clause = f"WHERE {' AND '.join(filter_conditions)}" if filter_conditions else ""

    # Fetch resolution reports
    resolution_reports = []
    if filter_conditions:
        query = f"""
        SELECT 
            c.id AS complaint_id,
            u.name AS complainant_name, 
            u.department, 
            c.location, 
            COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
            c.issue_raise_date AS date,
            c.due_date,
            c.status,
            c.resolution,
            c.description,
            c.complaint_type,
            CASE 
                WHEN c.status = 'resolved' THEN c.updated_at
                ELSE NULL
            END AS resolved_at,
            c.updated_due_date,
            c.reason
        FROM 
            complaint AS c
        JOIN 
            user AS u ON c.user_id = u.id
        LEFT JOIN 
            user AS u2 ON c.resolved_by_id = u2.id AND u2.role = 'admin'
        {where_clause}
        ORDER BY 
            c.issue_raise_date DESC;
        """
        cur.execute(query, query_params)
        resolution_reports = cur.fetchall()

    # Process the fetched data to format and calculate TAT
    report_data = []
    for report in resolution_reports:
        issue_raise_date = report[5]
        resolved_at = report[11]
        tat_duration = resolved_at - issue_raise_date if resolved_at else datetime.now() - issue_raise_date
        total_seconds = tat_duration.total_seconds()
        days = int(total_seconds // 86400)
        hours = int((total_seconds % 86400) // 3600)
        minutes = int((total_seconds % 3600) // 60)
        seconds = int(total_seconds % 60)
        tat_formatted = (
            f"{days}d {hours}h {minutes}m {seconds}s" if days > 0 else
            f"{hours}h {minutes}m {seconds}s" if hours > 0 else
            f"{minutes}m {seconds}s"
        )

        report_data.append({
            'id': report[0],
            'name': report[1],
            'department': report[2],
            'location': report[3].capitalize(),
            'resolved_by': report[4],
            'date': report[5].strftime("%Y-%m-%d %H:%M:%S"),
            'due_date': report[6].strftime("%Y-%m-%d %H:%M:%S"),
            'status': report[7].capitalize(),
            'resolution': report[8],
            'description': report[9],
            'complaint_type': report[10],
            'updated_at': report[11].strftime("%Y-%m-%d %H:%M:%S") if report[11] else 'None',
            'tat': tat_formatted if resolved_at else 'None',
            'updated_due_date': report[12].strftime("%Y-%m-%d %H:%M:%S") if report[12] else 'No revised due date assigned yet.',
            'reason': report[13] if report[13] else 'No reason specified.'
        })

    # Pass the processed data and filter options to the template
    return render(request, 'complaintsearch.html', {
        'complaints': report_data,
        'distinct_locations': distinct_locations,
        'distinct_departments': distinct_departments,
        'distinct_complaint_types': distinct_complaint_types,
        'statuses': statuses,
        'ids': ids,
        'names': names,
        'user_name': user_name,
        'user_role': user_role,
        'filters': filter_params,  # Pass the filters dictionary
    })

def user_complaint_search(request):
    # Check if the user is logged in
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')

    # Get the logged-in user's name and role from the session
    user_name = request.session.get('user_name')
    user_role = request.session.get('user_role')

    # Establish database connection
    cur, db = connection()

    # Fetch distinct values for dropdown filters
    location_query = "SELECT DISTINCT location FROM complaint where user_id =%s"
    department_query = "SELECT DISTINCT department FROM complaint where user_id = %s"
    complaint_type_query = "SELECT DISTINCT complaint_type FROM complaint where user_id = %s"
    status_query = "SELECT DISTINCT status FROM complaint where user_id = %s"
    ids_query = "SELECT id from complaint where user_id = %s"

    cur.execute(location_query, [user_id])
    distinct_locations = [row[0] for row in cur.fetchall()]

    cur.execute(department_query, [user_id])
    distinct_departments = [row[0] for row in cur.fetchall()]

    cur.execute(complaint_type_query, [user_id])
    distinct_complaint_types = [row[0] for row in cur.fetchall()]

    cur.execute(status_query, [user_id])
    statuses = [row[0] for row in cur.fetchall()]
    
    cur.execute(ids_query, [user_id])
    ids = [row[0] for row in cur.fetchall()]

    # Initialize report_data as an empty list for the default state
    report_data = []

    # Retrieve filter parameters from GET request
    filter_params = {
        'fromdate': request.GET.get('fromdate', ''),
        'todate': request.GET.get('todate', ''),
        'location': request.GET.get('location', ''),
        'department': request.GET.get('department', ''),
        'complaint_type': request.GET.get('complaint_type', ''),
        'status': request.GET.get('status', ''),
        'complaint_id': request.GET.get('complaint_id', '')
    }

    # Prepare filter conditions for user's complaints
    filter_conditions = []
    query_params = [user_id]  # Add the user_id as the first parameter

    if filter_params['fromdate'] and filter_params['todate']:
        filter_conditions.append("c.issue_raise_date BETWEEN %s AND %s")
        query_params.extend([f"{filter_params['fromdate']} 00:00:00", f"{filter_params['todate']} 23:59:59"])
    elif filter_params['fromdate']:
        filter_conditions.append("c.issue_raise_date >= %s")
        query_params.append(f"{filter_params['fromdate']} 00:00:00")
    elif filter_params['todate']:
        filter_conditions.append("c.issue_raise_date <= %s")
        query_params.append(f"{filter_params['todate']} 23:59:59")

    if filter_params['location']:
        filter_conditions.append("c.location = %s")
        query_params.append(filter_params['location'])

    if filter_params['department']:
        filter_conditions.append("u.department = %s")
        query_params.append(filter_params['department'])

    if filter_params['complaint_type']:
        filter_conditions.append("c.complaint_type = %s")
        query_params.append(filter_params['complaint_type'])

    if filter_params['status']:
        filter_conditions.append("LOWER(c.status) = LOWER(%s)")
        query_params.append(filter_params['status'])
        
    if filter_params['complaint_id']:
        # Split the complaint_id string into a list and filter by multiple IDs
        complaint_ids = filter_params['complaint_id'].split(',')
        filter_conditions.append("c.id IN (%s)" % ','.join(['%s'] * len(complaint_ids)))
        query_params.extend(complaint_ids)

    # Build the WHERE clause if there are any filter conditions
    where_clause = f"WHERE c.user_id = %s {'AND ' if filter_conditions else ''}{' AND '.join(filter_conditions)}" if filter_conditions else "WHERE user_id = %s"

    # Final query with filters applied for user-specific complaints
    query = f"""
        SELECT 
            c.id AS complaint_id,  -- Add the complaint ID here
            u.name AS complainant_name, 
            u.department, 
            c.location, 
            COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
            c.issue_raise_date AS date,
            c.due_date,
            c.status,
            c.resolution,
            c.description,
            c.complaint_type,
            CASE 
                WHEN c.status = 'resolved' THEN c.updated_at
                ELSE NULL
            END AS resolved_at,
            c.updated_due_date,
            c.reason
        FROM 
            complaint AS c
        JOIN 
            user AS u ON c.user_id = u.id
        LEFT JOIN 
            user AS u2 ON c.resolved_by_id = u2.id AND u2.role = 'admin'
        {where_clause}
        ORDER BY 
            c.issue_raise_date DESC;
    """

    # Only execute the query if filters are applied
    if filter_conditions:
        cur.execute(query, query_params)
        resolution_reports = cur.fetchall()

        # Process the fetched data to format and calculate TAT
        for report in resolution_reports:
            complaint_id = report[0]  # Extract the complaint ID
            complainant_name = report[1]
            department = report[2]
            location = report[3].capitalize()
            resolved_by = report[4]
            date = report[5]
            due_date = report[6]
            status = report[7].capitalize()
            resolution = report[8]
            description = report[9]
            complaint_type = report[10]
            resolved_at = report[11]
            updated_due_date = report[12]
            reason = report[13]

            # Calculate TAT
            issue_raise_date = date
            if resolved_at:
                tat_duration = resolved_at - issue_raise_date
            else:
                tat_duration = datetime.now() - issue_raise_date

            total_seconds = tat_duration.total_seconds()
            days = int(total_seconds // 86400)
            hours = int((total_seconds % 86400) // 3600)
            minutes = int((total_seconds % 3600) // 60)
            seconds = int(total_seconds % 60)

            # Format TAT based on conditions
            if days > 0:
                tat_formatted = f"{days}d {hours}h {minutes}m {seconds}s"
            elif hours > 0:
                tat_formatted = f"{hours}h {minutes}m {seconds}s"
            else:
                tat_formatted = f"{minutes}m {seconds}s"

            # Append the processed data
            report_data.append({
                'id': complaint_id,
                'name': complainant_name,
                'department': department,
                'location': location,
                'resolved_by': resolved_by,
                'date': date.strftime("%Y-%m-%d %H:%M:%S"),
                'due_date': due_date.strftime("%Y-%m-%d %H:%M:%S"),
                'status': status,
                'resolution': resolution,
                'description': description,
                'complaint_type': complaint_type,
                'updated_at': resolved_at.strftime("%Y-%m-%d %H:%M:%S") if resolved_at else 'None',
                'tat': tat_formatted if resolved_at else 'None',
                'updated_due_date': updated_due_date.strftime("%Y-%m-%d %H:%M:%S") if updated_due_date else 'No revised due date assigned yet.',
                'reason': reason if reason else 'No reason specified.'
            })

    # Fetch the logged-in user's details
    cur.execute("SELECT * FROM user WHERE id = %s", [user_id])
    user = cur.fetchone()

    # Prepare context to pass to the template
    context = {
        'complaints': report_data,
        'distinct_locations': distinct_locations,
        'distinct_departments': distinct_departments,
        'distinct_complaint_types': distinct_complaint_types,
        'statuses': statuses,
        'user_name': user_name,
        'user_role': user_role,
        'filter_params': filter_params,  # Pass filter parameters to template
        'user_details': user,  # Pass user details to the template
        'ids': ids,
    }

    # Pass the context to the template
    return render(request, 'usercomplaintsearch.html', context)
    
def resolver_complaint_search(request):
    # Check if the user is logged in
    resolver_id = request.session.get('user_id')
    if not resolver_id:
        return redirect('login')

    # Get the logged-in resolver's name and role from the session
    resolver_name = request.session.get('user_name')
    resolver_role = request.session.get('user_role')

    # Establish database connection
    cur, db = connection()

    # Fetch distinct values for dropdown filters
    location_query = "SELECT DISTINCT location FROM complaint where assigned_to = %s"
    department_query = "SELECT DISTINCT department FROM complaint where assigned_to = %s"
    complaint_type_query = "SELECT DISTINCT complaint_type FROM complaint where assigned_to = %s"
    status_query = "SELECT DISTINCT status FROM complaint where assigned_to=%s"
    ids_query = 'SELECT id from complaint where assigned_to = %s'
    name_query = 'select distinct name from complaint where assigned_to = %s'

    cur.execute(location_query, [resolver_id])
    distinct_locations = [row[0] for row in cur.fetchall()]

    cur.execute(department_query, [resolver_id])
    distinct_departments = [row[0] for row in cur.fetchall()]

    cur.execute(complaint_type_query, [resolver_id])
    distinct_complaint_types = [row[0] for row in cur.fetchall()]

    cur.execute(status_query, [resolver_id])
    statuses = [row[0] for row in cur.fetchall()]

    cur.execute(ids_query, [resolver_id])
    ids = [row[0] for row in cur.fetchall()]

    cur.execute(name_query, [resolver_id])
    names = [row[0] for row in cur.fetchall()]

    # Initialize report_data as an empty list for the default state
    report_data = []

    # Retrieve filter parameters from GET request
    filters = {
        'fromdate': request.GET.get('fromdate', ''),
        'todate': request.GET.get('todate', ''),
        'location': request.GET.get('location', ''),
        'department': request.GET.get('department', ''),
        'complaint_type': request.GET.get('complaint_type', ''),
        'status': request.GET.get('status', ''),
        'complaint_id': request.GET.get('complaint_id', ''),
        'complaint_name': request.GET.get('complaint_name', '')
    }

    # Prepare filter conditions for resolver's complaints
    filter_conditions = []
    query_params = [resolver_id]  # Add the resolver_id as the first parameter

    if filters['fromdate'] and filters['todate']:
        filter_conditions.append("c.issue_raise_date BETWEEN %s AND %s")
        query_params.extend([f"{filters['fromdate']} 00:00:00", f"{filters['todate']} 23:59:59"])
    elif filters['fromdate']:
        filter_conditions.append("c.issue_raise_date >= %s")
        query_params.append(f"{filters['fromdate']} 00:00:00")
    elif filters['todate']:
        filter_conditions.append("c.issue_raise_date <= %s")
        query_params.append(f"{filters['todate']} 23:59:59")

    if filters['location']:
        filter_conditions.append("c.location = %s")
        query_params.append(filters['location'])

    if filters['department']:
        filter_conditions.append("u.department = %s")
        query_params.append(filters['department'])

    if filters['complaint_type']:
        filter_conditions.append("c.complaint_type = %s")
        query_params.append(filters['complaint_type'])

    if filters['status']:
        filter_conditions.append("LOWER(c.status) = LOWER(%s)")
        query_params.append(filters['status'])

    if filters['complaint_id']:
        # Split the complaint_id string into a list and filter by multiple IDs
        complaint_ids = filters['complaint_id'].split(',')
        filter_conditions.append("c.id IN (%s)" % ','.join(['%s'] * len(complaint_ids)))
        query_params.extend(complaint_ids)

    if filters['complaint_name']:
        filter_conditions.append("LOWER(c.name) LIKE %s")
        query_params.append(f"%{filters['complaint_name'].lower()}%")

    # Build the WHERE clause if there are any filter conditions
    where_clause = f"WHERE assigned_to = %s {'AND ' if filter_conditions else ''}{' AND '.join(filter_conditions)}" if filter_conditions else "WHERE assigned_to = %s"

    # Final query with filters applied for resolver's complaints
    query = f"""
        SELECT
            c.id AS complaint_id,  -- Add the complaint ID here 
            u.name AS complainant_name, 
            u.department, 
            c.location, 
            COALESCE(u2.name, 'Not Resolved') AS resolved_by_name,  
            c.issue_raise_date AS date,
            c.due_date,
            c.status,
            c.resolution,
            c.description,
            c.complaint_type,
            CASE 
                WHEN c.status = 'resolved' THEN c.updated_at
                ELSE NULL
            END AS resolved_at,
            c.updated_due_date,
            c.reason
        FROM 
            complaint AS c
        JOIN 
            user AS u ON c.user_id = u.id
        LEFT JOIN 
            user AS u2 ON c.resolved_by_id = u2.id AND u2.role = 'admin'
        {where_clause}
        ORDER BY 
            c.issue_raise_date DESC;
    """

    if filter_conditions:
        cur.execute(query, query_params)
        resolution_reports = cur.fetchall()

        # Process the fetched data to format and calculate TAT
        for report in resolution_reports:
            complaint_id = report[0]
            complainant_name = report[1]
            department = report[2]
            location = report[3].capitalize()
            resolved_by = report[4]
            date = report[5]
            due_date = report[6]
            status = report[7].capitalize()
            resolution = report[8]
            description = report[9]
            complaint_type = report[10]
            resolved_at = report[11]
            updated_due_date = report[12]
            reason = report[13]

            # Calculate TAT
            issue_raise_date = date
            if resolved_at:
                tat_duration = resolved_at - issue_raise_date
            else:
                tat_duration = datetime.now() - issue_raise_date

            total_seconds = tat_duration.total_seconds()
            days = int(total_seconds // 86400)
            hours = int((total_seconds % 86400) // 3600)
            minutes = int((total_seconds % 3600) // 60)
            seconds = int(total_seconds % 60)

            # Format TAT based on conditions
            if days > 0:
                tat_formatted = f"{days}d {hours}h {minutes}m {seconds}s"
            elif hours > 0:
                tat_formatted = f"{hours}h {minutes}m {seconds}s"
            else:
                tat_formatted = f"{minutes}m {seconds}s"

            # Append the processed data
            report_data.append({
                'id': complaint_id,
                'name': complainant_name,
                'department': department,
                'location': location,
                'resolved_by': resolved_by,
                'date': date.strftime("%Y-%m-%d %H:%M:%S"),
                'due_date': due_date.strftime("%Y-%m-%d %H:%M:%S"),
                'status': status,
                'resolution': resolution,
                'description': description,
                'complaint_type': complaint_type,
                'updated_at': resolved_at.strftime("%Y-%m-%d %H:%M:%S") if resolved_at else 'None',
                'tat': tat_formatted if resolved_at else 'None',
                'updated_due_date': updated_due_date.strftime("%Y-%m-%d %H:%M:%S") if updated_due_date else 'No revised due date assigned yet.',
                'reason': reason if reason else 'No reason specified.'
            })

    # Fetch the logged-in resolver's details
    cur.execute("SELECT * FROM user WHERE id = %s", [resolver_id])
    resolver = cur.fetchone()

    # Pass the processed data and filter options to the template
    return render(request, 'resolvercomplaintsearch.html', {
        'complaints': report_data,
        'distinct_locations': distinct_locations,
        'distinct_departments': distinct_departments,
        'distinct_complaint_types': distinct_complaint_types,
        'statuses': statuses,
        'user_name': resolver_name,
        'user_role': resolver_role,
        'from_date': filters['fromdate'],
        'to_date': filters['todate'],
        'user_details': resolver,
        'ids': ids,
        'names': names,
        'filters': filters  # Pass the filters dictionary to the template
    })

# def signature_master(request):
#     user_name = request.session.get('username')
#     user_role = request.session.get('role')
#     cur, db = connection()  # Ensure your connection function returns both cursor and connection objects

#     if request.method == "POST":
#         # Fetch form data
#         name = request.POST.get('name')
#         designation = request.POST.get('designation')
#         contact = request.POST.get('contact')
#         image_file = request.FILES.get('image')  # Use FILES for file uploads

#         if name and designation and contact and image_file:
#             try:
#                 # Read image as binary data
#                 image_data = image_file.read()

#                 # Insert data into the table
#                 query = """
#                 INSERT INTO signature_master (name, designation, contact, image)
#                 VALUES (%s, %s, %s, %s)
#                 """
#                 cur.execute(query, (name, designation, contact, image_data))
#                 db.commit()

#                 messages.success(request, "Record added successfully!")
#             except Exception as e:
#                 db.rollback()
#                 messages.error(request, f"An error occurred: {e}")
#         else:
#             messages.error(request, "All fields are required.")

#     # Fetch all records for display, including the image
#     try:
#         cur.execute("SELECT id, name, designation, contact, image FROM signature_master")
#         records = cur.fetchall()
#     except Exception as e:
#         messages.error(request, f"An error occurred while fetching records: {e}")
#         records = []

#     # Convert binary image data to Base64
#     processed_records = []
#     for record in records:
#         image_data = record[4]
#         if image_data:
#             base64_image = f"data:image/jpeg;base64,{base64.b64encode(image_data).decode('utf-8')}"
#         else:
#             base64_image = None
#         processed_records.append({
#             'id': record[0],
#             'name': record[1],
#             'designation': record[2],
#             'contact': record[3],
#             'image': base64_image,
#         })

#     return render(request, 'signature_master.html', {
#         'user_name': user_name,
#         'user_role': user_role,
#         'records': processed_records
#     })

@csrf_protect
def signature_master(request):
 
    user_name = request.session.get('user_name')
    if request.method == "POST":
        try:
            # Parse the incoming JSON data
            data = json.loads(request.body)
            name = data.get('name')
            signature = data.get('signature')
            print(signature)
            
            
            # Ensure all required fields are provided
            if not name or not signature:
                return JsonResponse({'message': 'Please provide all required fields.'}, status=400)


            # Strip any HTML tags from the email message
            # message = strip_tags(message)


            # Insert the complaint into the database using a connection cursor
            cur, db = connection()  # Get the database cursor and connection
            query = """
                INSERT INTO signature_master (name, signature)
                VALUES (%s, %s)
            """
            cur.execute(query, (name, signature))
            insert_id = cur.lastrowid
            db.commit()  # Commit the transaction


            # Close the cursor and the database connection
            cur.close()
            db.close()


            # Return a success message
            return JsonResponse({'message': 'Complaint submitted successfully!','complaint':{'name':name,'signature':signature,'id':insert_id} })


        except Exception as e:
            print(f"Error while submitting complaint: {e}")
            return JsonResponse({'message': 'There was an error submitting your complaint.'}, status=500)


    # Fetch department, complaint types, and complaints from the database
    cur, db = connection()
    try:
        # Fetch Names
        cur.execute("SELECT name FROM signature_master")
        names = cur.fetchall()
        print(f'The Names are : {names}')
        

        # Fetch Signatures
        cur.execute("SELECT signature FROM signature_master")
        signatures = cur.fetchall()
        print(f'The Signatures are : {signatures}')


        # Fetch Signature Master Table Information
        cur.execute("SELECT * FROM signature_master")
        complaints = cur.fetchall()


        # Return the data to the template
        return render(request, 'signature_master.html', {
            'names': names,
            'signatures': signatures,
            'complaints': complaints,
            'user_name': user_name  # Pass user_name to the template
            
        })


    except Exception as e:
        print(f"Error fetching data: {e}")
        return JsonResponse({'message': 'There was an error fetching data from the database.'}, status=500)
    finally:
        # Ensure to close the cursor and database connection
        cur.close()
        db.close()


@csrf_protect
def edit_signature(request, id):
    cur, db = connection()

    if request.method == 'POST':
        data = json.loads(request.body)
        name = data.get('name')
        signature = data.get('signature')

        if not name or not signature:
            return JsonResponse({'message': 'Please provide all required fields.'}, status=400)

        try:
            cur.execute("""
                UPDATE signature_master 
                SET name = %s, signature = %s
                WHERE id = %s
            """, (name, signature, id))

            db.commit()

            return JsonResponse({'message': 'Complaint updated successfully!', 'id': id, 'name': name, 'signature': signature})
        except Exception as e:
            print(f"Error while updating complaint: {e}")
            return JsonResponse({'message': 'There was an error updating the complaint.'}, status=500)
        finally:
            cur.close()
            db.close()

@csrf_protect
def delete_signature(request, id):
    cur, db = connection()


    if request.method == 'POST':
        try:
            # Delete the email entry by id
            cur.execute("DELETE FROM signature_master WHERE id = %s", (id,))
            db.commit()


            return JsonResponse({'message': 'Complaint deleted successfully!', 'id': id})


        except Exception as e:
            print(f"Error while deleting complaint: {e}")
            return JsonResponse({'message': 'There was an error deleting the complaint.'}, status=500)

    
@csrf_exempt
def upload_image(request):
    if request.method == "POST" and request.FILES.get("upload"):
        uploaded_file = request.FILES["upload"]
        file_path = os.path.join(settings.MEDIA_ROOT, uploaded_file.name)

        with open(file_path, "wb+") as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

        # Construct the file URL
        file_url = os.path.join(settings.MEDIA_URL, uploaded_file.name)

        return JsonResponse({
            "uploaded": True,
            "url": file_url
        })

    return JsonResponse({
        "uploaded": False,
        "error": {"message": "No file uploaded"}
    })
    
def your_view(request):
    # Establish a connection and execute the query
    cur, db = connection()
    cur.execute("SELECT DISTINCT signature FROM signature_master")
    signatures = cur.fetchall()  # Fetch all results

    # Convert the result into a list of dictionaries or tuples for the template
    signature_list = [{'id': i, 'name': sig[0]} for i, sig in enumerate(signatures, start=1)]

    # Pass the data to the template
    context = {'signatures': signature_list}
    return render(request, 'sendemail.html', context)
